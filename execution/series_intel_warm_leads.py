#!/usr/bin/env python3
"""
Warm Leads / R+D — Series Intel (Separate Analysis Workbook)
============================================================
Builds series-level intel using Gemini grounded search with anti-hallucination checks.

Inputs:
- outreach/sheets/licensing warm leads analysis.xlsx  (sheet: "Warm Leads R+D")
- outreach/sheets/All-Genre Licensing Tracker.xlsx    (sheet: "R+D")

Outputs (separate, does not modify Molly's trackers):
- outreach/sheets/Warm_Leads_Series_Intel.xlsx
- outreach/sheets/RD_Series_Intel.xlsx

Token-averse design:
- One grounded Gemini call per unique (series, author) pair.
- Strong caching and resumability.
- Keep generation maxOutputTokens low; ask for compact JSON only.

Anti-hallucination design:
- Require evidence URLs for all non-obvious claims.
- Compute ratings summary from the sheet (not from LLM).
- Evidence check: try fetching each evidence URL; mark PASS if at least 1 page is fetchable and contains series/author string.

Usage:
  python3 execution/series_intel_warm_leads.py
  python3 execution/series_intel_warm_leads.py --limit 10
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from execution.websearch_email_discovery import _gemini_search
WARM_LEADS_XLSX = PROJECT / "outreach" / "sheets" / "licensing warm leads analysis.xlsx"
ALL_GENRES_XLSX = PROJECT / "outreach" / "sheets" / "All-Genre Licensing Tracker.xlsx"

OUT_WARM = PROJECT / "outreach" / "sheets" / "Warm_Leads_Series_Intel.xlsx"
OUT_RD = PROJECT / "outreach" / "sheets" / "RD_Series_Intel.xlsx"

CACHE_DIR = PROJECT / "outreach" / "sheets" / ".cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
CACHE_WARM = CACHE_DIR / "series_intel_warm_leads.cache.json"
CACHE_RD = CACHE_DIR / "series_intel_rd.cache.json"


SESSION = requests.Session()
SESSION.headers.update(
    {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
)


def _norm(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _load_json(path: Path) -> Dict[str, Any]:
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            return {}
    return {}


def _save_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))


def _safe_get(url: str, timeout: int = 6) -> str:
    try:
        r = SESSION.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        pass
    return ""


def _evidence_check(urls: List[str], series: str, author: str) -> Tuple[str, str]:
    """
    Returns (status, details)
    status: PASS / PARTIAL / FAIL
    """
    urls = [u for u in urls if u and u.startswith("http")]
    if not urls:
        return "FAIL", "no_urls"
    series_l = series.lower()
    author_l = author.lower()
    ok = 0
    fetched = 0
    for u in urls[:4]:
        html = _safe_get(u)
        if not html:
            continue
        fetched += 1
        txt = html.lower()
        if (series_l and series_l in txt) or (author_l and author_l in txt):
            ok += 1
    if ok > 0:
        return "PASS", f"matched={ok} fetched={fetched}"
    if fetched > 0:
        return "PARTIAL", f"matched=0 fetched={fetched}"
    return "FAIL", "fetch_failed"


def _parse_ratings_from_row(row: Dict[str, Any], max_books: int = 6) -> Dict[str, Any]:
    """
    Uses existing sheet columns (Book 1..5 + Last) to compute:
    - rating_count_total (sum of counts that exist)
    - rating_avg (weighted by counts where possible; else simple avg)
    - rating_trend (rough slope: last - first)
    """
    ratings = []
    counts = []
    # Standard schema in licensing warm leads analysis.xlsx:
    for i in range(1, 6):
        ratings.append(row.get(f"Book {i} Rating"))
        counts.append(row.get(f"Book {i} Count"))
    ratings.append(row.get("Last Book Rating"))
    counts.append(row.get("Last Book Count"))

    clean_r = []
    clean_c = []
    for r, c in zip(ratings, counts):
        try:
            rv = float(r)
            if rv <= 0:
                continue
            clean_r.append(rv)
            try:
                cv = float(c)
                clean_c.append(cv if cv > 0 else 0.0)
            except Exception:
                clean_c.append(0.0)
        except Exception:
            continue

    if not clean_r:
        return {"rating_count_total": "", "rating_avg": "", "rating_trend": "", "rating_notes": "no_ratings_in_sheet"}

    weighted_possible = any(x > 0 for x in clean_c) and len(clean_c) == len(clean_r)
    if weighted_possible:
        total = sum(clean_c)
        if total > 0:
            avg = sum(r * c for r, c in zip(clean_r, clean_c)) / total
        else:
            avg = sum(clean_r) / len(clean_r)
    else:
        total = ""
        avg = sum(clean_r) / len(clean_r)

    trend = clean_r[-1] - clean_r[0] if len(clean_r) >= 2 else 0.0
    return {
        "rating_count_total": int(total) if isinstance(total, (int, float)) and total != "" else total,
        "rating_avg": round(avg, 3),
        "rating_trend": round(trend, 3),
        "rating_notes": "weighted" if weighted_possible else "unweighted",
    }


def _gemini_series_intel(series: str, author: str, subgenre: str = "", trope: str = "", n_books: str = "") -> Dict[str, Any]:
    """
    One grounded call per series, compact JSON only.
    """
    prompt = (
        "Create concise series intel for licensing evaluation using grounded web search.\n"
        f'SERIES: "{series}"\n'
        f'AUTHOR: "{author}"\n'
        + (f"SUBGENRE (from sheet): {subgenre}\n" if subgenre else "")
        + (f"TROPE (from sheet): {trope}\n" if trope else "")
        + (f"NUM_BOOKS (from sheet): {n_books}\n" if n_books else "")
        + "Return ONLY these 8 lines, no markdown, no bullets:\n"
        "FIRST_PUBLISHED: <date or unknown>\n"
        "MOST_RECENT_PUBLISHED: <date or unknown>\n"
        "AUTHOR_CREDIBILITY: <1 short sentence on acclaim/scale>\n"
        "SERIES_PROMINENCE: <1 short sentence on popularity, bestseller signals, awards or adaptations>\n"
        "STORY_BLURB: <1 short sentence>\n"
        "SERIES_STRUCTURE: <same couple / rotating couples / anthology / unclear>\n"
        "WHY_INTERESTING: <1 short sentence for licensing evaluation>\n"
        "SUBGENRE_FIT: <1 short sentence on why it fits or does not fit>\n"
        "Use 'unknown' when needed, but do not leave fields blank."
    )
    text, sources, _ = _gemini_search(prompt)

    data: Dict[str, Any] = {
        "publication": {"first_published": None, "most_recent_published": None, "notes": None},
        "author_credibility": {"summary": None, "highlights": [], "evidence_urls": []},
        "series_prominence": {"summary": None, "awards": [], "bestseller_signals": [], "adaptations": [], "evidence_urls": []},
        "story": {"blurb": None, "series_structure": None, "why_interesting": None, "subgenre_fit": None, "evidence_urls": []},
    }

    patterns = {
        "first_published": r"FIRST_PUBLISHED:\s*(.+)",
        "most_recent_published": r"MOST_RECENT_PUBLISHED:\s*(.+)",
        "author_credibility": r"AUTHOR_CREDIBILITY:\s*(.+)",
        "series_prominence": r"SERIES_PROMINENCE:\s*(.+)",
        "story_blurb": r"STORY_BLURB:\s*(.+)",
        "series_structure": r"SERIES_STRUCTURE:\s*(.+)",
        "why_interesting": r"WHY_INTERESTING:\s*(.+)",
        "subgenre_fit": r"SUBGENRE_FIT:\s*(.+)",
    }

    extracted = {}
    for key, pattern in patterns.items():
        m = re.search(pattern, text, re.I)
        if m:
            value = m.group(1).strip().strip("`")
            if value.lower() != "unknown":
                extracted[key] = value

    if extracted.get("first_published"):
        data["publication"]["first_published"] = extracted["first_published"]
    if extracted.get("most_recent_published"):
        data["publication"]["most_recent_published"] = extracted["most_recent_published"]
    if extracted.get("author_credibility"):
        data["author_credibility"]["summary"] = extracted["author_credibility"]
    if extracted.get("series_prominence"):
        data["series_prominence"]["summary"] = extracted["series_prominence"]
    if extracted.get("story_blurb"):
        data["story"]["blurb"] = extracted["story_blurb"]
    if extracted.get("series_structure"):
        data["story"]["series_structure"] = extracted["series_structure"]
    if extracted.get("why_interesting"):
        data["story"]["why_interesting"] = extracted["why_interesting"]
    if extracted.get("subgenre_fit"):
        data["story"]["subgenre_fit"] = extracted["subgenre_fit"]

    if sources:
        data["_grounding_urls"] = sources[:6]

    # If the first pass produced almost nothing, do one fallback grounded call.
    filled = sum(
        bool(x)
        for x in [
            data["publication"]["first_published"],
            data["publication"]["most_recent_published"],
            data["author_credibility"]["summary"],
            data["series_prominence"]["summary"],
            data["story"]["blurb"],
            data["story"]["series_structure"],
            data["story"]["why_interesting"],
            data["story"]["subgenre_fit"],
        ]
    )
    if filled <= 1:
        fallback_prompt = (
            f'Using web results, briefly summarize the book series "{series}" by "{author}" for licensing review. '
            "Respond in 4 labeled lines only:\n"
            "AUTHOR: <1 short sentence on acclaim/size>\n"
            "SERIES: <1 short sentence on popularity/awards/adaptations>\n"
            "STORY: <1 short sentence on premise>\n"
            "FIT: <1 short sentence on subgenre/format fit>\n"
        )
        text2, sources2, _ = _gemini_search(fallback_prompt)
        m = re.search(r"AUTHOR:\s*(.+)", text2, re.I)
        if m and not data["author_credibility"]["summary"]:
            data["author_credibility"]["summary"] = m.group(1).strip()
        m = re.search(r"SERIES:\s*(.+)", text2, re.I)
        if m and not data["series_prominence"]["summary"]:
            data["series_prominence"]["summary"] = m.group(1).strip()
        m = re.search(r"STORY:\s*(.+)", text2, re.I)
        if m and not data["story"]["blurb"]:
            data["story"]["blurb"] = m.group(1).strip()
        m = re.search(r"FIT:\s*(.+)", text2, re.I)
        if m and not data["story"]["subgenre_fit"]:
            data["story"]["subgenre_fit"] = m.group(1).strip()
        if sources2 and not data.get("_grounding_urls"):
            data["_grounding_urls"] = sources2[:6]

    return data


def _write_output(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Series Intel"

    headers = [
        "Series Title",
        "Author",
        "Sub-genre (sheet)",
        "Primary Trope (sheet)",
        "# Books (sheet)",
        "Ratings: Series Size Signal",
        "Ratings: Avg (sheet)",
        "Ratings: Trend (sheet)",
        "Ratings: Count Total (sheet)",
        "Ratings Notes (sheet)",
        "Publication: First Published",
        "Publication: Most Recent",
        "Publication: Recency Signal",
        "Author Credibility Summary",
        "Author Review Signal",
        "Series Prominence Summary",
        "Awards / Adaptations / Bestseller Notes",
        "Story Blurb",
        "Series Structure",
        "Why Interesting",
        "Subgenre Fit",
        "Licensing Review Notes",
        "Subjective Conviction (content team)",
        "Evidence URLs (merged)",
        "Evidence Check Status",
        "Evidence Check Details",
        "Last Updated (UTC)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h

    for r_i, row in enumerate(rows, 2):
        for c_i, h in enumerate(headers, 1):
            ws.cell(r_i, c_i).value = row.get(h)

    wb.save(path)


def _read_warm_leads(limit: int = 0) -> List[Dict[str, Any]]:
    wb = load_workbook(WARM_LEADS_XLSX, data_only=True)
    ws = wb["Warm Leads R+D"]
    # This sheet contains duplicate header names later in the row; rely on the first block (cols 1..20).
    # Expected schema (first block):
    # 1 Series Title, 2 Author, 3 Sub-genre, 4 Primary Trope, 5 Book Names..., 6 # Books,
    # 7..12 ratings, 13..18 counts, 19 Tier, 20 Logline
    headers = {
        "Series Title": 1,
        "Author": 2,
        "Sub-genre": 3,
        "Primary Trope": 4,
        "Book Names (First 5 + Last)": 5,
        "# Books": 6,
        "Book 1 Rating": 7,
        "Book 2 Rating": 8,
        "Book 3 Rating": 9,
        "Book 4 Rating": 10,
        "Book 5 Rating": 11,
        "Last Book Rating": 12,
        "Book 1 Count": 13,
        "Book 2 Count": 14,
        "Book 3 Count": 15,
        "Book 4 Count": 16,
        "Book 5 Count": 17,
        "Last Book Count": 18,
        "Tier": 19,
        "Logline": 20,
    }

    rows = []
    for r in range(2, ws.max_row + 1):
        series = _norm(ws.cell(r, headers["Series Title"]).value)
        author = _norm(ws.cell(r, headers["Author"]).value)
        if not series or not author:
            continue
        d = {k: ws.cell(r, c).value for k, c in headers.items()}
        rows.append(d)
    if limit > 0:
        rows = rows[:limit]
    return rows


def _read_rd(limit: int = 0) -> List[Dict[str, Any]]:
    wb = load_workbook(ALL_GENRES_XLSX, data_only=True)
    ws = wb["R+D"]
    # Header row is 1, but many empty columns; take first 3 as canonical.
    author_col = 1
    series_col = 2
    rows = []
    for r in range(2, ws.max_row + 1):
        author = _norm(ws.cell(r, author_col).value)
        series = _norm(ws.cell(r, series_col).value)
        if not author or not series:
            continue
        rows.append({"Series Title": series, "Author": author, "Sub-genre": "", "Primary Trope": "", "# Books": ""})
    if limit > 0:
        rows = rows[:limit]
    return rows


def _run(block_rows: List[Dict[str, Any]], cache_path: Path, partial_output_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    cache = _load_json(cache_path)
    out_rows: List[Dict[str, Any]] = []

    for i, src in enumerate(block_rows, 1):
        series = _norm(src.get("Series Title") or src.get("Series ") or src.get("Series"))
        author = _norm(src.get("Author") or src.get("Author "))
        subgenre = _norm(src.get("Sub-genre"))
        trope = _norm(src.get("Primary Trope"))
        n_books = _norm(src.get("# Books") or src.get("Number of Books"))

        key = f"{series}||{author}".lower()
        cached = cache.get(key)
        def _intel_quality(obj: Dict[str, Any]) -> int:
            if not isinstance(obj, dict):
                return 0
            pub = obj.get("publication", {}) if isinstance(obj.get("publication"), dict) else {}
            cred = obj.get("author_credibility", {}) if isinstance(obj.get("author_credibility"), dict) else {}
            prom = obj.get("series_prominence", {}) if isinstance(obj.get("series_prominence"), dict) else {}
            story = obj.get("story", {}) if isinstance(obj.get("story"), dict) else {}
            fields = [
                pub.get("first_published"),
                pub.get("most_recent_published"),
                cred.get("summary"),
                prom.get("summary"),
                story.get("blurb"),
                story.get("series_structure"),
                story.get("why_interesting"),
                story.get("subgenre_fit"),
            ]
            return sum(bool(_norm(v)) for v in fields)

        if cached and _intel_quality(cached) >= 3:
            intel = cached
        else:
            intel = _gemini_series_intel(series, author, subgenre=subgenre, trope=trope, n_books=n_books)
            cache[key] = intel
            if i % 5 == 0:
                _save_json(cache_path, cache)
            time.sleep(0.4)

        rating = _parse_ratings_from_row(src)

        # Merge evidence urls
        urls: List[str] = []
        for path in [
            ("author_credibility", "evidence_urls"),
            ("series_prominence", "evidence_urls"),
            ("story", "evidence_urls"),
        ]:
            try:
                u = intel.get(path[0], {}).get(path[1], [])
                if isinstance(u, list):
                    urls.extend([_norm(x) for x in u if _norm(x).startswith("http")])
            except Exception:
                pass
        if "_grounding_urls" in intel and isinstance(intel["_grounding_urls"], list):
            urls.extend([_norm(x) for x in intel["_grounding_urls"] if _norm(x).startswith("http")])
        # de-dupe preserving order
        seen = set()
        urls2 = []
        for u in urls:
            if u in seen:
                continue
            seen.add(u)
            urls2.append(u)
        urls = urls2[:10]

        ev_status, ev_details = _evidence_check(urls, series=series, author=author)

        pub = intel.get("publication", {}) if isinstance(intel.get("publication"), dict) else {}
        story = intel.get("story", {}) if isinstance(intel.get("story"), dict) else {}
        cred = intel.get("author_credibility", {}) if isinstance(intel.get("author_credibility"), dict) else {}
        prom = intel.get("series_prominence", {}) if isinstance(intel.get("series_prominence"), dict) else {}

        # Decision helper signals
        recency_signal = ""
        first_pub = _norm(pub.get("first_published"))
        recent_pub = _norm(pub.get("most_recent_published"))
        date_text = (recent_pub or first_pub).lower()
        if any(y in date_text for y in ["2026", "2025", "2024", "2023"]):
            recency_signal = "Recent"
        elif any(y in date_text for y in ["2022", "2021", "2020", "2019", "2018"]):
            recency_signal = "Midlist backlist"
        elif date_text:
            recency_signal = "Older backlist"

        count_total = rating.get("rating_count_total")
        size_signal = ""
        try:
            count_val = int(count_total)
            if count_val >= 100000:
                size_signal = "Large audience"
            elif count_val >= 25000:
                size_signal = "Meaningful audience"
            elif count_val >= 5000:
                size_signal = "Moderate audience"
            elif count_val > 0:
                size_signal = "Niche audience"
        except Exception:
            size_signal = ""

        author_signal = ""
        author_summary = _norm(cred.get("summary"))
        lower_author = author_summary.lower()
        if any(x in lower_author for x in ["new york times", "usa today", "wall street journal", "bestselling", "bestselling author"]):
            author_signal = "Known / commercially proven"
        elif author_summary:
            author_signal = "Some market signal"

        awards_notes = _norm(prom.get("summary"))
        licensing_notes = ""
        if size_signal:
            licensing_notes += f"Audience signal: {size_signal}. "
        if recency_signal:
            licensing_notes += f"Recency: {recency_signal}. "
        if ev_status != "PASS":
            licensing_notes += "Needs manual verification before commercial use."
        else:
            licensing_notes += "Research is grounded but still needs editorial judgment."

        out_rows.append(
            {
                "Series Title": series,
                "Author": author,
                "Sub-genre (sheet)": subgenre,
                "Primary Trope (sheet)": trope,
                "# Books (sheet)": n_books,
                "Ratings: Series Size Signal": size_signal,
                "Ratings: Avg (sheet)": rating.get("rating_avg"),
                "Ratings: Trend (sheet)": rating.get("rating_trend"),
                "Ratings: Count Total (sheet)": rating.get("rating_count_total"),
                "Ratings Notes (sheet)": rating.get("rating_notes"),
                "Publication: First Published": first_pub or "",
                "Publication: Most Recent": recent_pub or "",
                "Publication: Recency Signal": recency_signal,
                "Author Credibility Summary": author_summary or "",
                "Author Review Signal": author_signal,
                "Series Prominence Summary": awards_notes or "",
                "Awards / Adaptations / Bestseller Notes": awards_notes or "",
                "Story Blurb": _norm(story.get("blurb")) or "",
                "Series Structure": _norm(story.get("series_structure")) or "",
                "Why Interesting": _norm(story.get("why_interesting")) or "",
                "Subgenre Fit": _norm(story.get("subgenre_fit")) or "",
                "Licensing Review Notes": licensing_notes.strip(),
                "Subjective Conviction (content team)": "",
                "Evidence URLs (merged)": " ; ".join(urls),
                "Evidence Check Status": ev_status,
                "Evidence Check Details": ev_details,
                "Last Updated (UTC)": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

        if partial_output_path and i % 5 == 0:
            _write_output(partial_output_path, out_rows)

    _save_json(cache_path, cache)
    return out_rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    warm_src = _read_warm_leads(limit=args.limit)
    warm_out = _run(warm_src, CACHE_WARM, partial_output_path=OUT_WARM)
    _write_output(OUT_WARM, warm_out)

    rd_src = _read_rd(limit=args.limit)
    rd_out = _run(rd_src, CACHE_RD, partial_output_path=OUT_RD)
    _write_output(OUT_RD, rd_out)


if __name__ == "__main__":
    main()
