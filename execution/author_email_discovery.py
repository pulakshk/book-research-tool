#!/usr/bin/env python3
"""
AUTHOR EMAIL DISCOVERY — Find contact info for self-pub authors.
Scrapes Goodreads author pages, author websites, Amazon author pages,
and uses Gemini as a knowledge-based fallback.

Usage:
    python execution/author_email_discovery.py \
        --input "sub genre analysis/Sub genre analysis- Self Pub universe.xlsx" \
        --sheet "Ice Hockey and sports" \
        --output "data/author_contacts_ice_hockey.csv"
"""

import asyncio
import argparse
import csv
import json
import os
import random
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin, urlparse

import pandas as pd
from loguru import logger
from playwright.async_api import async_playwright
from dotenv import load_dotenv

# Load env
PROJECT_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
load_dotenv(os.path.join(PROJECT_ROOT, '.env'))
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# Gemini setup
try:
    import google.generativeai as genai
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model = genai.GenerativeModel('gemini-2.5-flash')
    else:
        gemini_model = None
        logger.warning("No GEMINI_API_KEY — Gemini fallback disabled")
except ImportError:
    gemini_model = None
    logger.warning("google-generativeai not installed — Gemini fallback disabled")

# ============================================================================
# CONFIGURATION
# ============================================================================

WORKER_COUNT = 4
SAVE_INTERVAL = 10
SLEEP_MIN = 2
SLEEP_MAX = 5
HEADLESS = True
CONTEXT_ROTATION_INTERVAL = 15

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2.1 Safari/605.1.15",
]

# Email regex
EMAIL_PATTERN = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

# Domains to exclude from email results
EXCLUDED_EMAIL_DOMAINS = {
    'example.com', 'email.com', 'sentry.io', 'wixpress.com', 'amazon.com',
    'goodreads.com', 'facebook.com', 'twitter.com', 'instagram.com',
    'googleapis.com', 'google.com', 'w3.org', 'schema.org', 'wordpress.org',
    'wordpress.com', 'gravatar.com', 'cloudflare.com', 'jquery.com',
    'bootstrapcdn.com', 'fontawesome.com', 'gstatic.com', 'cloudfront.net',
    'squarespace.com', 'shopify.com', 'mailchimp.com',
}

# File extensions to exclude (image/asset emails)
EXCLUDED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.css', '.js'}

# Contact page paths to try on author websites
CONTACT_PATHS = [
    '/contact', '/contact-me', '/contact-us', '/about', '/about-me',
    '/connect', '/reach-out', '/get-in-touch', '/hire-me',
    '/representation', '/literary-agent', '/agent',
]

# Keywords near emails that indicate quality contact info
CONTACT_KEYWORDS = [
    'email', 'contact', 'reach', 'inquir', 'represent', 'agent',
    'booking', 'business', 'press', 'media', 'rights', 'licensing',
    'manage', 'publicist', 'publicity',
]

# Social media URL patterns
SOCIAL_PATTERNS = {
    'twitter': re.compile(r'https?://(?:www\.)?(twitter\.com|x\.com)/[a-zA-Z0-9_]+', re.I),
    'instagram': re.compile(r'https?://(?:www\.)?instagram\.com/[a-zA-Z0-9_.]+', re.I),
    'facebook': re.compile(r'https?://(?:www\.)?facebook\.com/[a-zA-Z0-9.]+', re.I),
    'bookbub': re.compile(r'https?://(?:www\.)?bookbub\.com/(?:authors|profile)/[a-zA-Z0-9\-]+', re.I),
    'tiktok': re.compile(r'https?://(?:www\.)?tiktok\.com/@[a-zA-Z0-9_.]+', re.I),
}


# ============================================================================
# HELPERS
# ============================================================================

def is_valid_email(email):
    """Filter out junk emails."""
    email = email.lower().strip()
    domain = email.split('@')[-1]
    if domain in EXCLUDED_EMAIL_DOMAINS:
        return False
    if any(email.endswith(ext) for ext in EXCLUDED_EXTENSIONS):
        return False
    if len(email) < 6 or len(email) > 100:
        return False
    # Skip noreply / system emails
    if any(x in email for x in ['noreply', 'no-reply', 'donotreply', 'mailer-daemon', 'postmaster']):
        return False
    return True


def extract_emails_from_text(text):
    """Extract valid emails from text content."""
    if not text:
        return []
    raw = EMAIL_PATTERN.findall(text)
    return list(set(e for e in raw if is_valid_email(e)))


def extract_socials_from_text(text):
    """Extract social media URLs from text/HTML."""
    socials = {}
    for platform, pattern in SOCIAL_PATTERNS.items():
        matches = pattern.findall(text)
        if matches:
            # Take the first valid match
            full_match = pattern.search(text)
            if full_match:
                socials[platform] = full_match.group(0)
    return socials


async def create_stealth_context(browser):
    """Create a stealth browser context with anti-detection."""
    ua = random.choice(USER_AGENTS)
    context = await browser.new_context(
        user_agent=ua,
        viewport={"width": 1920, "height": 1080},
        locale="en-US",
        timezone_id="America/New_York",
        extra_http_headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "DNT": "1",
        }
    )
    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        window.chrome = { runtime: {} };
    """)
    return context


async def safe_goto(page, url, timeout=30000, retries=3):
    """Network-resilient navigation with retry."""
    for attempt in range(1, retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            await asyncio.sleep(random.uniform(0.5, 1.5))
            return True
        except Exception as e:
            logger.debug(f"  goto attempt {attempt}/{retries} failed for {url[:60]}: {e}")
            await asyncio.sleep(2 * attempt + random.uniform(1, 3))
    return False


# ============================================================================
# STAGE 1: GOODREADS AUTHOR PAGE
# ============================================================================

async def discover_from_goodreads(page, author_name, gr_book_link=None):
    """Find author's website and socials from Goodreads."""
    result = {'website': '', 'socials': {}}

    try:
        # Strategy A: If we have a Goodreads book link, navigate to author from there
        if gr_book_link and str(gr_book_link).startswith('http'):
            if not await safe_goto(page, gr_book_link):
                # Fall through to search
                pass
            else:
                await asyncio.sleep(random.uniform(1, 2))
                # Find author link on book page
                author_el = await page.query_selector(
                    "a.ContributorLink[href*='/author/show/'], "
                    "span.ContributorLinkWithAvatar a[href*='/author/']"
                )
                if author_el:
                    author_href = await author_el.get_attribute("href")
                    if author_href:
                        author_url = author_href if author_href.startswith('http') else f"https://www.goodreads.com{author_href}"
                        if await safe_goto(page, author_url):
                            await asyncio.sleep(random.uniform(1, 2))
                            return await _extract_gr_author_info(page)

        # Strategy B: Search Goodreads for the author
        search_url = f"https://www.goodreads.com/search?q={author_name.replace(' ', '+')}&search_type=authors"
        if not await safe_goto(page, search_url):
            return result

        await asyncio.sleep(random.uniform(1, 2))

        # Click first author result
        author_link = await page.query_selector("a.authorName[href*='/author/show/']")
        if author_link:
            href = await author_link.get_attribute("href")
            author_url = href if href.startswith('http') else f"https://www.goodreads.com{href}"
            if await safe_goto(page, author_url):
                await asyncio.sleep(random.uniform(1, 2))
                return await _extract_gr_author_info(page)

    except Exception as e:
        logger.debug(f"  GR author discovery error for {author_name}: {e}")

    return result


async def _extract_gr_author_info(page):
    """Extract website and social links from a Goodreads author profile page."""
    result = {'website': '', 'socials': {}}

    try:
        # Get full page HTML for social link extraction
        html = await page.content()
        result['socials'] = extract_socials_from_text(html)

        # Look for author website link
        # GR author pages have a "Website" link in the author info section
        website_selectors = [
            "a[href*='://'][rel='nofollow noopener noreferrer']:not([href*='goodreads']):not([href*='amazon'])",
            "div.dataItem a[href*='://']:not([href*='goodreads'])",
            "a.authorWebsite",
        ]
        for sel in website_selectors:
            els = await page.query_selector_all(sel)
            for el in els:
                href = await el.get_attribute("href")
                if href and not any(x in href.lower() for x in ['goodreads.com', 'amazon.com', 'facebook.com', 'twitter.com', 'instagram.com', 'tiktok.com', 'bookbub.com']):
                    text = (await el.text_content() or '').strip().lower()
                    # Likely a personal website
                    if any(x in text for x in ['website', 'site', 'blog', 'home']) or (href.startswith('http') and '.' in urlparse(href).netloc):
                        result['website'] = href
                        break
            if result['website']:
                break

        # If no website found via selectors, check for any external link in the info section
        if not result['website']:
            all_links = await page.query_selector_all("div.rightContainer a[href*='://'], div.aboutAuthor a[href*='://']")
            for link in all_links:
                href = await link.get_attribute("href")
                if href and not any(x in href.lower() for x in ['goodreads', 'amazon', 'facebook', 'twitter', 'instagram', 'tiktok', 'bookbub', 'youtube']):
                    parsed = urlparse(href)
                    if parsed.scheme in ('http', 'https') and '.' in parsed.netloc:
                        result['website'] = href
                        break

    except Exception as e:
        logger.debug(f"  GR author info extraction error: {e}")

    return result


# ============================================================================
# STAGE 2: AUTHOR WEBSITE EMAIL SCRAPING
# ============================================================================

async def discover_from_website(page, website_url):
    """Visit author website and find email/contact info."""
    result = {'emails': [], 'agent': '', 'socials': {}}

    if not website_url or not website_url.startswith('http'):
        return result

    try:
        # First, scrape the homepage
        if not await safe_goto(page, website_url):
            return result

        await asyncio.sleep(random.uniform(1, 2))
        html = await page.content()
        text = await page.evaluate("() => document.body ? document.body.innerText : ''")

        # Extract emails and socials from homepage
        result['emails'] = extract_emails_from_text(html)
        result['socials'] = extract_socials_from_text(html)

        # Check for literary agent/representation info
        agent_patterns = [
            r'(?:represented by|literary agent|agent:?|representation:?)\s*([A-Z][a-z]+ [A-Z][a-z]+(?:\s+(?:at|of|@)\s+[A-Za-z\s&]+)?)',
            r'(?:agent|representative|manager):\s*(.+?)(?:\n|<|$)',
        ]
        for pat in agent_patterns:
            m = re.search(pat, text, re.I)
            if m:
                result['agent'] = m.group(1).strip()[:100]
                break

        # If no email found on homepage, try contact pages
        if not result['emails']:
            base_url = f"{urlparse(website_url).scheme}://{urlparse(website_url).netloc}"
            for path in CONTACT_PATHS:
                contact_url = base_url + path
                try:
                    if await safe_goto(page, contact_url):
                        await asyncio.sleep(random.uniform(0.5, 1))
                        contact_html = await page.content()
                        emails = extract_emails_from_text(contact_html)
                        if emails:
                            result['emails'] = emails
                            # Also check for agent info on contact page
                            contact_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
                            for pat in agent_patterns:
                                m = re.search(pat, contact_text, re.I)
                                if m and not result['agent']:
                                    result['agent'] = m.group(1).strip()[:100]
                            break
                except Exception:
                    continue

        # Also look for mailto: links specifically (more reliable than regex on full HTML)
        try:
            mailto_els = await page.query_selector_all("a[href^='mailto:']")
            for el in mailto_els:
                href = await el.get_attribute("href")
                if href:
                    email = href.replace('mailto:', '').split('?')[0].strip()
                    if is_valid_email(email) and email not in result['emails']:
                        result['emails'].insert(0, email)  # Priority: mailto links first
        except Exception:
            pass

    except Exception as e:
        logger.debug(f"  Website discovery error for {website_url}: {e}")

    return result


# ============================================================================
# STAGE 3: AMAZON AUTHOR PAGE
# ============================================================================

async def discover_from_amazon(page, author_name):
    """Search Amazon for author page and extract contact/social links."""
    result = {'website': '', 'socials': {}}

    try:
        search_url = f"https://www.amazon.com/s?k={author_name.replace(' ', '+')}&i=digital-text&ref=nb_sb_noss"
        if not await safe_goto(page, search_url):
            return result

        await asyncio.sleep(random.uniform(2, 4))

        # Find author link in search results
        author_link = await page.query_selector(
            "a.a-link-normal[href*='/e/'], "
            "a[href*='/author/'], "
            "a.a-size-base.a-link-normal[href*='field-author']"
        )
        if not author_link:
            return result

        href = await author_link.get_attribute("href")
        if not href:
            return result

        author_url = href if href.startswith('http') else f"https://www.amazon.com{href}"

        # Navigate to author page
        if not await safe_goto(page, author_url):
            return result

        await asyncio.sleep(random.uniform(1, 2))
        html = await page.content()

        # Extract website from Amazon author page
        website_el = await page.query_selector(
            "a.a-link-normal[href*='://']:not([href*='amazon']):not([href*='audible'])"
        )
        if website_el:
            website_href = await website_el.get_attribute("href")
            if website_href and not any(x in website_href.lower() for x in ['amazon', 'audible', 'facebook', 'twitter', 'instagram']):
                result['website'] = website_href

        # Extract social links
        result['socials'] = extract_socials_from_text(html)

    except Exception as e:
        logger.debug(f"  Amazon author discovery error for {author_name}: {e}")

    return result


# ============================================================================
# STAGE 4: GEMINI FALLBACK
# ============================================================================

async def discover_from_gemini(author_name, series_name=''):
    """Use Gemini knowledge to find author contact info."""
    if not gemini_model:
        return {}

    try:
        prompt = f"""You are a book industry research assistant. I need to find the public contact information for the author "{author_name}".
{f'They wrote the "{series_name}" book series.' if series_name else ''}

Please provide any of the following that you know with HIGH CONFIDENCE (publicly available info only):
1. Their official website URL
2. Their public email address (from their website or social profiles)
3. Their literary agent or representation (agent name + agency)
4. Their Twitter/X handle
5. Their Instagram handle

Return ONLY valid JSON:
{{"website": "url or empty string", "email": "email or empty string", "agent": "agent info or empty string", "twitter": "handle or empty string", "instagram": "handle or empty string"}}

CRITICAL: Only include information you are CONFIDENT about. Return empty strings for anything uncertain."""

        response = await asyncio.to_thread(gemini_model.generate_content, prompt)
        text = response.text.strip()
        if '```json' in text:
            text = text.split('```json')[1].split('```')[0]
        elif '```' in text:
            text = text.split('```')[1].split('```')[0]
        return json.loads(text)

    except Exception as e:
        if "429" in str(e):
            logger.warning("  Gemini rate limit — waiting 5s")
            await asyncio.sleep(5)
        else:
            logger.debug(f"  Gemini fallback error for {author_name}: {e}")
        return {}


# ============================================================================
# MAIN DISCOVERY PIPELINE PER AUTHOR
# ============================================================================

async def discover_author_contact(page, author_name, gr_book_link=None, series_name=''):
    """Run the full 4-stage discovery pipeline for one author."""
    contact = {
        'Author Name': author_name,
        'Email': '',
        'Website': '',
        'Twitter': '',
        'Instagram': '',
        'Facebook': '',
        'BookBub': '',
        'TikTok': '',
        'Literary Agent': '',
        'Contact Source': '',
        'Discovery Status': 'Not Found',
    }

    all_socials = {}

    # --- STAGE 1: Goodreads Author Page ---
    logger.debug(f"  Stage 1: Goodreads for {author_name}")
    gr_result = await discover_from_goodreads(page, author_name, gr_book_link)
    website = gr_result.get('website', '')
    all_socials.update(gr_result.get('socials', {}))

    await asyncio.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    # --- STAGE 2: Author Website ---
    if website:
        logger.debug(f"  Stage 2: Website {website} for {author_name}")
        web_result = await discover_from_website(page, website)
        if web_result.get('emails'):
            contact['Email'] = web_result['emails'][0]
            contact['Contact Source'] = f"Website ({website})"
            contact['Discovery Status'] = 'Found Email'
        if web_result.get('agent'):
            contact['Literary Agent'] = web_result['agent']
        all_socials.update(web_result.get('socials', {}))
        contact['Website'] = website
        await asyncio.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    # --- STAGE 3: Amazon Author Page (if no email yet) ---
    if not contact['Email']:
        logger.debug(f"  Stage 3: Amazon for {author_name}")
        amz_result = await discover_from_amazon(page, author_name)
        if amz_result.get('website') and not website:
            website = amz_result['website']
            contact['Website'] = website
            # Try scraping this newly found website
            web_result = await discover_from_website(page, website)
            if web_result.get('emails'):
                contact['Email'] = web_result['emails'][0]
                contact['Contact Source'] = f"Amazon -> Website ({website})"
                contact['Discovery Status'] = 'Found Email'
            if web_result.get('agent') and not contact['Literary Agent']:
                contact['Literary Agent'] = web_result['agent']
            all_socials.update(web_result.get('socials', {}))
        all_socials.update(amz_result.get('socials', {}))
        await asyncio.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    # --- STAGE 4: Gemini Fallback (if no email yet) ---
    if not contact['Email']:
        logger.debug(f"  Stage 4: Gemini for {author_name}")
        gemini_result = await discover_from_gemini(author_name, series_name)
        if gemini_result:
            if gemini_result.get('email') and is_valid_email(gemini_result['email']):
                contact['Email'] = gemini_result['email']
                contact['Contact Source'] = 'Gemini (knowledge-based)'
                contact['Discovery Status'] = 'Found Email'
            if gemini_result.get('website') and not contact['Website']:
                contact['Website'] = gemini_result['website']
            if gemini_result.get('agent') and not contact['Literary Agent']:
                contact['Literary Agent'] = gemini_result['agent']
            # Map Gemini social results
            if gemini_result.get('twitter'):
                handle = gemini_result['twitter'].lstrip('@')
                all_socials.setdefault('twitter', f"https://x.com/{handle}")
            if gemini_result.get('instagram'):
                handle = gemini_result['instagram'].lstrip('@')
                all_socials.setdefault('instagram', f"https://instagram.com/{handle}")

    # --- Finalize ---
    contact['Twitter'] = all_socials.get('twitter', '')
    contact['Instagram'] = all_socials.get('instagram', '')
    contact['Facebook'] = all_socials.get('facebook', '')
    contact['BookBub'] = all_socials.get('bookbub', '')
    contact['TikTok'] = all_socials.get('tiktok', '')

    # Determine final status
    if contact['Email']:
        contact['Discovery Status'] = 'Found Email'
    elif contact['Website']:
        contact['Discovery Status'] = 'Found Website Only'
    elif any(contact[k] for k in ['Twitter', 'Instagram', 'Facebook', 'BookBub', 'TikTok']):
        contact['Discovery Status'] = 'Social Only'
    else:
        contact['Discovery Status'] = 'Not Found'

    return contact


# ============================================================================
# WORKER
# ============================================================================

async def worker(worker_id, browser, queue, results, lock, save_path, author_gr_map):
    """Async worker that processes authors from the queue."""
    context = await create_stealth_context(browser)
    page = await context.new_page()
    processed = 0

    try:
        while True:
            try:
                author_info = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            author_name = author_info['author']
            gr_link = author_info.get('gr_link', '')
            series_name = author_info.get('series', '')

            logger.info(f"  Worker {worker_id}: {author_name}")

            try:
                contact = await discover_author_contact(page, author_name, gr_link, series_name)

                async with lock:
                    results.append(contact)
                    processed += 1

                    # Save checkpoint
                    if processed % SAVE_INTERVAL == 0:
                        _save_results(results, save_path)
                        logger.info(f"  Worker {worker_id}: Saved checkpoint ({processed} authors)")

                status_icon = {
                    'Found Email': 'Y',
                    'Found Website Only': '~',
                    'Social Only': '~',
                    'Not Found': 'X',
                }.get(contact['Discovery Status'], '?')

                logger.info(
                    f"  Worker {worker_id}: [{status_icon}] {author_name} -> "
                    f"Email: {contact['Email'][:30] or 'N/A'} | "
                    f"Website: {'Yes' if contact['Website'] else 'No'} | "
                    f"Socials: {sum(1 for k in ['Twitter','Instagram','Facebook','BookBub','TikTok'] if contact[k])}"
                )

            except Exception as e:
                logger.error(f"  Worker {worker_id}: Error on {author_name}: {e}")

            # Rotate context periodically
            if processed % CONTEXT_ROTATION_INTERVAL == 0 and processed > 0:
                logger.info(f"  Worker {worker_id}: Rotating context...")
                await page.close()
                await context.close()
                context = await create_stealth_context(browser)
                page = await context.new_page()

            await asyncio.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    finally:
        await page.close()
        await context.close()

    logger.info(f"  Worker {worker_id}: Done — processed {processed} authors")


# ============================================================================
# I/O HELPERS
# ============================================================================

def load_authors_from_xlsx(xlsx_path, sheet_name):
    """Load unique authors from an xlsx sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    if not rows:
        return []

    header = rows[0]
    # Find column indices by name
    col_map = {}
    for i, h in enumerate(header):
        if h:
            h_lower = str(h).lower().strip()
            if 'author name' in h_lower:
                col_map['author'] = i
            elif 'book series name' in h_lower:
                col_map['series'] = i
            elif 'goodreads link' in h_lower:
                col_map['gr_link'] = i

    author_col = col_map.get('author', 1)
    series_col = col_map.get('series', 0)
    gr_col = col_map.get('gr_link', None)

    # Collect unique authors with their first GR link + series name
    seen = set()
    authors = []
    for row in rows[1:]:
        author = str(row[author_col] or '').strip()
        if not author or author.lower() in ('nan', 'none', ''):
            continue
        # Normalize for dedup
        author_key = author.lower()
        if author_key in seen:
            continue
        seen.add(author_key)

        series = str(row[series_col] or '').strip() if series_col is not None else ''
        gr_link = ''
        if gr_col is not None and len(row) > gr_col:
            gr_link = str(row[gr_col] or '').strip()
            if not gr_link.startswith('http'):
                gr_link = ''

        authors.append({
            'author': author,
            'series': series,
            'gr_link': gr_link,
        })

    wb.close()
    return authors


def load_authors_from_csv(csv_path):
    """Load unique authors from a CSV."""
    df = pd.read_csv(csv_path)
    seen = set()
    authors = []

    for _, row in df.iterrows():
        author = str(row.get('Author Name', '')).strip()
        if not author or author.lower() in ('nan', 'none', ''):
            continue
        if author.lower() in seen:
            continue
        seen.add(author.lower())

        gr_link = str(row.get('Goodreads Link', '')).strip()
        if not gr_link.startswith('http'):
            gr_link = ''
        series = str(row.get('Series Name', row.get('Book Series Name', ''))).strip()

        authors.append({
            'author': author,
            'series': series,
            'gr_link': gr_link,
        })

    return authors


def load_existing_results(output_path):
    """Load already-processed authors for resume capability."""
    if not os.path.exists(output_path):
        return set(), []
    df = pd.read_csv(output_path)
    processed = set(df['Author Name'].str.lower().str.strip())
    existing = df.to_dict('records')
    return processed, existing


def _save_results(results, output_path):
    """Save results to CSV."""
    if not results:
        return
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    df = pd.DataFrame(results)
    cols = [
        'Author Name', 'Email', 'Website', 'Twitter', 'Instagram',
        'Facebook', 'BookBub', 'TikTok', 'Literary Agent',
        'Contact Source', 'Discovery Status',
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ''
    df = df[cols]
    df.to_csv(output_path, index=False)


# ============================================================================
# MAIN
# ============================================================================

async def run_discovery(authors, output_path, resume=True):
    """Run the full discovery pipeline."""

    # Resume support
    already_processed = set()
    existing_results = []
    if resume:
        already_processed, existing_results = load_existing_results(output_path)
        if already_processed:
            logger.info(f"Resuming: {len(already_processed)} authors already processed")

    # Filter to unprocessed authors
    to_process = [a for a in authors if a['author'].lower().strip() not in already_processed]

    if not to_process:
        logger.success("All authors already processed!")
        return

    logger.info(f"\n{'='*60}")
    logger.info(f"AUTHOR EMAIL DISCOVERY")
    logger.info(f"Total authors: {len(authors)} | Already done: {len(already_processed)} | To process: {len(to_process)}")
    logger.info(f"Workers: {WORKER_COUNT} | Output: {output_path}")
    logger.info(f"{'='*60}\n")

    # Build queue
    queue = asyncio.Queue()
    for a in to_process:
        queue.put_nowait(a)

    results = list(existing_results)  # Start with existing
    lock = asyncio.Lock()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)

        workers = [
            worker(i, browser, queue, results, lock, output_path, {})
            for i in range(min(WORKER_COUNT, len(to_process)))
        ]

        await asyncio.gather(*workers)
        await browser.close()

    # Final save
    _save_results(results, output_path)

    # Stats
    total = len(results)
    emails_found = sum(1 for r in results if r.get('Email'))
    websites_found = sum(1 for r in results if r.get('Website'))
    socials_found = sum(1 for r in results if any(r.get(k) for k in ['Twitter', 'Instagram', 'Facebook', 'BookBub', 'TikTok']))

    logger.success(f"\nDiscovery Complete!")
    logger.info(f"  Total authors: {total}")
    logger.info(f"  Emails found: {emails_found} ({emails_found/total*100:.0f}%)" if total else "  No authors")
    logger.info(f"  Websites found: {websites_found} ({websites_found/total*100:.0f}%)" if total else "")
    logger.info(f"  With socials: {socials_found} ({socials_found/total*100:.0f}%)" if total else "")
    logger.info(f"  Output: {output_path}")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Author Email/Contact Discovery")
    parser.add_argument("--input", type=str, help="Path to xlsx file")
    parser.add_argument("--sheet", type=str, default="Ice Hockey and sports", help="Sheet name in xlsx")
    parser.add_argument("--input-csv", type=str, help="Path to CSV file (alternative to xlsx)")
    parser.add_argument("--output", type=str, default="data/author_contacts.csv", help="Output CSV path")
    parser.add_argument("--workers", type=int, default=4, help="Number of parallel workers")
    parser.add_argument("--visible", action="store_true", help="Run browser in visible mode")
    parser.add_argument("--no-resume", action="store_true", help="Start fresh (don't resume)")
    args = parser.parse_args()

    global WORKER_COUNT, HEADLESS  # noqa
    WORKER_COUNT = args.workers
    if args.visible:
        HEADLESS = False

    # Load authors
    if args.input_csv:
        authors = load_authors_from_csv(args.input_csv)
    elif args.input:
        authors = load_authors_from_xlsx(args.input, args.sheet)
    else:
        # Default
        default_xlsx = os.path.join(PROJECT_ROOT, "sub genre analysis", "Sub genre analysis- Self Pub universe.xlsx")
        if os.path.exists(default_xlsx):
            authors = load_authors_from_xlsx(default_xlsx, args.sheet)
        else:
            parser.print_help()
            return

    logger.info(f"Loaded {len(authors)} unique authors")

    if not authors:
        logger.error("No authors found in input!")
        return

    asyncio.run(run_discovery(authors, args.output, resume=not args.no_resume))


if __name__ == "__main__":
    main()
