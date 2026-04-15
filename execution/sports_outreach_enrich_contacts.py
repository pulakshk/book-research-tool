#!/usr/bin/env python3
"""
outreach/sports-romance.xlsx — Contact Enrichment
================================================
Fills missing/unclear author + agency email IDs for rows where:
- Outreach Status is one of: NEED CONTACT / NEW — NEEDS ENRICHMENT / BELOW LENGTH THRESHOLD
- Email Status does NOT contain "sent" (case-insensitive)

Anti-hallucination rules:
- Never invent emails.
- Only write an email if we have a source URL and the email string is found on that page.
- If no email is found, write a contact page URL (contact form) as a fallback with a clear type.

Outputs:
- Updates the workbook in-place (after making a timestamped backup in .tmp/backups/)
- Regenerates a sent-outreach view tab: "Clean Sent Outreach (Auto)"
- Creates a to-send view tab: "Outreach Ready (Auto)"

Usage:
  python3 execution/sports_outreach_enrich_contacts.py
  python3 execution/sports_outreach_enrich_contacts.py --limit-authors 25
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import shutil
import time
from pathlib import Path
from typing import Dict, Any, List, Tuple

import requests
from openpyxl import load_workbook

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from execution.websearch_email_discovery import discover_author_contact, _valid_email
DEFAULT_WB = PROJECT / "outreach" / "sheets" / "Sports Romance_ Outreach.xlsx"
CACHE_PATH = PROJECT / "outreach" / "sports-romance" / "source" / "sports_outreach_contact_cache.json"
SEED_CACHE_PATH = PROJECT / "outreach" / "sports-romance" / "source" / "email_discovery_cache.json"
BACKUP_DIR = PROJECT / ".tmp" / "backups"
LOG_PATH = PROJECT / "outreach/sheets" / "docs" / "SPORTS_EMAIL_VALIDATION_LOG.md"

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def _norm(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def _contains_sent(status: str) -> bool:
    s = (status or "").strip().lower()
    return "sent" in s


def _eligible_outreach_status(v: str) -> bool:
    v = (v or "").strip()
    return v in {"NEED CONTACT", "NEW — NEEDS ENRICHMENT", "NEW - NEEDS ENRICHMENT", "BELOW LENGTH THRESHOLD"}


def _load_cache() -> Dict[str, Any]:
    cache: Dict[str, Any] = {}
    # Seed from previous pipeline cache when available (reduces Gemini calls).
    if SEED_CACHE_PATH.exists():
        try:
            seed = json.loads(SEED_CACHE_PATH.read_text())
            if isinstance(seed, dict):
                for k, v in seed.items():
                    if isinstance(v, dict):
                        cache[str(k).strip().lower()] = v
        except Exception:
            pass
    if CACHE_PATH.exists():
        try:
            newer = json.loads(CACHE_PATH.read_text())
            if isinstance(newer, dict):
                for k, v in newer.items():
                    if isinstance(v, dict):
                        cache[str(k).strip().lower()] = v
        except Exception:
            pass
    return cache


def _has_contact_signal(record: Dict[str, Any]) -> bool:
    if not isinstance(record, dict):
        return False
    return any(
        _norm(record.get(key))
        for key in ["email", "agent_email", "contact_page", "website"]
    )


def _save_cache(cache: Dict[str, Any]) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, indent=2, ensure_ascii=False))


def _safe_get(url: str, timeout: int = 15) -> str:
    try:
        r = requests.get(url, timeout=timeout, allow_redirects=True, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        })
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        pass
    return ""


def _verify_email_on_page(email: str, source_url: str) -> bool:
    if not email or not source_url:
        return False
    if not _valid_email(email):
        return False
    html = _safe_get(source_url)
    if not html:
        return False
    # Basic check: email appears somewhere in visible text or HTML.
    return email.lower() in html.lower()


def _pick_confidence(result: Dict[str, Any]) -> str:
    if result.get("email") and result.get("email_source"):
        return "High"
    if result.get("agent_email") and result.get("agent_source"):
        return "Medium"
    if result.get("contact_page"):
        return "Not Found"
    return "Not Found"


def _backup_file(path: Path) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def _append_log(line: str) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(f"- [{timestamp}] {line}\n")


def _sheet_headers(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[str(v).strip()] = c
    return headers


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        # clear contents (keep sheet)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None
        return ws
    return wb.create_sheet(title=name)


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    for r_i, row in enumerate(rows, 2):
        for c_i, val in enumerate(row, 1):
            ws.cell(r_i, c_i).value = val


def _refresh_views_and_save(wb, ws, idx: Dict[str, int], wb_path: Path, cache: Dict[str, Any], note: str = "") -> None:
    _save_cache(cache)
    sent_headers, sent_rows = _build_sent_view(ws, idx)
    ws_sent = _ensure_sheet(wb, "Clean Sent Outreach (Auto)")
    _write_table(ws_sent, sent_headers, sent_rows)

    ready_headers, ready_rows = _build_ready_view(ws, idx)
    ws_ready = _ensure_sheet(wb, "Outreach Ready (Auto)")
    _write_table(ws_ready, ready_headers, ready_rows)

    wb.save(wb_path)
    if note:
        _append_log(note)


def _audit_existing_results(ws, idx: Dict[str, int], cache: Dict[str, Any]) -> Tuple[int, int]:
    """
    Normalize stale rows so the workbook reflects validated state, not older mixed logic.
    Returns: (rows_changed, critical_flags)
    """
    rows_changed = 0
    critical_flags = 0
    for r in range(2, ws.max_row + 1):
        author = _norm(ws.cell(r, idx["Author Name"]).value)
        if not author:
            continue
        outstat = _norm(ws.cell(r, idx["Outreach Status"]).value)
        if not _eligible_outreach_status(outstat):
            continue
        es = _norm(ws.cell(r, idx["Email Status"]).value)
        if _contains_sent(es):
            continue

        author_email = _norm(ws.cell(r, idx["Author Email ID"]).value)
        agency_email = _norm(ws.cell(r, idx["Agency Email ID"]).value)
        found_email = _norm(ws.cell(r, idx["Found Email"]).value)
        email_type = _norm(ws.cell(r, idx["Email Type"]).value)
        source_url = _norm(ws.cell(r, idx["Email Source URL"]).value)
        notes = _norm(ws.cell(r, idx["Research Notes"]).value)
        cached = cache.get(author.lower(), {}) if author else {}
        verified_author = bool(cached.get("verified_author_email"))
        verified_agent = bool(cached.get("verified_agent_email"))
        contact_page = _norm(cached.get("contact_page"))

        changed = False

        # If no verified email exists, we should not present a direct email as validated.
        if not verified_author and not verified_agent:
            if email_type in {"author_email", "author_direct", "agent_email", "agency_general"}:
                if author_email or agency_email or found_email:
                    ws.cell(r, idx["Author Email ID"]).value = None
                    ws.cell(r, idx["Agency Email ID"]).value = None
                    ws.cell(r, idx["Found Email"]).value = None
                    changed = True
            if contact_page or email_type == "contact_form_url":
                ws.cell(r, idx["Email Type"]).value = "contact_form_url"
                if contact_page and not source_url:
                    ws.cell(r, idx["Email Source URL"]).value = contact_page
                ws.cell(r, idx["Email Confidence"]).value = "Not Found"
                changed = True

        # If an author mismatch warning exists, keep it flagged for manual review.
        if "DATA CORRECTION" in notes or "Verify attribution" in notes:
            critical_flags += 1
            if ws.cell(r, idx["Email Confidence"]).value != "Medium":
                ws.cell(r, idx["Email Confidence"]).value = "Medium"
                changed = True

        # Inconsistent state: contact form type should not carry Found Email.
        if _norm(ws.cell(r, idx["Email Type"]).value) == "contact_form_url" and _norm(ws.cell(r, idx["Found Email"]).value):
            ws.cell(r, idx["Found Email"]).value = None
            changed = True

        if changed:
            rows_changed += 1
            if "Audit note:" not in notes:
                extra = "Audit note: normalized row to validated contact state."
                ws.cell(r, idx["Research Notes"]).value = (notes + " | " + extra).strip(" |") if notes else extra
    return rows_changed, critical_flags


def _build_sent_view(ws_src, idx: Dict[str, int]) -> Tuple[List[str], List[List[Any]]]:
    cols = [
        "Priority Band", "Priority Order", "Series Name", "Author Name", "Sub Genre",
        "Series Type", "# of Books in Series", "First Book Rating (Stars)", "First Book GR Ratings (#)",
        "Publisher / Author Name", "Author Email ID", "Agency Email ID", "Agent Name",
        "Email Status", "Outreach date", "Response Status", "Response Date", "Warm Lead?", "Boilerplate Sent?"
    ]
    rows = []
    es_col = idx.get("Email Status")
    if not es_col:
        return cols, rows
    for r in range(2, ws_src.max_row + 1):
        es = _norm(ws_src.cell(r, es_col).value)
        if not _contains_sent(es):
            continue
        rows.append([ws_src.cell(r, idx.get(c)).value if idx.get(c) else None for c in cols])
    # Stable ordering: Priority Band then Priority Order
    def _sort_key(row):
        pb = _norm(row[0])
        po = row[1] if isinstance(row[1], (int, float)) else 10**9
        rank = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "P5": 4}.get(pb, 9)
        return (rank, po)
    rows.sort(key=_sort_key)
    return cols, rows


def _build_ready_view(ws_src, idx: Dict[str, int]) -> Tuple[List[str], List[List[Any]]]:
    cols = [
        "Outreach Status", "Priority Band", "Priority Order",
        "Series Name", "Author Name",
        "Author Email ID", "Agency Email ID", "Agent Name",
        "Author Website", "Found Email", "Email Type", "Email Source URL", "Email Confidence",
        "Research Notes"
    ]
    rows = []
    for r in range(2, ws_src.max_row + 1):
        outstat = _norm(ws_src.cell(r, idx["Outreach Status"]).value)
        if not _eligible_outreach_status(outstat):
            continue
        es = _norm(ws_src.cell(r, idx["Email Status"]).value)
        if _contains_sent(es):
            continue
        author_email = _norm(ws_src.cell(r, idx["Author Email ID"]).value)
        agency_email = _norm(ws_src.cell(r, idx["Agency Email ID"]).value)
        found_email = _norm(ws_src.cell(r, idx["Found Email"]).value)
        source_url = _norm(ws_src.cell(r, idx["Email Source URL"]).value)
        if not (author_email or agency_email or found_email):
            continue
        if not source_url:
            continue
        rows.append([ws_src.cell(r, idx.get(c)).value if idx.get(c) else None for c in cols])
    return cols, rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=str, default=str(DEFAULT_WB))
    ap.add_argument("--limit-authors", type=int, default=0)
    ap.add_argument("--sleep", type=float, default=0.3)
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        raise SystemExit(f"Workbook not found: {wb_path}")

    backup_path = _backup_file(wb_path)
    _append_log(f"Started validation run for `{wb_path.name}`. Backup created at `{backup_path.name}`.")

    wb = load_workbook(wb_path)
    ws = wb["Picks for Licensing Outreach"]
    idx = _sheet_headers(ws)

    required = [
        "Series Name", "Author Name", "Outreach Status", "Email Status",
        "Author Email ID", "Agency Email ID", "Agent Name",
        "Author Website", "Contact Description",
        "Found Email", "Email Type", "Email Source URL", "Email Confidence", "Research Notes",
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit(f"Missing required columns in sheet header row: {missing}")

    # Build author -> has_sent mapping for cross-reference notes
    sent_series_by_author: Dict[str, List[str]] = {}
    for r in range(2, ws.max_row + 1):
        author = _norm(ws.cell(r, idx["Author Name"]).value)
        if not author:
            continue
        es = _norm(ws.cell(r, idx["Email Status"]).value)
        if _contains_sent(es):
            series = _norm(ws.cell(r, idx["Series Name"]).value)
            if series:
                sent_series_by_author.setdefault(author.lower(), []).append(series)

    # Collect eligible rows and unique authors to research
    author_to_series: Dict[str, str] = {}
    row_targets: List[int] = []
    for r in range(2, ws.max_row + 1):
        outstat = _norm(ws.cell(r, idx["Outreach Status"]).value)
        if not _eligible_outreach_status(outstat):
            continue
        es = _norm(ws.cell(r, idx["Email Status"]).value)
        if _contains_sent(es):
            continue

        # Needs research if no contact info or previously marked as not found/unclear.
        author_email = _norm(ws.cell(r, idx["Author Email ID"]).value)
        agency_email = _norm(ws.cell(r, idx["Agency Email ID"]).value)
        found_email = _norm(ws.cell(r, idx["Found Email"]).value)
        confidence = _norm(ws.cell(r, idx["Email Confidence"]).value).strip()
        if (author_email or agency_email or found_email) and confidence in {"High", "Medium"}:
            continue

        author = _norm(ws.cell(r, idx["Author Name"]).value)
        series = _norm(ws.cell(r, idx["Series Name"]).value)
        if not author:
            continue
        row_targets.append(r)
        author_key = author.lower()
        if author_key not in author_to_series and series:
            author_to_series[author_key] = series

    cache = _load_cache()
    authors = sorted(author_to_series.keys())
    if args.limit_authors > 0:
        authors = authors[: args.limit_authors]

    processed = 0
    for akey in authors:
        author = akey  # we'll store pretty author from any row later
        # Look up a canonical author string + series for prompt context
        series = author_to_series.get(akey, "")
        # Find a pretty author name from sheet
        pretty = None
        for r in row_targets:
            a = _norm(ws.cell(r, idx["Author Name"]).value)
            if a and a.lower() == akey:
                pretty = a
                break
        pretty = pretty or akey

        cached = cache.get(akey)
        if cached and _has_contact_signal(cached):
            result = cached
        else:
            result = discover_author_contact(pretty, series)
            result["last_checked_utc"] = dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            cache[akey] = result
        processed += 1

        # Verify emails against their source URLs before writing to sheet.
        verified_author = False
        verified_agent = False
        if result.get("email") and result.get("email_source"):
            verified_author = _verify_email_on_page(result["email"], result["email_source"])
            if not verified_author:
                result["email"] = ""
                result["email_source"] = ""
        if result.get("agent_email") and result.get("agent_source"):
            verified_agent = _verify_email_on_page(result["agent_email"], result["agent_source"])
            if not verified_agent:
                result["agent_email"] = ""
                result["agent_source"] = ""
        result["verified_author_email"] = verified_author
        result["verified_agent_email"] = verified_agent

        contact_method = "not_found"
        if verified_author:
            contact_method = "author_email_verified"
        elif verified_agent:
            contact_method = "agent_email_verified"
        elif result.get("contact_page"):
            contact_method = "contact_form_only"

        reasoning_flags = []
        if not verified_author and _norm(cache.get(akey, {}).get("email")):
            reasoning_flags.append("author_email_failed_source_check")
        if not verified_agent and _norm(cache.get(akey, {}).get("agent_email")):
            reasoning_flags.append("agent_email_failed_source_check")

        # Apply to all matching rows (same author), but only for eligible statuses.
        for r in row_targets:
            row_author = _norm(ws.cell(r, idx["Author Name"]).value).lower()
            if row_author != akey:
                continue
            outstat = _norm(ws.cell(r, idx["Outreach Status"]).value)
            if not _eligible_outreach_status(outstat):
                continue
            es = _norm(ws.cell(r, idx["Email Status"]).value)
            if _contains_sent(es):
                continue

            existing_author = _norm(ws.cell(r, idx["Author Email ID"]).value)
            existing_agency = _norm(ws.cell(r, idx["Agency Email ID"]).value)
            existing_found = _norm(ws.cell(r, idx["Found Email"]).value)

            # Only fill blanks or low-confidence records.
            if not existing_author and result.get("email"):
                ws.cell(r, idx["Author Email ID"]).value = result["email"]
            if not existing_agency and result.get("agent_email"):
                ws.cell(r, idx["Agency Email ID"]).value = result["agent_email"]
            if not _norm(ws.cell(r, idx["Agent Name"]).value) and result.get("agent_name"):
                ws.cell(r, idx["Agent Name"]).value = result["agent_name"]
            if not _norm(ws.cell(r, idx["Author Website"]).value) and result.get("website"):
                ws.cell(r, idx["Author Website"]).value = result["website"]

            # Choose primary contact for Found Email.
            if not existing_found:
                primary = result.get("email") or result.get("agent_email") or ""
                if primary:
                    ws.cell(r, idx["Found Email"]).value = primary

            # Email type & source URL
            email_type = ""
            source_url = ""
            if result.get("email"):
                email_type = "author_email"
                source_url = result.get("email_source", "")
            elif result.get("agent_email"):
                email_type = "agent_email"
                source_url = result.get("agent_source", "")
            elif result.get("contact_page"):
                email_type = "contact_form_url"
                source_url = result.get("contact_page", "")

            if email_type and not _norm(ws.cell(r, idx["Email Type"]).value):
                ws.cell(r, idx["Email Type"]).value = email_type
            if source_url and not _norm(ws.cell(r, idx["Email Source URL"]).value):
                ws.cell(r, idx["Email Source URL"]).value = source_url

            confidence = _pick_confidence(result)
            ws.cell(r, idx["Email Confidence"]).value = confidence

            # Research notes: cross-reference outreach done for same author
            notes = _norm(ws.cell(r, idx["Research Notes"]).value)
            sent_series = sent_series_by_author.get(akey, [])
            if sent_series:
                msg = f"Already outreached for this author on: {', '.join(sorted(set(sent_series)))}"
                if msg not in notes:
                    notes = (notes + " | " + msg).strip(" |") if notes else msg
            if reasoning_flags:
                flag_msg = "Validation flags: " + ", ".join(reasoning_flags)
                if flag_msg not in notes:
                    notes = (notes + " | " + flag_msg).strip(" |") if notes else flag_msg
            ws.cell(r, idx["Research Notes"]).value = notes

        if processed <= 5 or processed % 10 == 0:
            _append_log(
                f"Processed author `{pretty}` for series `{series}` -> {contact_method}."
            )
        if processed % 5 == 0:
            _refresh_views_and_save(
                wb, ws, idx, wb_path, cache,
                note=f"Checkpoint saved after {processed} authors."
            )
        time.sleep(args.sleep)

    audited_rows, critical_flags = _audit_existing_results(ws, idx, cache)
    _append_log(
        f"Post-run audit normalized {audited_rows} rows and flagged {critical_flags} rows for manual review."
    )
    _refresh_views_and_save(
        wb, ws, idx, wb_path, cache,
        note=f"Validation run completed. Processed {processed} unique authors."
    )


if __name__ == "__main__":
    main()
