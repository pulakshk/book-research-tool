#!/usr/bin/env python3
"""
Repair the Ice Hockey outreach workbook with evidence-backed contact and
Goodreads series verification.

Scope:
- Prioritize the "Ice Hockey Outreach Updates" sheet for outreach readiness.
- Reuse the validated author contacts to clean the master
  "Ice Hockey & Sports Romance" sheet as well.
- Never invent emails: only keep an email if it is present on a public page we
  fetched successfully.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qs, quote_plus, unquote, urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = PROJECT_ROOT / "outreach" / "ice-hockey"
SOURCE_DIR = OUT_DIR / "source"
REPORTS_DIR = OUT_DIR / "reports"
VERIFIED_DIR = OUT_DIR / "verified"
EXPORTS_DIR = OUT_DIR / "exports"
PROGRESS_DIR = OUT_DIR / "progress"
WORKBOOK_PATH = SOURCE_DIR / "Final self-pub scored.xlsx"

OUTREACH_SHEET = "Ice Hockey Outreach Updates"
MASTER_SHEET = "Ice Hockey & Sports Romance"

GOODREADS_BASE = "https://www.goodreads.com"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
RATING_RE = re.compile(r"([0-9.]+)\s+avg rating\s+[—-]\s+([0-9,]+)\s+ratings?", re.I)
PRIMARY_WORKS_RE = re.compile(r"(\d+)\s+primary works", re.I)
TOTAL_WORKS_RE = re.compile(r"(\d+)\s+total works", re.I)

SOCIAL_HOSTS = {
    "facebook.com",
    "twitter.com",
    "x.com",
    "instagram.com",
    "tiktok.com",
    "bookbub.com",
    "goodreads.com",
    "amazon.com",
    "wikipedia.org",
    "linktr.ee",
    "mailchi.mp",
    "substack.com",
    "youtube.com",
}

BAD_EMAILS = {
    "user@domain.com",
    "email@example.com",
    "example@example.com",
    "author@directauthor.com",
}

BAD_EMAIL_DOMAINS = {
    "example.com",
    "domain.com",
    "email.com",
}

BAD_URL_PARTS = [
    "goodreads.com",
    "amazon.com",
    "facebook.com",
    "instagram.com",
    "twitter.com",
    "x.com",
    "bookbub.com",
]

CONTACT_PATHS = [
    "",
    "/contact",
    "/about",
    "/representation",
    "/rights",
]

LOCAL_BOOK_SOURCES = [
    PROJECT_ROOT / "data" / "final_v1_sports_hockey_dump.csv",
    PROJECT_ROOT / "data" / "unified_book_data_enriched_ultra.csv",
    PROJECT_ROOT / "data" / "master_backups" / "ice_hockey_724_complete_database.xlsx",
]

LOCAL_AUTHOR_SOURCES = [
    PROJECT_ROOT / "data" / "author_contacts_ice_hockey.csv",
    PROJECT_ROOT / "data" / "ice_hockey_full_merged.csv",
]

VERIFICATION_COLUMNS = [
    "Validated_Email",
    "Email_Verified",
    "Email_Source_URL",
    "Email_Source_Type",
    "Validated_Website",
    "Agency_Contact",
    "Agency_Source",
    "Verified_Series_Name",
    "Verified_Goodreads_Series_URL",
    "Verified_Books_in_Series",
    "Verified_Type",
    "Verified_First_Book_Name",
    "Verified_Last_Book_Name",
    "Verified_Total_Pages",
    "Verified_Length_of_Adaption_in_Hours",
    "Series_Primary_Works",
    "Series_Total_Works",
    "Data_Quality_Flag",
    "Verification_Notes",
]

CHECKPOINT_AUTHOR_CSV = PROGRESS_DIR / "author_contacts.partial.csv"
CHECKPOINT_OUTREACH_CSV = PROGRESS_DIR / "outreach.partial.csv"
CHECKPOINT_MASTER_CSV = PROGRESS_DIR / "master.partial.csv"


def norm(text: object) -> str:
    text = "" if text is None else str(text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def norm_key(text: object) -> str:
    text = norm(text)
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"\bseries\b", "", text)
    text = re.sub(r"\bvol(?:ume)?\.?\s*\d+\b", "", text)
    text = re.sub(r"#\s*\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?", "", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_series_label(text: object) -> str:
    text = "" if text is None else str(text).strip()
    text = re.sub(r"\s*\([^)]*#\s*\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?[^)]*\)\s*$", "", text)
    text = re.sub(r"\s*,?\s*#\s*\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?\s*$", "", text)
    return text.strip(" ,")


def title_similarity(a: object, b: object) -> float:
    from difflib import SequenceMatcher

    aa = norm_key(a)
    bb = norm_key(b)
    if not aa or not bb:
        return 0.0
    if aa == bb:
        return 1.0
    if aa in bb or bb in aa:
        return 0.9
    return SequenceMatcher(None, aa, bb).ratio()


def safe_float(value: object) -> Optional[float]:
    try:
        val = float(value)
        if math.isnan(val):
            return None
        return val
    except Exception:
        return None


def safe_int(value: object) -> Optional[int]:
    val = safe_float(value)
    if val is None:
        return None
    return int(round(val))


def ensure_url(url: object) -> str:
    raw = norm(url)
    if raw in {"", "nan", "none", "null"}:
        return ""
    if raw.startswith("http://") or raw.startswith("https://"):
        return str(url).strip()
    return f"https://{str(url).strip()}"


def host_from_url(url: str) -> str:
    if not url:
        return ""
    return urlparse(url).netloc.lower().replace("www.", "")


def strip_goodreads_tracking(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    path = parsed.path
    clean = f"{parsed.scheme}://{parsed.netloc}{path}"
    return clean


def goodreads_id(url: str) -> str:
    m = re.search(r"/(?:book/show|series)/(\d+)", url or "")
    return m.group(1) if m else ""


def goodreads_slug_matches_title(url: str, title: object) -> bool:
    if "goodreads.com/book/show/" not in (url or ""):
        return False
    path = urlparse(url).path
    slug = path.split("/")[-1]
    slug = re.sub(r"^\d+-?", "", slug)
    lowered = slug.lower()
    if any(token in lowered for token in ["summary-of-", "analysis-of-", "study-guide", "book-review", "quick-read"]):
        return False
    slug = slug.replace("-", " ")
    return title_similarity(slug, title) >= 0.55


def clean_email(email: str) -> str:
    email = email.strip().strip(".;,)")
    if email.lower().startswith("mailto:"):
        email = email[7:]
    return email


def is_valid_public_email(email: str) -> bool:
    email = clean_email(email).lower()
    if not email or email in BAD_EMAILS:
        return False
    if "@" not in email:
        return False
    domain = email.split("@", 1)[1]
    if domain in BAD_EMAIL_DOMAINS:
        return False
    if domain.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".css", ".js")):
        return False
    if any(token in domain for token in ["sentry", "wixpress", "cloudfront", "gr-assets", "amazonaws", "mailchimp"]):
        return False
    if any(token in email for token in ["badge@", "@2x-", "static", "sprite"]):
        return False
    if any(token in email for token in ["noreply", "no-reply", "donotreply", "example"]):
        return False
    return True


def extract_emails_from_html(html: str) -> List[str]:
    return sorted({clean_email(e) for e in EMAIL_RE.findall(html) if is_valid_public_email(clean_email(e))})


def decode_ddg_href(href: str) -> str:
    if not href:
        return ""
    if href.startswith("http://") or href.startswith("https://"):
        return href
    parsed = urlparse(href)
    query = parse_qs(parsed.query)
    if "uddg" in query:
        return unquote(query["uddg"][0])
    return href


def is_candidate_website(url: str) -> bool:
    host = host_from_url(url)
    if not host or host == "nan":
        return False
    if any(host.endswith(bad) for bad in SOCIAL_HOSTS):
        return False
    return True


def page_mentions_author(soup: BeautifulSoup, author: str) -> bool:
    tokens = [tok for tok in norm_key(author).split() if len(tok) > 2]
    if not tokens:
        return True
    text_parts = []
    if soup.title:
        text_parts.append(soup.title.get_text(" ", strip=True))
    for selector in ["h1", "h2", "meta[name='description']"]:
        for el in soup.select(selector)[:5]:
            content = el.get("content") if selector.startswith("meta") else el.get_text(" ", strip=True)
            if content:
                text_parts.append(content)
    sample = " ".join(text_parts).lower()
    if not sample:
        return False
    matches = sum(1 for tok in tokens if tok in sample)
    return matches >= min(2, len(tokens))


def is_plausible_agent_text(text: object) -> bool:
    raw = str(text or "").strip()
    if not raw:
        return False
    tokens = re.findall(r"[A-Za-z]+", raw)
    if not tokens:
        return False
    stopwords = {"the", "and", "into", "for", "with", "of", "at", "to", "from", "on", "in"}
    if all(tok.lower() in stopwords for tok in tokens):
        return False
    if raw.lower() == raw and len(tokens) <= 2:
        return False
    return True


@dataclass
class LocalBook:
    author: str
    author_key: str
    series: str
    series_key: str
    title: str
    title_key: str
    goodreads_link: str
    pages: Optional[int]
    rating: Optional[float]
    rating_count: Optional[int]
    book_number: Optional[float]
    publication_date: str
    total_books_in_series: Optional[int]
    source: str


@dataclass
class LocalSeries:
    author: str
    author_key: str
    series: str
    series_key: str
    books: List[LocalBook] = field(default_factory=list)


@dataclass
class AuthorCandidate:
    author: str
    author_key: str
    websites: set = field(default_factory=set)
    emails: set = field(default_factory=set)
    literary_agents: set = field(default_factory=set)


class CachedHttp:
    def __init__(self, delay: float = 0.15):
        self._cache: Dict[str, Optional[str]] = {}
        self._lock = threading.Lock()
        self.delay = delay

    def get(self, url: str, timeout: int = 12) -> Optional[str]:
        url = strip_goodreads_tracking(url) if "goodreads.com" in url else url
        with self._lock:
            if url in self._cache:
                return self._cache[url]

        try:
            time.sleep(self.delay)
            response = requests.get(
                url,
                timeout=timeout,
                headers={"User-Agent": USER_AGENT},
                allow_redirects=True,
            )
            if response.status_code >= 400:
                html = None
            else:
                html = response.text
        except Exception:
            html = None

        with self._lock:
            self._cache[url] = html
        return html


class GoodreadsClient:
    def __init__(self, http: CachedHttp):
        self.http = http
        self.book_cache: Dict[str, Dict[str, object]] = {}
        self.series_cache: Dict[str, Dict[str, object]] = {}
        self.search_cache: Dict[str, List[Dict[str, str]]] = {}

    def search_books(self, query: str) -> List[Dict[str, str]]:
        query = query.strip()
        if not query:
            return []
        if query in self.search_cache:
            return self.search_cache[query]

        url = f"{GOODREADS_BASE}/search?q={quote_plus(query)}&search_type=books"
        html = self.http.get(url)
        if not html:
            self.search_cache[query] = []
            return []

        soup = BeautifulSoup(html, "html.parser")
        results: List[Dict[str, str]] = []
        for row in soup.select("tr[itemtype]"):
            title_el = row.select_one("a.bookTitle")
            author_el = row.select_one("a.authorName span[itemprop='name'], a.authorName")
            if not title_el or not author_el:
                continue
            title_text = title_el.get_text(" ", strip=True)
            href = title_el.get("href") or ""
            full_url = urljoin(GOODREADS_BASE, href)
            title_base = re.sub(r"\s*\(.*?\)\s*$", "", title_text).strip()
            results.append(
                {
                    "title": title_text,
                    "title_base": title_base,
                    "author": author_el.get_text(" ", strip=True),
                    "url": full_url,
                }
            )
        self.search_cache[query] = results
        return results

    def fetch_book(self, url: str) -> Dict[str, object]:
        url = strip_goodreads_tracking(url)
        if url in self.book_cache:
            return self.book_cache[url]

        html = self.http.get(url)
        if not html:
            self.book_cache[url] = {}
            return {}

        soup = BeautifulSoup(html, "html.parser")
        data: Dict[str, object] = {"url": url}

        title_el = soup.select_one("h1[data-testid='bookTitle']")
        if title_el:
            data["title"] = title_el.get_text(" ", strip=True)

        author_el = soup.select_one("span.ContributorLink__name, a.authorName span[itemprop='name'], a.ContributorLink")
        if author_el:
            data["author"] = author_el.get_text(" ", strip=True)

        series_el = soup.select_one("a[href*='/series/']")
        if series_el:
            data["series_url"] = urljoin(GOODREADS_BASE, series_el.get("href") or "")
            series_text = series_el.get_text(" ", strip=True)
            data["series_text"] = series_text
            data["series_name"] = re.sub(r"\s*#.*$", "", series_text).strip()

        pages_el = soup.select_one("p[data-testid='pagesFormat']")
        if pages_el:
            m = re.search(r"(\d+)\s+pages", pages_el.get_text(" ", strip=True), re.I)
            if m:
                data["pages"] = int(m.group(1))

        rating_el = soup.select_one("div.RatingStatistics__rating, [data-testid='averageRating']")
        if rating_el:
            try:
                data["rating"] = float(rating_el.get_text(" ", strip=True))
            except Exception:
                pass

        rating_count_text = " ".join(
            el.get_text(" ", strip=True)
            for el in soup.select("[data-testid='ratingsCount'], span[data-testid='ratingsCount']")
        )
        m = re.search(r"([0-9,]+)\s+ratings", rating_count_text, re.I)
        if m:
            data["rating_count"] = int(m.group(1).replace(",", ""))

        self.book_cache[url] = data
        return data

    def fetch_series(self, url: str) -> Dict[str, object]:
        url = strip_goodreads_tracking(url)
        if url in self.series_cache:
            return self.series_cache[url]

        html = self.http.get(url)
        if not html:
            self.series_cache[url] = {}
            return {}

        soup = BeautifulSoup(html, "html.parser")
        all_text = " ".join(soup.stripped_strings)
        name_el = soup.select_one("h1")
        heading = name_el.get_text(" ", strip=True) if name_el else ""
        name = re.sub(r"\s+Series$", "", heading).strip() or heading

        primary_works = None
        total_works = None
        m = PRIMARY_WORKS_RE.search(all_text)
        if m:
            primary_works = int(m.group(1))
        m = TOTAL_WORKS_RE.search(all_text)
        if m:
            total_works = int(m.group(1))

        books: List[Dict[str, object]] = []
        for item in soup.select("div.listWithDividers__item"):
            number_text = ""
            title = ""
            book_url = ""
            author = ""
            rating = None
            rating_count = None

            h3 = item.select_one("h3")
            if h3:
                number_text = h3.get_text(" ", strip=True)

            title_el = item.select_one("a[itemprop='url'], a.gr-h3--serif")
            if title_el:
                title = title_el.get_text(" ", strip=True)
                book_url = urljoin(GOODREADS_BASE, title_el.get("href") or "")

            author_el = item.select_one("span[itemprop='author'] span[itemprop='name'], a.authorName")
            if author_el:
                author = author_el.get_text(" ", strip=True)

            item_text = " ".join(item.stripped_strings)
            m = RATING_RE.search(item_text)
            if m:
                try:
                    rating = float(m.group(1))
                except Exception:
                    rating = None
                rating_count = int(m.group(2).replace(",", ""))

            num = None
            m = re.search(r"Book\s+(\d+(?:\.\d+)?)", number_text, re.I)
            if m:
                num = float(m.group(1))

            if title and book_url:
                books.append(
                    {
                        "number": num,
                        "title": title,
                        "url": book_url,
                        "author": author,
                        "rating": rating,
                        "rating_count": rating_count,
                    }
                )

        result = {
            "url": url,
            "series_name": name,
            "primary_works": primary_works,
            "total_works": total_works,
            "books": books,
        }
        self.series_cache[url] = result
        return result


def load_sheet_rows(sheet_name: str) -> List[Dict[str, object]]:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_row = 2 if sheet_name == OUTREACH_SHEET else 1
    data_start = 3 if sheet_name == OUTREACH_SHEET else 2
    headers = [cell for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    rows: List[Dict[str, object]] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        first = values[0] if len(values) > 0 else None
        author = values[1] if len(values) > 1 else None
        if first is None and author is None:
            continue
        record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        record["_sheet"] = sheet_name
        record["_excel_row"] = excel_row
        record["_author_key"] = norm_key(author)
        record["_series_key"] = norm_key(first)
        first_book_col = "First Book Name"
        record["_first_book_key"] = norm_key(record.get(first_book_col))
        rows.append(record)
    return rows


def load_backup_books() -> Iterable[LocalBook]:
    workbook = load_workbook(LOCAL_BOOK_SOURCES[2], read_only=True, data_only=True)
    ws = workbook["All 724 Books"]
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: headers.index(name) for name in headers if name}
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[idx["Title"]]
        author = row[idx["Author"]]
        series = clean_series_label(row[idx["Series"]])
        link = row[idx["Goodreads_Link"]]
        if not title or not author or not series or not link:
            continue
        if not goodreads_slug_matches_title(str(link), title):
            continue
        yield LocalBook(
            author=str(author).strip(),
            author_key=norm_key(author),
            series=str(series).strip(),
            series_key=norm_key(series),
            title=str(title).strip(),
            title_key=norm_key(title),
            goodreads_link=strip_goodreads_tracking(str(link).strip()),
            pages=None,
            rating=None,
            rating_count=None,
            book_number=safe_float(row[idx.get("Series_Number", 0)] if "Series_Number" in idx else None),
            publication_date="",
            total_books_in_series=None,
            source="ice_hockey_724_complete_database.xlsx",
        )


def load_local_books() -> Tuple[Dict[Tuple[str, str], LocalSeries], Dict[Tuple[str, str], LocalBook]]:
    book_candidates: Dict[str, LocalBook] = {}

    def add_book(book: LocalBook) -> None:
        key = goodreads_id(book.goodreads_link) or f"{book.title_key}|{book.author_key}"
        existing = book_candidates.get(key)
        if existing is None:
            book_candidates[key] = book
            return
        existing_score = score_local_book(existing)
        new_score = score_local_book(book)
        if new_score > existing_score:
            book_candidates[key] = book

    final_dump = pd.read_csv(LOCAL_BOOK_SOURCES[0], low_memory=False)
    for _, row in final_dump.iterrows():
        title = row.get("Book Name")
        author = row.get("Author Name")
        series = clean_series_label(row.get("Series Name"))
        link = row.get("Goodreads Link")
        if not title or not author or not series or not link:
            continue
        if not goodreads_slug_matches_title(str(link), title):
            continue
        add_book(
            LocalBook(
                author=str(author).strip(),
                author_key=norm_key(author),
                series=str(series).strip(),
                series_key=norm_key(series),
                title=str(title).strip(),
                title_key=norm_key(title),
                goodreads_link=strip_goodreads_tracking(str(link).strip()),
                pages=safe_int(row.get("Pages")),
                rating=safe_float(row.get("Goodreads Rating")),
                rating_count=safe_int(row.get("Goodreads # of Ratings")),
                book_number=safe_float(row.get("Book Number")),
                publication_date=str(row.get("Publication Date") or "").strip(),
                total_books_in_series=safe_int(row.get("Total Books in Series")),
                source="final_v1_sports_hockey_dump.csv",
            )
        )

    ultra = pd.read_csv(LOCAL_BOOK_SOURCES[1], low_memory=False)
    ultra = ultra[ultra["Primary Subgenre"].astype(str).str.contains("Hockey", case=False, na=False)].copy()
    for _, row in ultra.iterrows():
        title = row.get("Book Name")
        author = row.get("Author Name")
        series = clean_series_label(row.get("Series Name"))
        link = row.get("Goodreads Link")
        if not title or not author or not series or not link:
            continue
        if not goodreads_slug_matches_title(str(link), title):
            continue
        add_book(
            LocalBook(
                author=str(author).strip(),
                author_key=norm_key(author),
                series=str(series).strip(),
                series_key=norm_key(series),
                title=str(title).strip(),
                title_key=norm_key(title),
                goodreads_link=strip_goodreads_tracking(str(link).strip()),
                pages=safe_int(row.get("Pages")),
                rating=safe_float(row.get("Goodreads Rating")),
                rating_count=safe_int(row.get("Goodreads # of Ratings")),
                book_number=safe_float(row.get("Book Number")),
                publication_date=str(row.get("Publication Date") or "").strip(),
                total_books_in_series=safe_int(row.get("Total Books in Series")),
                source="unified_book_data_enriched_ultra.csv",
            )
        )

    for book in load_backup_books():
        add_book(book)

    series_map: Dict[Tuple[str, str], LocalSeries] = {}
    by_book_author: Dict[Tuple[str, str], LocalBook] = {}
    for book in book_candidates.values():
        series_key = (book.series_key, book.author_key)
        if series_key not in series_map:
            series_map[series_key] = LocalSeries(
                author=book.author,
                author_key=book.author_key,
                series=book.series,
                series_key=book.series_key,
            )
        series_map[series_key].books.append(book)
        by_book_author[(book.title_key, book.author_key)] = book
    return series_map, by_book_author


def score_local_book(book: LocalBook) -> int:
    score = 0
    if book.pages:
        score += 3
    if book.rating_count:
        score += 2
    if book.book_number is not None:
        score += 2
    if book.source == "final_v1_sports_hockey_dump.csv":
        score += 3
    if book.source == "ice_hockey_724_complete_database.xlsx":
        score += 1
    return score


def build_author_candidates(sheet_rows: Iterable[Dict[str, object]]) -> Dict[str, AuthorCandidate]:
    authors: Dict[str, AuthorCandidate] = {}

    def ensure(author: object) -> AuthorCandidate:
        key = norm_key(author)
        if key not in authors:
            authors[key] = AuthorCandidate(author=str(author).strip(), author_key=key)
        return authors[key]

    for row in sheet_rows:
        author = row.get("Author Name")
        if not author:
            continue
        entry = ensure(author)
        for email in [row.get("Email")]:
            if email and is_valid_public_email(str(email)):
                entry.emails.add(clean_email(str(email)))
        website = ensure_url(row.get("Website"))
        if website and is_candidate_website(website):
            entry.websites.add(website)
        agent = row.get("Literary Agent")
        if agent and norm(agent) and is_plausible_agent_text(agent):
            entry.literary_agents.add(str(agent).strip())

    for path in LOCAL_AUTHOR_SOURCES:
        df = pd.read_csv(path, low_memory=False)
        for _, row in df.iterrows():
            author = row.get("Author Name")
            if not author:
                continue
            key = norm_key(author)
            if key not in authors:
                continue
            entry = authors[key]
            email = row.get("Email")
            if email and is_valid_public_email(str(email)):
                entry.emails.add(clean_email(str(email)))
            website = ensure_url(row.get("Website"))
            if website and is_candidate_website(website):
                entry.websites.add(website)
            agent = row.get("Literary Agent")
            if agent and norm(agent) and is_plausible_agent_text(agent):
                entry.literary_agents.add(str(agent).strip())

    return authors


def choose_best_email(
    emails: Iterable[str],
    base_host: str,
    candidate_emails: Iterable[str],
    mailto_emails: Iterable[str],
) -> str:
    candidate_set = {clean_email(e).lower() for e in candidate_emails if is_valid_public_email(clean_email(e))}
    mailto_set = {clean_email(e).lower() for e in mailto_emails if is_valid_public_email(clean_email(e))}
    scored: List[Tuple[int, str]] = []
    for email in emails:
        clean = clean_email(email)
        if not is_valid_public_email(clean):
            continue
        score = 0
        lower = clean.lower()
        domain = lower.split("@", 1)[1]
        if lower in candidate_set:
            score += 6
        if lower in mailto_set:
            score += 5
        if base_host and domain.endswith(base_host):
            score += 4
        if any(token in lower for token in ["contact", "hello", "rights", "media", "office", "books"]):
            score += 2
        if lower.startswith("info@"):
            score += 1
        scored.append((score, clean))
    if not scored:
        return ""
    scored.sort(key=lambda item: (-item[0], item[1]))
    return scored[0][1]


def extract_agency_hint(
    soup: BeautifulSoup,
    literary_agents: Iterable[str],
) -> Tuple[str, str]:
    text = " ".join(soup.stripped_strings)
    text_lower = text.lower()
    for agent in literary_agents:
        agent_text = str(agent).strip()
        if not agent_text:
            continue
        tokens = [tok for tok in norm_key(agent_text).split() if len(tok) > 3]
        if tokens and all(tok in text_lower for tok in tokens[:2]):
            return agent_text, "official-site"
    if soup.select_one("form"):
        return "Contact form on official website", "official-site"
    return "", ""


def search_official_site(http: CachedHttp, author: str) -> str:
    query = f'{author} author official website'
    url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    html = http.get(url)
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    for link in soup.select("a.result__a, a[href]"):
        href = decode_ddg_href(link.get("href") or "")
        if not href.startswith("http"):
            continue
        host = host_from_url(href)
        if not host or any(host.endswith(bad) for bad in SOCIAL_HOSTS):
            continue
        return href
    return ""


def expand_contact_urls(base_url: str) -> List[str]:
    base_url = ensure_url(base_url)
    if not base_url:
        return []
    urls = []
    for suffix in CONTACT_PATHS:
        urls.append(urljoin(base_url.rstrip("/") + "/", suffix.lstrip("/")))
    return list(dict.fromkeys(urls))


def validate_author_contact(http: CachedHttp, candidate: AuthorCandidate) -> Dict[str, object]:
    websites = [ensure_url(url) for url in candidate.websites if ensure_url(url) and is_candidate_website(ensure_url(url))]
    if not websites:
        fallback = search_official_site(http, candidate.author)
        if fallback:
            websites = [fallback]

    best_email = ""
    email_source = ""
    email_source_type = ""
    agency_contact = ""
    agency_source = ""
    validated_website = websites[0] if websites else ""
    notes: List[str] = []

    for website in websites[:1]:
        validated_website = website
        base_host = host_from_url(website)
        for page_url in expand_contact_urls(website):
            html = http.get(page_url)
            if not html:
                continue
            page_host = host_from_url(page_url)
            soup = BeautifulSoup(html, "html.parser")
            if not page_mentions_author(soup, candidate.author):
                continue
            mailto_emails = [clean_email(a.get("href", "")) for a in soup.select("a[href^='mailto:']")]
            page_emails = extract_emails_from_html(html)
            chosen = ""
            if is_candidate_website(page_url) or page_host == base_host:
                chosen = choose_best_email(page_emails + mailto_emails, base_host, candidate.emails, mailto_emails)
            if chosen:
                best_email = chosen
                email_source = page_url
                email_source_type = "official-website"
                notes.append(f"email found on {host_from_url(page_url)}")
                break
            if not agency_contact:
                agency_contact, agency_source = extract_agency_hint(soup, candidate.literary_agents)
                if agency_contact:
                    if agency_contact == "Contact form on official website":
                        notes.append("official site has contact form but no public email")
                    else:
                        notes.append("agency or representation found on official site")
            if "linktr.ee" in page_url:
                notes.append("linktr.ee profile found")
        if best_email:
            break

    if best_email:
        data_quality = "GREEN"
    elif agency_contact or validated_website:
        data_quality = "YELLOW"
    else:
        data_quality = "RED"
    return {
        "Author Name": candidate.author,
        "Author_Checked": True,
        "Validated_Email": best_email,
        "Email_Verified": bool(best_email),
        "Email_Source_URL": email_source,
        "Email_Source_Type": email_source_type,
        "Validated_Website": validated_website if is_candidate_website(validated_website) else "",
        "Agency_Contact": agency_contact,
        "Agency_Source": agency_source,
        "Author_Data_Quality_Flag": data_quality,
        "Author_Verification_Notes": "; ".join(dict.fromkeys(notes)),
    }


def match_local_series(
    row: Dict[str, object],
    series_map: Dict[Tuple[str, str], LocalSeries],
    by_book_author: Dict[Tuple[str, str], LocalBook],
) -> Tuple[Optional[LocalSeries], str]:
    author_key = row["_author_key"]
    series_key = row["_series_key"]
    first_book_key = row["_first_book_key"]

    if first_book_key and (first_book_key, author_key) in by_book_author:
        book = by_book_author[(first_book_key, author_key)]
        return series_map.get((book.series_key, author_key)), "local-first-book"

    if (series_key, author_key) in series_map:
        return series_map[(series_key, author_key)], "local-series"

    if (series_key, author_key) in by_book_author:
        book = by_book_author[(series_key, author_key)]
        return series_map.get((book.series_key, author_key)), "local-series-is-book"

    author_series = [series for (s_key, a_key), series in series_map.items() if a_key == author_key]
    best = None
    best_score = 0.0
    for series in author_series:
        score = max(title_similarity(series.series, row.get("Book Series Name")), title_similarity(series.series, row.get("First Book Name")))
        if score > best_score:
            best_score = score
            best = series
    if best and best_score >= 0.78:
        return best, "local-fuzzy"
    return None, ""


def is_collection_title(title: object) -> bool:
    text = norm(title)
    patterns = [
        r"\bbox set\b",
        r"\bcollection\b",
        r"\bbooks?\s+\d+(?:-\d+)?\b",
        r"\bvol(?:ume)?\.?\b",
        r"\bomnibus\b",
    ]
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def summarize_local_series(local_series: Optional[LocalSeries]) -> Optional[Dict[str, object]]:
    if not local_series or not local_series.books:
        return None
    best_by_title: Dict[str, LocalBook] = {}
    for book in local_series.books:
        existing = best_by_title.get(book.title_key)
        if existing is None or score_local_book(book) > score_local_book(existing):
            best_by_title[book.title_key] = book

    main_books = [book for book in best_by_title.values() if not is_collection_title(book.title)]
    if not main_books:
        main_books = list(best_by_title.values())

    sorted_books = sorted(
        main_books,
        key=lambda book: (
            book.book_number is None or (book.book_number is not None and book.book_number <= 0),
            book.book_number or 0,
            book.publication_date or "",
            book.title.lower(),
        ),
    )
    declared_counts = [book.total_books_in_series for book in main_books if book.total_books_in_series]
    declared = max(declared_counts) if declared_counts else None
    observed = len({book.title_key for book in sorted_books})
    books_in_series = max(observed, declared or 0)
    pages = [book.pages for book in sorted_books if book.pages is not None]
    complete = bool(sorted_books) and books_in_series == observed
    total_pages = sum(pages) if complete and len(pages) == observed else None
    return {
        "series_name": local_series.series,
        "books_in_series": books_in_series,
        "first_book": sorted_books[0].title,
        "last_book": sorted_books[-1].title,
        "total_pages": total_pages,
        "complete": complete,
        "first_link": sorted_books[0].goodreads_link,
    }


def choose_search_result(
    results: List[Dict[str, str]],
    author_name: str,
    preferred_title: str,
    series_hint: str,
) -> Optional[Dict[str, str]]:
    author_key = norm_key(author_name)
    best = None
    best_score = 0.0
    for item in results:
        score = 0.0
        if norm_key(item["author"]) == author_key:
            score += 5.0
        else:
            score += 2.0 * title_similarity(item["author"], author_name)
        score += 3.0 * title_similarity(item["title_base"], preferred_title)
        score += 1.5 * title_similarity(item["title"], series_hint)
        if score > best_score:
            best_score = score
            best = item
    if best_score < 4.5:
        return None
    return best


def local_pages_for_series(
    series_data: Dict[str, object],
    local_series: Optional[LocalSeries],
    gr: GoodreadsClient,
) -> Tuple[Optional[int], str]:
    books = list(series_data.get("books", []))
    main_books = select_primary_books(series_data)
    page_total = 0
    matched = 0

    local_by_id: Dict[str, LocalBook] = {}
    local_by_title: Dict[str, LocalBook] = {}
    if local_series:
        for book in local_series.books:
            gid = goodreads_id(book.goodreads_link)
            if gid:
                local_by_id[gid] = book
            local_by_title[book.title_key] = book

    for item in main_books:
        gid = goodreads_id(str(item.get("url", "")))
        local_book = local_by_id.get(gid) if gid else None
        if local_book is None:
            local_book = local_by_title.get(norm_key(item.get("title")))
        pages = local_book.pages if local_book else None
        if pages is None:
            live_book = gr.fetch_book(str(item.get("url")))
            pages = safe_int(live_book.get("pages"))
        if pages is not None:
            page_total += pages
            matched += 1

    if matched == len(main_books) and main_books:
        return page_total, f"pages resolved for {matched}/{len(main_books)} primary works"
    return None, f"pages resolved for {matched}/{len(main_books)} primary works"


def select_primary_books(series_data: Dict[str, object]) -> List[Dict[str, object]]:
    books = list(series_data.get("books", []))
    if not books:
        return []
    primary_works = safe_int(series_data.get("primary_works")) or len(books)
    numbered_books = []
    for item in books:
        number = safe_float(item.get("number"))
        if number is None:
            continue
        if number >= 1 and float(number).is_integer():
            numbered_books.append(item)
    if len(numbered_books) >= primary_works:
        return numbered_books[:primary_works]
    return books[:primary_works]


def verify_outreach_row(
    row: Dict[str, object],
    author_result: Dict[str, object],
    series_map: Dict[Tuple[str, str], LocalSeries],
    by_book_author: Dict[Tuple[str, str], LocalBook],
    gr: GoodreadsClient,
) -> Dict[str, object]:
    notes: List[str] = []
    local_series, local_reason = match_local_series(row, series_map, by_book_author)
    local_summary = summarize_local_series(local_series)
    if local_reason:
        notes.append(local_reason)

    preferred_title = ""
    first_book = str(row.get("First Book Name") or "").strip()
    current_series = str(row.get("Book Series Name") or "").strip()

    if local_summary:
        preferred_title = str(local_summary["first_book"])
    elif first_book:
        preferred_title = first_book
    else:
        preferred_title = current_series

    book_data: Dict[str, object] = {}
    if local_summary and local_summary.get("first_link"):
        book_data = gr.fetch_book(str(local_summary["first_link"]))
        if norm_key(book_data.get("author")) != row["_author_key"]:
            book_data = {}

    if not book_data:
        queries = []
        if preferred_title:
            queries.append(f'"{preferred_title}" "{row.get("Author Name")}"')
        if first_book and first_book.lower() != preferred_title.lower():
            queries.append(f'"{first_book}" "{row.get("Author Name")}"')
        if current_series and current_series.lower() not in {preferred_title.lower(), first_book.lower()}:
            queries.append(f'"{current_series}" "{row.get("Author Name")}"')
        for query in queries:
            results = gr.search_books(query)
            chosen = choose_search_result(results, str(row.get("Author Name")), preferred_title, current_series)
            if not chosen:
                continue
            book_data = gr.fetch_book(chosen["url"])
            if norm_key(book_data.get("author")) == row["_author_key"]:
                notes.append("goodreads-search")
                break
            book_data = {}

    verified_series_name = ""
    verified_series_url = ""
    verified_books = None
    verified_type = ""
    verified_first = ""
    verified_last = ""
    verified_total_pages = None
    verified_length = None
    primary_works = None
    total_works = None

    if local_summary and local_summary.get("complete"):
        verified_series_name = str(local_summary["series_name"])
        verified_books = safe_int(local_summary.get("books_in_series"))
        verified_first = str(local_summary["first_book"])
        verified_last = str(local_summary["last_book"])
        verified_total_pages = safe_int(local_summary.get("total_pages"))
        if verified_total_pages is not None:
            verified_length = round(verified_total_pages * 300 / 10000, 2)
        primary_works = verified_books
        total_works = verified_books
        if verified_books and verified_books <= 1:
            verified_type = "Standalone"
        elif verified_books and verified_books <= 3:
            verified_type = "Short Series"
        elif verified_books and verified_books <= 5:
            verified_type = "Series"
        elif verified_books:
            verified_type = "Long Series"
        notes.append("local-goodreads-catalog")

    if book_data and book_data.get("series_url"):
        verified_series_url = str(book_data.get("series_url") or "")
        series_name_from_book = clean_series_label(book_data.get("series_name"))
        if series_name_from_book:
            verified_series_name = series_name_from_book

        series_data = gr.fetch_series(str(book_data["series_url"]))
        primary_books = select_primary_books(series_data)
        if series_data:
            primary_works = safe_int(series_data.get("primary_works")) or len(primary_books)
            total_works = safe_int(series_data.get("total_works")) or len(series_data.get("books", []))
            verified_series_name = clean_series_label(series_data.get("series_name") or verified_series_name or current_series)
            verified_series_url = str(series_data.get("url") or verified_series_url)
            verified_books = primary_works
            if primary_books:
                verified_first = primary_books[0]["title"]
                verified_last = primary_books[-1]["title"]
            elif not verified_first:
                verified_first = str(book_data.get("title") or "")
                verified_last = verified_first
            pages_total, page_note = local_pages_for_series(series_data, local_series, gr)
            notes.append(page_note)
            if pages_total is not None:
                verified_total_pages = pages_total
                verified_length = round(pages_total * 300 / 10000, 2)
            if verified_books <= 1:
                verified_type = "Standalone"
            elif verified_books <= 3:
                verified_type = "Short Series"
            elif verified_books <= 5:
                verified_type = "Series"
            else:
                verified_type = "Long Series"
            notes.append("goodreads-series")
        elif local_summary:
            notes.append("goodreads-book-page")
    elif not verified_series_name and book_data:
        verified_series_name = str(book_data.get("title") or current_series)
        verified_first = str(book_data.get("title") or preferred_title)
        verified_last = verified_first
        verified_books = 1
        primary_works = 1
        total_works = 1
        verified_type = "Standalone"
        pages = safe_int(book_data.get("pages"))
        if pages is not None:
            verified_total_pages = pages
            verified_length = round(pages * 300 / 10000, 2)
        notes.append("goodreads-standalone")

    has_series = bool(verified_series_name and verified_books)
    has_email = bool(author_result.get("Validated_Email"))
    has_agency = bool(author_result.get("Agency_Contact"))
    if has_series and has_email:
        flag = "GREEN"
    elif has_series and (has_agency or author_result.get("Validated_Website")):
        flag = "YELLOW"
    elif has_email:
        flag = "YELLOW"
    else:
        flag = "RED"

    return {
        **author_result,
        "Verified_Series_Name": verified_series_name,
        "Verified_Goodreads_Series_URL": verified_series_url,
        "Verified_Books_in_Series": verified_books,
        "Verified_Type": verified_type,
        "Verified_First_Book_Name": verified_first,
        "Verified_Last_Book_Name": verified_last,
        "Verified_Total_Pages": verified_total_pages,
        "Verified_Length_of_Adaption_in_Hours": verified_length,
        "Series_Primary_Works": primary_works,
        "Series_Total_Works": total_works,
        "Data_Quality_Flag": flag,
        "Verification_Notes": "; ".join(dict.fromkeys([n for n in notes if n])),
    }


def author_contact_source_text(row: pd.Series) -> str:
    if not row.get("Author_Checked"):
        return row.get("Contact Source", "") or ""
    if row.get("Validated_Email"):
        return f"Validated ({row.get('Email_Source_URL', '')})".strip()
    if row.get("Agency_Contact"):
        return f"Agency/Contact ({row.get('Agency_Source', '')})".strip()
    if row.get("Validated_Website"):
        return f"Website only ({row.get('Validated_Website', '')})".strip()
    return ""


def apply_author_updates(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    checked = df.get("Author_Checked", pd.Series(False, index=df.index)).astype("boolean").fillna(False)
    if "Validated_Email" in df.columns:
        df.loc[checked, "Email"] = df.loc[checked, "Validated_Email"].fillna("")
    if "Validated_Website" in df.columns:
        mask = checked & df["Validated_Website"].fillna("").astype(str).str.strip().ne("")
        df.loc[mask, "Website"] = df.loc[mask, "Validated_Website"]
    if "Agency_Contact" in df.columns and "Literary Agent" in df.columns:
        mask = checked & df["Agency_Contact"].fillna("").astype(str).str.strip().ne("")
        df.loc[mask, "Literary Agent"] = df.loc[mask, "Agency_Contact"]
    if "Contact Source" in df.columns and "Email_Source_URL" in df.columns:
        df["Contact Source"] = df.apply(author_contact_source_text, axis=1)
    return df


def apply_series_updates(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    updates = {
        "Book Series Name": "Verified_Series_Name",
        "Type": "Verified_Type",
        "Books in Series": "Verified_Books_in_Series",
        "Total Pages": "Verified_Total_Pages",
        "Length of Adaption in Hours": "Verified_Length_of_Adaption_in_Hours",
        "First Book Name": "Verified_First_Book_Name",
        "Last Book Name": "Verified_Last_Book_Name",
    }
    for target, source in updates.items():
        if source in df.columns:
            mask = df[source].notna() & (df[source].astype(str).str.strip() != "")
            df.loc[mask, target] = df.loc[mask, source]
    return df


def apply_author_results_to_master(master_rows: List[Dict[str, object]], author_lookup: Dict[str, Dict[str, object]]) -> pd.DataFrame:
    records = []
    for row in master_rows:
        author_result = author_lookup.get(row["_author_key"], {})
        record = dict(row)
        record.update(author_result)
        records.append(record)
    df = pd.DataFrame(records)
    checked = df.get("Author_Checked", pd.Series(False, index=df.index)).astype("boolean").fillna(False)
    if "Validated_Email" in df.columns:
        df.loc[checked, "Email"] = df.loc[checked, "Validated_Email"].fillna("")
    if "Validated_Website" in df.columns:
        mask = checked & df["Validated_Website"].fillna("").astype(str).str.strip().ne("")
        df.loc[mask, "Website"] = df.loc[mask, "Validated_Website"]
    if "Agency_Contact" in df.columns:
        mask = checked & df["Agency_Contact"].fillna("").astype(str).str.strip().ne("")
        df.loc[mask, "Literary Agent"] = df.loc[mask, "Agency_Contact"]
    if "Email_Source_URL" in df.columns and "Contact Source" in df.columns:
        def source_text(row: pd.Series) -> str:
            if not row.get("Author_Checked"):
                return row.get("Contact Source", "") or ""
            if row.get("Validated_Email"):
                return f"Validated ({row.get('Email_Source_URL', '')})".strip()
            if row.get("Agency_Contact"):
                return f"Agency/Contact ({row.get('Agency_Source', '')})".strip()
            if row.get("Validated_Website"):
                return f"Website only ({row.get('Validated_Website', '')})".strip()
            return ""
        df["Contact Source"] = df.apply(source_text, axis=1)
    return df


def write_csv_outputs(outreach_df: pd.DataFrame, master_df: pd.DataFrame, author_df: pd.DataFrame) -> Tuple[Path, Path, Path]:
    ensure_output_dirs()
    outreach_csv = EXPORTS_DIR / "ice_hockey_outreach_verified.csv"
    master_csv = EXPORTS_DIR / "ice_hockey_master_contacts_verified.csv"
    author_csv = EXPORTS_DIR / "ice_hockey_author_contacts_verified.csv"
    outreach_df.to_csv(outreach_csv, index=False)
    master_df.to_csv(master_csv, index=False)
    author_df.to_csv(author_csv, index=False)
    return outreach_csv, master_csv, author_csv


def ensure_output_dirs() -> None:
    for path in [SOURCE_DIR, REPORTS_DIR, VERIFIED_DIR, EXPORTS_DIR, PROGRESS_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def ensure_progress_dir() -> None:
    ensure_output_dirs()


def write_partial_author_output(author_df: pd.DataFrame) -> None:
    ensure_progress_dir()
    author_df.sort_values("Author Name").to_csv(CHECKPOINT_AUTHOR_CSV, index=False)


def write_partial_outreach_output(outreach_df: pd.DataFrame) -> None:
    ensure_progress_dir()
    outreach_df.sort_values("_excel_row").to_csv(CHECKPOINT_OUTREACH_CSV, index=False)


def write_partial_master_output(master_df: pd.DataFrame) -> None:
    ensure_progress_dir()
    master_df.sort_values("_excel_row").to_csv(CHECKPOINT_MASTER_CSV, index=False)


def log_progress(stage: str, current: int, total: int) -> None:
    print(f"[progress] {stage}: {current}/{total}", flush=True)


def update_sheet_from_dataframe(ws, df: pd.DataFrame, header_row: int, start_row: int, columns_to_update: List[str]) -> None:
    headers = [cell.value for cell in ws[header_row]]
    header_index = {value: idx + 1 for idx, value in enumerate(headers) if value}
    last_col = len(headers)
    for col in VERIFICATION_COLUMNS:
        if col not in header_index:
            last_col += 1
            ws.cell(row=header_row, column=last_col, value=col)
            header_index[col] = last_col
    if header_row == 2:
        ws.cell(row=1, column=min(header_index[col] for col in VERIFICATION_COLUMNS), value="Verification")

    row_lookup = {int(row["_excel_row"]): row for _, row in df.iterrows()}
    for excel_row, row in row_lookup.items():
        for col in columns_to_update + VERIFICATION_COLUMNS:
            if col not in header_index:
                continue
            value = row.get(col)
            if pd.isna(value):
                value = ""
            ws.cell(row=excel_row, column=header_index[col], value=value)


def write_workbook_copy(outreach_df: pd.DataFrame, master_df: pd.DataFrame) -> Path:
    ensure_output_dirs()
    output_path = VERIFIED_DIR / "Final self-pub scored.ice-hockey-verified.xlsx"
    shutil.copy2(WORKBOOK_PATH, output_path)
    wb = load_workbook(output_path)
    ws_out = wb[OUTREACH_SHEET]
    update_sheet_from_dataframe(
        ws_out,
        outreach_df,
        header_row=2,
        start_row=3,
        columns_to_update=[
            "Book Series Name",
            "Type",
            "Books in Series",
            "Total Pages",
            "Length of Adaption in Hours",
            "First Book Name",
            "Last Book Name",
            "Email",
            "Website",
        ],
    )

    ws_master = wb[MASTER_SHEET]
    update_sheet_from_dataframe(
        ws_master,
        master_df,
        header_row=1,
        start_row=2,
        columns_to_update=[
            "Book Series Name",
            "Type",
            "Books in Series",
            "Total Pages",
            "Length of Adaption in Hours",
            "First Book Name",
            "Last Book Name",
            "Email",
            "Website",
            "Literary Agent",
            "Contact Source",
        ],
    )
    wb.save(output_path)
    return output_path


def build_author_results(
    authors: Dict[str, AuthorCandidate],
    http: CachedHttp,
    limit: Optional[int],
    checkpoint_every: int = 25,
) -> pd.DataFrame:
    items = list(authors.items())
    if limit is not None:
        items = items[:limit]

    results: List[Dict[str, object]] = []
    total = len(items)
    with ThreadPoolExecutor(max_workers=12) as pool:
        futures = {
            pool.submit(validate_author_contact, http, candidate): key
            for key, candidate in items
        }
        for index, future in enumerate(as_completed(futures), start=1):
            results.append(future.result())
            if index == total or index % checkpoint_every == 0:
                log_progress("authors", index, total)
                write_partial_author_output(pd.DataFrame(results))
    return pd.DataFrame(results)


def build_verified_sheet_results(
    sheet_rows: List[Dict[str, object]],
    author_lookup: Dict[str, Dict[str, object]],
    series_map: Dict[Tuple[str, str], LocalSeries],
    by_book_author: Dict[Tuple[str, str], LocalBook],
    gr: GoodreadsClient,
    limit: Optional[int],
    stage_name: str,
    partial_writer,
    checkpoint_every: int = 25,
) -> pd.DataFrame:
    rows = sheet_rows[:limit] if limit is not None else sheet_rows
    records: List[Dict[str, object]] = []
    total = len(rows)
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {
            pool.submit(
                verify_outreach_row,
                row,
                author_lookup.get(row["_author_key"], {}),
                series_map,
                by_book_author,
                gr,
            ): row
            for row in rows
        }
        for index, future in enumerate(as_completed(futures), start=1):
            row = futures[future]
            result = dict(row)
            result.update(future.result())
            records.append(result)
            if index == total or index % checkpoint_every == 0:
                log_progress(stage_name, index, total)
                partial_df = pd.DataFrame(records)
                partial_df = apply_author_updates(partial_df)
                partial_df = apply_series_updates(partial_df)
                partial_writer(partial_df)
    df = pd.DataFrame(records)
    df = apply_author_updates(df)
    df = apply_series_updates(df)
    return df.sort_values("_excel_row")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit-authors", type=int, default=None)
    parser.add_argument("--limit-rows", type=int, default=None)
    parser.add_argument("--limit-master-rows", type=int, default=None)
    parser.add_argument("--skip-workbook", action="store_true")
    parser.add_argument("--checkpoint-every", type=int, default=25)
    args = parser.parse_args()

    outreach_rows = load_sheet_rows(OUTREACH_SHEET)
    master_rows = load_sheet_rows(MASTER_SHEET)
    all_rows = outreach_rows + master_rows

    http = CachedHttp()
    gr = GoodreadsClient(http)
    series_map, by_book_author = load_local_books()
    author_candidates = build_author_candidates(all_rows)

    print(
        json.dumps(
            {
                "stage": "start",
                "outreach_rows": len(outreach_rows),
                "master_rows": len(master_rows),
                "author_candidates": len(author_candidates),
            }
        ),
        flush=True,
    )

    author_df = build_author_results(author_candidates, http, args.limit_authors, args.checkpoint_every)
    author_lookup = {norm_key(row["Author Name"]): row for row in author_df.to_dict("records")}
    write_partial_author_output(author_df)
    master_author_df = apply_author_results_to_master(master_rows, author_lookup).sort_values("_excel_row")
    write_partial_master_output(master_author_df)
    print(
        json.dumps(
            {
                "stage": "authors_complete",
                "authors_total": len(author_df),
                "authors_with_validated_email": int(author_df["Validated_Email"].astype(str).str.strip().ne("").sum()),
            }
        ),
        flush=True,
    )

    outreach_df = build_verified_sheet_results(
        outreach_rows,
        author_lookup,
        series_map,
        by_book_author,
        gr,
        args.limit_rows,
        "outreach",
        write_partial_outreach_output,
        args.checkpoint_every,
    )
    write_partial_outreach_output(outreach_df)
    master_df = build_verified_sheet_results(
        master_rows,
        author_lookup,
        series_map,
        by_book_author,
        gr,
        args.limit_master_rows,
        "master",
        write_partial_master_output,
        args.checkpoint_every,
    )
    write_partial_master_output(master_df)
    print(
        json.dumps(
            {
                "stage": "outreach_complete",
                "outreach_rows_total": len(outreach_df),
                "outreach_rows_series_verified": int(outreach_df["Verified_Series_Name"].astype(str).str.strip().ne("").sum()),
            }
        ),
        flush=True,
    )
    print(
        json.dumps(
            {
                "stage": "master_complete",
                "master_rows_total": len(master_df),
                "master_rows_series_verified": int(master_df["Verified_Series_Name"].astype(str).str.strip().ne("").sum()),
            }
        ),
        flush=True,
    )

    outreach_csv, master_csv, author_csv = write_csv_outputs(outreach_df, master_df, author_df)
    workbook_path = None if args.skip_workbook else write_workbook_copy(outreach_df, master_df)

    summary = {
        "authors_total": len(author_df),
        "authors_with_validated_email": int(author_df["Validated_Email"].astype(str).str.strip().ne("").sum()),
        "outreach_rows_total": len(outreach_df),
        "outreach_rows_series_verified": int(outreach_df["Verified_Series_Name"].astype(str).str.strip().ne("").sum()),
        "master_rows_total": len(master_df),
        "master_rows_series_verified": int(master_df["Verified_Series_Name"].astype(str).str.strip().ne("").sum()),
        "outreach_rows_green": int((outreach_df["Data_Quality_Flag"] == "GREEN").sum()),
        "outreach_rows_yellow": int((outreach_df["Data_Quality_Flag"] == "YELLOW").sum()),
        "outreach_rows_red": int((outreach_df["Data_Quality_Flag"] == "RED").sum()),
        "master_rows_green": int((master_df["Data_Quality_Flag"] == "GREEN").sum()),
        "master_rows_yellow": int((master_df["Data_Quality_Flag"] == "YELLOW").sum()),
        "master_rows_red": int((master_df["Data_Quality_Flag"] == "RED").sum()),
        "outputs": {
            "outreach_csv": str(outreach_csv),
            "master_csv": str(master_csv),
            "author_csv": str(author_csv),
            "workbook": str(workbook_path) if workbook_path else "",
        },
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
