#!/usr/bin/env python3
"""
EXPAND GENRE DISCOVERY — Gemini-powered series discovery for 9 subgenres.
Discovers 100+ new self-pub book series per subgenre, validates via Goodreads,
filters out fantasy/romantasy, deduplicates, and enriches to match the xlsx format.

Usage:
    python execution/expand_genre_discovery.py --all
    python execution/expand_genre_discovery.py --genre "Dark Romance"
    python execution/expand_genre_discovery.py --phase discover
    python execution/expand_genre_discovery.py --phase validate
    python execution/expand_genre_discovery.py --phase enrich
"""

import asyncio
import argparse
import json
import os
import random
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin

import pandas as pd
import numpy as np
from loguru import logger
from playwright.async_api import async_playwright
from dotenv import load_dotenv

# Load env
PROJECT_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
load_dotenv(os.path.join(PROJECT_ROOT, '.env'))
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# Gemini setup
try:
    import google.generativeai as genai
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model = genai.GenerativeModel('gemini-2.5-flash')
    else:
        gemini_model = None
        logger.error("No GEMINI_API_KEY — cannot run discovery")
except ImportError:
    gemini_model = None
    logger.error("google-generativeai not installed")

# ============================================================================
# CONFIGURATION
# ============================================================================

OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data", "genre_expansion")
WORKER_COUNT = 4
SAVE_INTERVAL = 10
SLEEP_MIN = 2
SLEEP_MAX = 5
HEADLESS = True
WORDS_PER_PAGE = 300
WORDS_PER_HOUR = 10000

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2.1 Safari/605.1.15",
]

TRADITIONAL_PUBLISHERS = [
    'penguin', 'harpercollins', 'simon & schuster', 'hachette', 'macmillan',
    'random house', 'scholastic', 'wiley', 'bloomsbury', 'sourcebooks',
    'berkley', 'avon', 'ballantine', 'bantam', 'dell', 'tor', 'forge',
    'st. martin', 'entangled', 'montlake', 'forever', 'grand central',
    'kensington', 'zebra', 'dafina', 'carina', 'harlequin', 'mira',
    'mills & boon', 'william morrow', 'putnam', 'dutton', 'atria',
    'gallery', 'scribner', 'vintage', 'anchor', 'crown', 'knopf',
    'doubleday', 'little brown', 'orbit', 'ace', 'roc', 'daw',
    'del rey', 'piatkus', 'hodder', 'headline', 'hq',
]

SELF_PUB_KEYWORDS = [
    'independently published', 'self-published', 'createspace', 'draft2digital',
    'smashwords', 'kindle direct', 'kdp', 'lulu', 'blurb', 'author house',
    'authorhouse', 'xlibris', 'iuniverse', 'trafford', 'balboa press',
]

FANTASY_EXCLUSION_KEYWORDS = [
    'fantasy', 'romantasy', 'fae', 'faerie', 'dragon', 'vampire', 'werewolf',
    'shifter', 'witch', 'wizard', 'magic', 'supernatural', 'paranormal',
    'demon', 'angel', 'realm', 'kingdom', 'throne', 'elves', 'orc',
    'sorcerer', 'enchanted', 'spellbound', 'mythical', 'dystopian',
    'post-apocalyptic', 'sci-fi', 'space', 'alien', 'cyberpunk', 'steampunk',
    'shapeshifter', 'lycan', 'immortal', 'necromancer', 'warlock',
    'young adult', ' ya ',
]

# The 9 subgenres
SUBGENRES = [
    "Ice Hockey & Sports Romance",
    "Dark Romance",
    "Forbidden Romance",
    "Mafia Drama/Romance",
    "Military Drama/Romance",
    "Political Drama/Romance",
    "Historical Romance/Fiction",
    "Small Town Drama/Romance",
    "Christian Drama/Romance",
    "Romantic Suspense/Psychological Thriller",
]

# Gemini discovery prompt templates — 6 variations per subgenre for breadth
DISCOVERY_PROMPTS = [
    """List 30 popular self-published {subgenre} book series with 3 or more books each. Focus on Kindle Unlimited and indie authors who are NOT with major traditional publishers. For each series provide: series_name, author_name, estimated_number_of_books.

CRITICAL: Do NOT include any fantasy, romantasy, paranormal, sci-fi, or YA titles. Only contemporary/realistic drama and romance.
CRITICAL: Do NOT include authors published by Penguin, HarperCollins, Simon & Schuster, Hachette, Macmillan, or Random House.

Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",

    """List 30 lesser-known but highly-rated self-published {subgenre} book series on Amazon Kindle with at least 3 books. These should be hidden gems that avid romance/drama readers love. Only contemporary drama/romance — no fantasy, paranormal, or sci-fi elements.

For each: series_name, author_name, estimated_number_of_books.
Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",

    """List 30 trending self-published {subgenre} series from 2022-2025 with 3+ books. Only contemporary drama/romance that are popular on BookTok, Bookstagram, or Kindle Unlimited. No fantasy, romantasy, paranormal, or sci-fi.

For each: series_name, author_name, estimated_number_of_books.
Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",

    """List 30 self-published {subgenre} series that have high Goodreads ratings (3.8+) and large followings (1000+ ratings). Only realistic fiction/romance — no paranormal, fantasy, or sci-fi. Focus on indie/self-pub authors only.

For each: series_name, author_name, estimated_number_of_books.
Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",

    """What are the most successful self-published {subgenre} series on Amazon Kindle that have 5+ books and would make great audio adaptations? List 30. Only drama/romance genre — no fantasy elements. Focus on series with strong narrative arcs and cliffhangers between books.

For each: series_name, author_name, estimated_number_of_books.
Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",

    """List 30 self-published {subgenre} book series that are well-known in the romance/drama reading community. These should be series with dedicated fan bases, available on Kindle Unlimited. The series must have at least 3 books, and be contemporary/realistic — absolutely no fantasy, paranormal, romantasy, or sci-fi.

Include both well-established series and newer breakout hits from 2023-2025.
For each: series_name, author_name, estimated_number_of_books.
Return ONLY valid JSON array: [{{"series_name": "...", "author_name": "...", "estimated_books": N}}, ...]""",
]

# ============================================================================
# SCORING / COMMISSIONING (mirrors analysis.py logic)
# ============================================================================

WEIGHTS = {'volume': 0.30, 'quality_first': 0.15, 'quality_avg': 0.10, 'retention': 0.25, 'appeal': 0.20}
RANK_THRESHOLDS = {'P0': 90, 'P1': 80, 'P2': 70, 'P3': 60, 'P4': 50, 'P5': 0}


def compute_commissioning_score(series_data):
    """Compute commissioning score for a validated series."""
    n_books = series_data.get('num_books', 0)
    first_rating = series_data.get('first_book_rating', 0) or 0
    avg_rating = series_data.get('avg_rating', 0) or 0
    first_count = series_data.get('first_book_count', 0) or 0
    last_count = series_data.get('last_book_count', 0) or 0

    # Volume score
    if n_books == 1:
        norm_vol = 40
    elif n_books < 3:
        norm_vol = 60
    elif n_books <= 5:
        norm_vol = 85
    else:
        norm_vol = 100

    # Type label
    if n_books == 1:
        type_label = "Standalone"
    elif n_books < 3:
        type_label = "Short Series"
    elif n_books <= 5:
        type_label = "Series"
    else:
        type_label = "Long Series"

    # Quality scores (benefit of doubt: if 0, give 80)
    norm_q1 = min(first_rating / 5.0, 1.0) * 100 if first_rating > 0.1 else 80
    norm_qavg = min(avg_rating / 5.0, 1.0) * 100 if avg_rating > 0.1 else 80

    # Retention
    retention_raw = last_count / first_count if first_count > 0 else 0
    norm_retention = min(retention_raw / 0.5, 1.0) * 100 if retention_raw > 0.01 else 80

    # Appeal
    norm_appeal = min(first_count / 10000, 1.0) * 100 if first_count > 1 else 80

    base_score = (
        (norm_q1 * WEIGHTS['quality_first']) +
        (norm_qavg * WEIGHTS['quality_avg']) +
        (norm_retention * WEIGHTS['retention']) +
        (norm_appeal * WEIGHTS['appeal']) +
        (norm_vol * WEIGHTS['volume'])
    )

    # Self-pub bonus
    self_pub = series_data.get('self_pub_flag', 'Indie')
    if self_pub in ('Self-Pub', 'Indie'):
        base_score *= 1.20

    # Rank
    rank = 'P5'
    for r, thresh in RANK_THRESHOLDS.items():
        if base_score >= thresh:
            rank = r
            break

    return {
        'score': round(base_score, 1),
        'rank': rank,
        'type': type_label,
        'rationale': f"Score: {int(base_score)} (Vol:{int(norm_vol)} Q1:{int(norm_q1)} QAvg:{int(norm_qavg)} R:{int(norm_retention)} A:{int(norm_appeal)})",
    }


# ============================================================================
# HELPERS
# ============================================================================

def normalize_name(name):
    """Normalize a series/author name for dedup."""
    if not name:
        return ''
    return re.sub(r'[^a-z0-9\s]', '', str(name).lower()).strip()


def determine_self_pub(publisher):
    """Determine self-pub status from publisher name."""
    if not publisher or pd.isna(publisher) or str(publisher).strip() == '':
        return 'Indie'  # Default assumption for unknown
    p_lower = str(publisher).lower()
    if any(kw in p_lower for kw in SELF_PUB_KEYWORDS):
        return 'Self-Pub'
    if any(pub in p_lower for pub in TRADITIONAL_PUBLISHERS):
        return 'Big Pub'
    return 'Indie'


def is_fantasy_or_excluded(text):
    """Check if text contains fantasy/romantasy/sci-fi indicators."""
    if not text:
        return False
    text_lower = str(text).lower()
    return any(kw in text_lower for kw in FANTASY_EXCLUSION_KEYWORDS)


async def create_stealth_context(browser):
    """Create a stealth browser context."""
    ua = random.choice(USER_AGENTS)
    context = await browser.new_context(
        user_agent=ua,
        viewport={"width": 1920, "height": 1080},
        locale="en-US",
        timezone_id="America/New_York",
        extra_http_headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "DNT": "1",
        }
    )
    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        window.chrome = { runtime: {} };
    """)
    return context


async def safe_goto(page, url, timeout=45000, retries=3):
    """Network-resilient navigation."""
    for attempt in range(1, retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            await asyncio.sleep(random.uniform(0.5, 1.5))
            return True
        except Exception as e:
            logger.debug(f"  goto attempt {attempt}/{retries} failed: {e}")
            await asyncio.sleep(2 * attempt + random.uniform(1, 3))
    return False


# ============================================================================
# PHASE 1: GEMINI DISCOVERY
# ============================================================================

async def gemini_discover_series(subgenre):
    """Use Gemini to discover series for a subgenre using multiple prompt variations."""
    if not gemini_model:
        logger.error("Gemini model not available!")
        return []

    all_series = []
    seen_keys = set()

    for i, prompt_template in enumerate(DISCOVERY_PROMPTS):
        logger.info(f"  Gemini prompt {i+1}/{len(DISCOVERY_PROMPTS)} for {subgenre}...")

        prompt = prompt_template.format(subgenre=subgenre)

        for attempt in range(3):
            try:
                response = await asyncio.to_thread(gemini_model.generate_content, prompt)
                text = response.text.strip()

                # Extract JSON
                if '```json' in text:
                    text = text.split('```json')[1].split('```')[0]
                elif '```' in text:
                    text = text.split('```')[1].split('```')[0]

                series_list = json.loads(text)

                if not isinstance(series_list, list):
                    logger.warning(f"  Unexpected response format for prompt {i+1}")
                    break

                new_count = 0
                for s in series_list:
                    name = str(s.get('series_name', '')).strip()
                    author = str(s.get('author_name', '')).strip()
                    est_books = s.get('estimated_books', 0)

                    if not name or not author:
                        continue

                    # Quick fantasy filter on series name
                    if is_fantasy_or_excluded(name):
                        continue

                    # Dedup key
                    key = f"{normalize_name(name)}|||{normalize_name(author)}"
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)

                    all_series.append({
                        'series_name': name,
                        'author_name': author,
                        'estimated_books': int(est_books) if est_books else 3,
                        'source_prompt': i + 1,
                    })
                    new_count += 1

                logger.info(f"    Got {new_count} new unique series (total: {len(all_series)})")
                break  # Success, move to next prompt

            except json.JSONDecodeError as e:
                logger.warning(f"  JSON parse error on prompt {i+1}, attempt {attempt+1}: {e}")
                if attempt < 2:
                    await asyncio.sleep(2)
            except Exception as e:
                if "429" in str(e):
                    logger.warning(f"  Rate limit on prompt {i+1}, waiting 10s...")
                    await asyncio.sleep(10)
                else:
                    logger.error(f"  Gemini error on prompt {i+1}: {e}")
                    break

        # Small delay between prompts
        await asyncio.sleep(random.uniform(2, 4))

    logger.info(f"  Total discovered for {subgenre}: {len(all_series)} unique series")
    return all_series


# ============================================================================
# PHASE 2: GOODREADS VALIDATION
# ============================================================================

async def validate_series_on_goodreads(page, series_name, author_name):
    """Search Goodreads to validate a series and get metadata."""
    result = {
        'validated': False,
        'gr_series_name': '',
        'gr_series_url': '',
        'gr_author': '',
        'num_books': 0,
        'books': [],  # list of {name, number, rating, count, pages, pub_date, gr_link}
        'publisher': '',
        'genres': '',
        'description': '',
    }

    try:
        # Search for the series on Goodreads
        query = f"{series_name} {author_name}".strip()
        search_url = f"https://www.goodreads.com/search?q={query.replace(' ', '+')}"

        if not await safe_goto(page, search_url):
            return result

        await asyncio.sleep(random.uniform(1, 2))

        # Find the first book result
        items = await page.query_selector_all("tr[itemtype='http://schema.org/Book']")
        if not items:
            return result

        first_item = items[0]
        title_el = await first_item.query_selector("a.bookTitle")
        if not title_el:
            return result

        href = await title_el.get_attribute("href")
        book_url = f"https://www.goodreads.com{href}"

        # Navigate to the book page
        if not await safe_goto(page, book_url):
            return result

        await asyncio.sleep(random.uniform(1, 2))

        # Check for series link
        series_el = await page.query_selector("h3.Text__italic a, div.BookPageTitleSection a[href*='/series/']")
        if not series_el:
            # Not part of a series — still capture single-book data
            return result

        series_text = (await series_el.text_content()).strip()
        series_href = await series_el.get_attribute("href")

        # Parse series name from "Series Name #1"
        m = re.match(r'(.+?)(?:\s*#[\d.]+)?$', series_text)
        if m:
            result['gr_series_name'] = m.group(1).strip()

        if series_href:
            result['gr_series_url'] = series_href if series_href.startswith('http') else f"https://www.goodreads.com{series_href}"

        # Get genres from book page (for fantasy filtering)
        try:
            genre_els = await page.query_selector_all("span.BookPageMetadataSection__genreButton a .Button__labelItem")
            genres = []
            for g in genre_els[:8]:
                genres.append((await g.text_content()).strip())
            result['genres'] = ', '.join(genres)
        except:
            pass

        # Get description
        try:
            desc_el = await page.query_selector("div.BookPageMetadataSection__description span.Formatted")
            if desc_el:
                result['description'] = (await desc_el.text_content()).strip()[:1000]
        except:
            pass

        # Get publisher
        try:
            pub_details = await page.query_selector("div.FeaturedDetails p[data-testid='publicationInfo']")
            if pub_details:
                pub_text = (await pub_details.text_content()).strip()
                pm = re.search(r'(?:Published|published)\s+.*?\s+by\s+(.+?)(?:\s*$)', pub_text)
                if pm:
                    result['publisher'] = pm.group(1).strip()
        except:
            pass

        # Navigate to series page to get all books
        if result['gr_series_url']:
            if not await safe_goto(page, result['gr_series_url']):
                return result

            await asyncio.sleep(random.uniform(1.5, 3))

            # Extract books from series page
            book_els = await page.query_selector_all("div.listWithDividers__item, div[itemtype='http://schema.org/Book']")
            if not book_els:
                # Try alternate selector
                book_els = await page.query_selector_all("div.responsiveBook")

            books = []
            for bel in book_els:
                try:
                    # Book name
                    name_el = await bel.query_selector("a span[itemprop='name'], a.gr-h3, span[role='heading']")
                    if not name_el:
                        name_el = await bel.query_selector("a span, h3 a")
                    if not name_el:
                        continue
                    book_name = (await name_el.text_content()).strip()
                    if not book_name:
                        continue

                    # Book number
                    num_el = await bel.query_selector("span:has-text('Book'), h3")
                    book_num = ''
                    if num_el:
                        num_text = (await num_el.text_content()).strip()
                        nm = re.search(r'Book\s+([\d.]+)', num_text)
                        if nm:
                            book_num = nm.group(1)

                    # Rating
                    rating_el = await bel.query_selector("span.minirating, span.br-currentRating")
                    rating = 0
                    count = 0
                    if rating_el:
                        rating_text = (await rating_el.text_content()).strip()
                        rm = re.search(r'([\d.]+)\s+avg', rating_text)
                        if rm:
                            rating = float(rm.group(1))
                        cm = re.search(r'([\d,]+)\s+rating', rating_text)
                        if cm:
                            count = int(cm.group(1).replace(',', ''))

                    # GR Link
                    link_el = await bel.query_selector("a[href*='/book/show/']")
                    gr_link = ''
                    if link_el:
                        lhref = await link_el.get_attribute("href")
                        gr_link = lhref if lhref and lhref.startswith('http') else f"https://www.goodreads.com{lhref}" if lhref else ''

                    # Only include numbered main books (skip companion novels, etc.)
                    if book_num or len(books) < 20:
                        books.append({
                            'name': book_name,
                            'number': book_num,
                            'rating': rating,
                            'count': count,
                            'gr_link': gr_link,
                        })
                except Exception:
                    continue

            # Filter to likely main series books (numbered or first N)
            numbered = [b for b in books if b['number']]
            result['books'] = numbered if numbered else books[:20]
            result['num_books'] = len(result['books'])
            result['validated'] = result['num_books'] >= 1

    except Exception as e:
        logger.debug(f"  GR validation error for {series_name}: {e}")

    return result


# ============================================================================
# PHASE 2 WORKER
# ============================================================================

async def validation_worker(worker_id, browser, queue, results_list, lock, save_path):
    """Worker that validates series via Goodreads."""
    context = await create_stealth_context(browser)
    page = await context.new_page()
    processed = 0

    try:
        while True:
            try:
                idx, series = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            logger.info(f"  Worker {worker_id} [{idx}]: Validating {series['series_name']} by {series['author_name']}")

            try:
                gr_data = await validate_series_on_goodreads(
                    page, series['series_name'], series['author_name']
                )

                series['gr_validated'] = gr_data['validated']
                series['gr_series_name'] = gr_data.get('gr_series_name', series['series_name'])
                series['gr_series_url'] = gr_data.get('gr_series_url', '')
                series['num_books'] = gr_data.get('num_books', 0)
                series['books'] = gr_data.get('books', [])
                series['publisher'] = gr_data.get('publisher', '')
                series['genres'] = gr_data.get('genres', '')
                series['description'] = gr_data.get('description', '')
                series['self_pub_flag'] = determine_self_pub(gr_data.get('publisher', ''))

                async with lock:
                    results_list.append(series)
                    processed += 1

                    if processed % SAVE_INTERVAL == 0:
                        _save_checkpoint(results_list, save_path)
                        logger.info(f"  Worker {worker_id}: Checkpoint ({processed} series)")

                status = 'Y' if gr_data['validated'] else 'X'
                logger.info(
                    f"  Worker {worker_id}: [{status}] {series['series_name']} -> "
                    f"{gr_data['num_books']} books | "
                    f"Publisher: {gr_data.get('publisher', 'N/A')[:20]}"
                )

            except Exception as e:
                logger.error(f"  Worker {worker_id}: Error on {series['series_name']}: {e}")

            # Rotate context
            if processed % 15 == 0 and processed > 0:
                await page.close()
                await context.close()
                context = await create_stealth_context(browser)
                page = await context.new_page()

            await asyncio.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    finally:
        await page.close()
        await context.close()

    logger.info(f"  Worker {worker_id}: Done — validated {processed} series")


def _save_checkpoint(results_list, path):
    """Save checkpoint of validated results."""
    if not results_list:
        return
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    # Save a simplified version (without nested books list)
    rows = []
    for s in results_list:
        rows.append({
            'series_name': s.get('gr_series_name', s.get('series_name', '')),
            'author_name': s.get('author_name', ''),
            'num_books': s.get('num_books', 0),
            'gr_series_url': s.get('gr_series_url', ''),
            'publisher': s.get('publisher', ''),
            'self_pub_flag': s.get('self_pub_flag', ''),
            'genres': s.get('genres', ''),
            'description': s.get('description', ''),
            'gr_validated': s.get('gr_validated', False),
            'books_json': json.dumps(s.get('books', [])),
        })
    pd.DataFrame(rows).to_csv(path, index=False)


# ============================================================================
# PHASE 3: FILTERING
# ============================================================================

def load_existing_series(xlsx_path):
    """Load all existing series from the xlsx for deduplication."""
    existing = set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
            if not rows:
                continue
            header = rows[0]
            # Find series name and author columns
            series_col = None
            author_col = None
            for i, h in enumerate(header):
                if h and 'series name' in str(h).lower():
                    series_col = i
                elif h and 'author name' in str(h).lower():
                    author_col = i
            if series_col is not None:
                for row in rows[1:]:
                    s = normalize_name(str(row[series_col] or ''))
                    a = normalize_name(str(row[author_col] or '')) if author_col is not None else ''
                    if s:
                        existing.add(f"{s}|||{a}")
                        existing.add(s)  # Also match by series name alone
        wb.close()
    except Exception as e:
        logger.warning(f"Could not load existing data for dedup: {e}")
    return existing


def filter_validated_series(validated_series, existing_keys):
    """Apply all filters to validated series."""
    filtered = []
    seen = set()

    for s in validated_series:
        series_name = s.get('gr_series_name', s.get('series_name', ''))
        author = s.get('author_name', '')
        num_books = s.get('num_books', 0)
        genres = s.get('genres', '')
        description = s.get('description', '')
        publisher = s.get('publisher', '')
        self_pub = s.get('self_pub_flag', '')

        # Filter 1: Must have been validated
        if not s.get('gr_validated'):
            continue

        # Filter 2: Minimum 3 books
        if num_books < 3:
            continue

        # Filter 3: Not a big publisher
        if self_pub == 'Big Pub':
            continue

        # Filter 4: Not fantasy/romantasy/sci-fi
        if is_fantasy_or_excluded(genres) or is_fantasy_or_excluded(series_name) or is_fantasy_or_excluded(description):
            continue

        # Filter 5: Not already in existing data
        norm_series = normalize_name(series_name)
        norm_author = normalize_name(author)
        dedup_key = f"{norm_series}|||{norm_author}"
        if dedup_key in existing_keys or norm_series in existing_keys:
            continue

        # Filter 6: Cross-subgenre dedup
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        # Filter 7: Check book ratings — at least one book should have a decent rating
        books = s.get('books', [])
        ratings = [b.get('rating', 0) for b in books if b.get('rating', 0) > 0]
        if ratings and max(ratings) < 3.0:
            continue

        filtered.append(s)

    return filtered


# ============================================================================
# PHASE 4: ENRICHMENT (Gemini subjective + format to xlsx schema)
# ============================================================================

async def gemini_enrich_series(series_name, author_name, description, subgenre):
    """Use Gemini for subjective analysis and differentiator."""
    if not gemini_model:
        return {'subjective': '', 'differentiator': ''}

    try:
        prompt = f"""You are a book industry analyst specializing in {subgenre}. Analyze this series:
Series: "{series_name}" by {author_name}
Description: {description[:600] if description else "No description available"}

Provide:
1. SUBJECTIVE_ANALYSIS: 1-2 sentence market analysis of this series' appeal and audience.
2. DIFFERENTIATOR: What makes this series unique or stand out in the {subgenre} subgenre.

Return ONLY valid JSON:
{{"subjective": "...", "differentiator": "..."}}"""

        response = await asyncio.to_thread(gemini_model.generate_content, prompt)
        text = response.text.strip()
        if '```json' in text:
            text = text.split('```json')[1].split('```')[0]
        elif '```' in text:
            text = text.split('```')[1].split('```')[0]
        return json.loads(text)
    except Exception as e:
        if "429" in str(e):
            await asyncio.sleep(5)
        return {'subjective': '', 'differentiator': ''}


def format_to_xlsx_schema(series, subgenre, gemini_enrichment=None):
    """Format a validated+filtered series into the xlsx column schema."""
    books = series.get('books', [])
    if not books:
        return None

    # Sort by book number
    for b in books:
        try:
            b['_num'] = float(b.get('number', 0) or 0)
        except:
            b['_num'] = 999
    books.sort(key=lambda x: x['_num'])

    first = books[0]
    last = books[-1]
    n_books = len(books)

    # Find highest and lowest rated
    rated_books = [b for b in books if b.get('rating', 0) > 0]
    if rated_books:
        highest = max(rated_books, key=lambda x: (x['rating'], x.get('count', 0)))
        lowest = min(rated_books, key=lambda x: x['rating'])
    else:
        highest = first
        lowest = first

    # Ratings
    ratings = [b.get('rating', 0) for b in books if b.get('rating', 0) > 0]
    avg_rating = sum(ratings) / len(ratings) if ratings else 0

    # Pages estimate (avg 300 pages per book if unknown)
    avg_pages = 300
    total_pages = n_books * avg_pages
    adaptation_hours = (total_pages * WORDS_PER_PAGE) / WORDS_PER_HOUR

    # Books list string
    books_list = ', '.join(b.get('name', '') for b in books if b.get('name'))

    # Scoring
    score_data = {
        'num_books': n_books,
        'first_book_rating': first.get('rating', 0),
        'avg_rating': avg_rating,
        'first_book_count': first.get('count', 0),
        'last_book_count': last.get('count', 0),
        'self_pub_flag': series.get('self_pub_flag', 'Indie'),
    }
    scoring = compute_commissioning_score(score_data)

    # Derived flags
    first_count = first.get('count', 0)
    first_rating = first.get('rating', 0)
    lowest_rating = lowest.get('rating', 0) if lowest else 0

    adaptation_flag = 'Very High' if adaptation_hours > 50 else ('High' if adaptation_hours >= 40 else ('Medium' if adaptation_hours >= 30 else 'Low'))
    rating_flag = 'Very High' if first_rating > 4.2 else ('High' if first_rating >= 3.9 else ('Medium' if first_rating >= 3.5 else 'Low'))
    appeal_flag = 'Very High' if first_count > 50000 else ('High' if first_count >= 10000 else ('Medium' if first_count >= 5000 else 'Low'))
    lowest_flag = 'Very High' if lowest_rating > 4.0 else ('High' if lowest_rating >= 3.8 else ('Medium' if lowest_rating >= 3.5 else 'Low'))

    # Rating stability
    if ratings and len(ratings) > 1:
        stability = max(ratings) - min(ratings)
        stability_flag = 'Very High' if stability < 0.2 else ('High' if stability < 0.4 else ('Medium' if stability < 0.6 else 'Low'))
    else:
        stability_flag = 'Medium'

    # Era
    # We don't have pub dates from this flow, so estimate based on Gemini's knowledge
    series_era = 'After 2020'  # Default for self-pub discovery

    enrichment = gemini_enrichment or {}

    return {
        'Book Series Name': series.get('gr_series_name', series.get('series_name', '')),
        'Author Name': series.get('author_name', ''),
        'Type': scoring['type'],
        'Books_In_Series_List': books_list,
        'Universe Type': '',
        'Universe Reasoning': '',
        'Verfied Flag': 'Yes' if series.get('gr_validated') else '',
        'Books in Series': n_books,
        'Total Pages': total_pages,
        'Length of Adaption in Hours': round(adaptation_hours, 1),
        'First Book Name': first.get('name', ''),
        'First Book Rating': first.get('rating', ''),
        'First Book Rating Count': first.get('count', ''),
        'Last Book Name': last.get('name', ''),
        'Last Book Rating': last.get('rating', ''),
        'Last Book Rating Count': last.get('count', ''),
        'Highest Rated Book Name': highest.get('name', ''),
        'Highest Rated Book Rating': highest.get('rating', ''),
        'Highest Rated Book Rating Count': highest.get('count', ''),
        'Lowest Rated Book Name': lowest.get('name', ''),
        'Lowest Rated Book Rating': lowest.get('rating', ''),
        'Lowest Rated Book Rating Count': lowest.get('count', ''),
        'Publisher Name': series.get('publisher', ''),
        'Self Pub Flag': series.get('self_pub_flag', 'Indie'),
        'Subjective Analysis': enrichment.get('subjective', ''),
        'Differentiator': enrichment.get('differentiator', ''),
        'Books_Featured_Rank_Validation': '',
        'Num_Books_Featured': 0,
        'First_Book_Pub_Year': '',
        'T100_Mapping': '',
        'Adaptation_Length_Flag': adaptation_flag,
        'First_Book_Rating_Flag': rating_flag,
        'Appeal Flag': appeal_flag,
        'Lowest_Book_Rating_Flag': lowest_flag,
        'Rating_Stability_Flag': stability_flag,
        'Series_Era': series_era,
        'Commissioning_Score': scoring['score'],
        'Commissioning_Rank': scoring['rank'],
        'Primary Subgenre': subgenre,
        'Rationale': scoring['rationale'],
        'Goodreads Series URL': series.get('gr_series_url', ''),
    }


# ============================================================================
# FULL PIPELINE PER SUBGENRE
# ============================================================================

async def run_subgenre_pipeline(subgenre, existing_keys, phase=None):
    """Run the full pipeline for a single subgenre."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = re.sub(r'[/\\:*?"<>| &]', '_', subgenre)

    discovery_path = os.path.join(OUTPUT_DIR, f"{safe_name}_discovered.json")
    validated_path = os.path.join(OUTPUT_DIR, f"{safe_name}_validated.csv")
    final_path = os.path.join(OUTPUT_DIR, f"{safe_name}_expanded.csv")

    # ---- Phase 1: Discover ----
    if phase is None or phase == 'discover':
        logger.info(f"\n{'='*60}")
        logger.info(f"PHASE 1: GEMINI DISCOVERY — {subgenre}")
        logger.info(f"{'='*60}")

        # Check for existing discovery
        if os.path.exists(discovery_path) and phase != 'discover':
            logger.info(f"  Loading existing discovery from {discovery_path}")
            with open(discovery_path) as f:
                discovered = json.load(f)
        else:
            discovered = await gemini_discover_series(subgenre)
            with open(discovery_path, 'w') as f:
                json.dump(discovered, f, indent=2)
            logger.success(f"  Saved {len(discovered)} discovered series to {discovery_path}")

        if phase == 'discover':
            return

    # Load discovery if we skipped phase 1
    if not os.path.exists(discovery_path):
        logger.error(f"No discovery file found at {discovery_path} — run --phase discover first")
        return
    with open(discovery_path) as f:
        discovered = json.load(f)

    # ---- Phase 2: Validate ----
    if phase is None or phase == 'validate':
        logger.info(f"\n{'='*60}")
        logger.info(f"PHASE 2: GOODREADS VALIDATION — {subgenre} ({len(discovered)} series)")
        logger.info(f"{'='*60}")

        # Check for already-validated
        already_validated = set()
        existing_validated = []
        if os.path.exists(validated_path):
            vdf = pd.read_csv(validated_path)
            already_validated = set(vdf['series_name'].str.lower().str.strip())
            # Reconstruct validated list from CSV
            for _, row in vdf.iterrows():
                entry = {
                    'series_name': row.get('series_name', ''),
                    'author_name': row.get('author_name', ''),
                    'gr_series_name': row.get('series_name', ''),
                    'gr_series_url': row.get('gr_series_url', ''),
                    'num_books': int(row.get('num_books', 0)),
                    'publisher': row.get('publisher', ''),
                    'self_pub_flag': row.get('self_pub_flag', ''),
                    'genres': row.get('genres', ''),
                    'description': row.get('description', ''),
                    'gr_validated': bool(row.get('gr_validated', False)),
                    'books': json.loads(row.get('books_json', '[]')),
                }
                existing_validated.append(entry)
            logger.info(f"  Resuming: {len(already_validated)} already validated")

        to_validate = [
            s for s in discovered
            if normalize_name(s['series_name']) not in already_validated
        ]

        if to_validate:
            queue = asyncio.Queue()
            for i, s in enumerate(to_validate):
                queue.put_nowait((i, s))

            validated_results = list(existing_validated)
            lock = asyncio.Lock()

            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=HEADLESS)
                workers = [
                    validation_worker(i, browser, queue, validated_results, lock, validated_path)
                    for i in range(min(WORKER_COUNT, len(to_validate)))
                ]
                await asyncio.gather(*workers)
                await browser.close()

            _save_checkpoint(validated_results, validated_path)
        else:
            validated_results = existing_validated
            logger.info("  All series already validated")

        if phase == 'validate':
            return

    # Load validated data
    if not os.path.exists(validated_path):
        logger.error(f"No validated file found — run --phase validate first")
        return

    vdf = pd.read_csv(validated_path)
    validated_results = []
    for _, row in vdf.iterrows():
        validated_results.append({
            'series_name': row.get('series_name', ''),
            'author_name': row.get('author_name', ''),
            'gr_series_name': row.get('series_name', ''),
            'gr_series_url': row.get('gr_series_url', ''),
            'num_books': int(row.get('num_books', 0)),
            'publisher': row.get('publisher', ''),
            'self_pub_flag': row.get('self_pub_flag', ''),
            'genres': row.get('genres', ''),
            'description': row.get('description', ''),
            'gr_validated': bool(row.get('gr_validated', False)),
            'books': json.loads(row.get('books_json', '[]')),
        })

    # ---- Phase 3: Filter ----
    logger.info(f"\n{'='*60}")
    logger.info(f"PHASE 3: FILTERING — {subgenre}")
    logger.info(f"{'='*60}")

    filtered = filter_validated_series(validated_results, existing_keys)
    logger.info(f"  {len(validated_results)} validated -> {len(filtered)} after filtering")
    logger.info(f"  Removed: {len(validated_results) - len(filtered)} (big pub, fantasy, <3 books, duplicates, low ratings)")

    if not filtered:
        logger.warning(f"  No series passed filtering for {subgenre}")
        return

    # ---- Phase 4: Enrich + Format ----
    logger.info(f"\n{'='*60}")
    logger.info(f"PHASE 4: ENRICHMENT — {subgenre} ({len(filtered)} series)")
    logger.info(f"{'='*60}")

    final_rows = []
    for i, series in enumerate(filtered):
        # Gemini enrichment for subjective analysis
        enrichment = await gemini_enrich_series(
            series.get('gr_series_name', series.get('series_name', '')),
            series.get('author_name', ''),
            series.get('description', ''),
            subgenre,
        )
        await asyncio.sleep(random.uniform(0.5, 1.5))

        row = format_to_xlsx_schema(series, subgenre, enrichment)
        if row:
            final_rows.append(row)

        if (i + 1) % 10 == 0:
            logger.info(f"  Enriched {i+1}/{len(filtered)} series")

    if final_rows:
        result_df = pd.DataFrame(final_rows)
        # Sort by commissioning score
        result_df = result_df.sort_values('Commissioning_Score', ascending=False)
        result_df.to_csv(final_path, index=False)
        logger.success(f"\n  Saved {len(result_df)} series to {final_path}")

        # Stats
        for rank in ['P0', 'P1', 'P2', 'P3', 'P4', 'P5']:
            count = len(result_df[result_df['Commissioning_Rank'] == rank])
            if count > 0:
                logger.info(f"    {rank}: {count} series")
    else:
        logger.warning(f"  No series to save for {subgenre}")


# ============================================================================
# MAIN ORCHESTRATOR
# ============================================================================

async def run_all(subgenres, existing_data_path, phase=None):
    """Run pipeline for all specified subgenres."""
    # Load existing data for dedup
    existing_keys = set()
    if existing_data_path and os.path.exists(existing_data_path):
        logger.info(f"Loading existing data for dedup from {existing_data_path}")
        existing_keys = load_existing_series(existing_data_path)
        logger.info(f"  {len(existing_keys)} existing series/author keys loaded")

    # Also add cross-subgenre dedup from our own output
    for sg in subgenres:
        safe_name = re.sub(r'[/\\:*?"<>| &]', '_', sg)
        prev_path = os.path.join(OUTPUT_DIR, f"{safe_name}_expanded.csv")
        if os.path.exists(prev_path):
            try:
                prev_df = pd.read_csv(prev_path)
                for _, row in prev_df.iterrows():
                    s = normalize_name(str(row.get('Book Series Name', '')))
                    a = normalize_name(str(row.get('Author Name', '')))
                    existing_keys.add(f"{s}|||{a}")
                    existing_keys.add(s)
            except:
                pass

    for subgenre in subgenres:
        try:
            await run_subgenre_pipeline(subgenre, existing_keys, phase)

            # Add newly found series to existing_keys for cross-genre dedup
            safe_name = re.sub(r'[/\\:*?"<>| &]', '_', subgenre)
            final_path = os.path.join(OUTPUT_DIR, f"{safe_name}_expanded.csv")
            if os.path.exists(final_path):
                try:
                    df = pd.read_csv(final_path)
                    for _, row in df.iterrows():
                        s = normalize_name(str(row.get('Book Series Name', '')))
                        a = normalize_name(str(row.get('Author Name', '')))
                        existing_keys.add(f"{s}|||{a}")
                        existing_keys.add(s)
                except:
                    pass

        except Exception as e:
            logger.error(f"Pipeline failed for {subgenre}: {e}")
            continue

    # Combine all outputs into a master CSV
    logger.info(f"\n{'='*60}")
    logger.info("COMBINING ALL SUBGENRE OUTPUTS")
    logger.info(f"{'='*60}")

    all_dfs = []
    for sg in subgenres:
        safe_name = re.sub(r'[/\\:*?"<>| &]', '_', sg)
        path = os.path.join(OUTPUT_DIR, f"{safe_name}_expanded.csv")
        if os.path.exists(path):
            df = pd.read_csv(path)
            all_dfs.append(df)
            logger.info(f"  {sg}: {len(df)} series")

    if all_dfs:
        master = pd.concat(all_dfs, ignore_index=True)

        # Final cross-genre dedup on the master
        master['_dedup_key'] = master['Book Series Name'].apply(normalize_name)
        master = master.drop_duplicates(subset='_dedup_key', keep='first').drop(columns='_dedup_key')

        master_path = os.path.join(OUTPUT_DIR, "all_genres_expanded_master.csv")
        master.to_csv(master_path, index=False)
        logger.success(f"\nMaster file: {master_path} ({len(master)} total series)")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Expand Genre Discovery — Gemini-powered series finder")
    parser.add_argument("--all", action="store_true", help="Process all subgenres")
    parser.add_argument("--genre", type=str, help="Single subgenre to process")
    parser.add_argument("--phase", type=str, choices=['discover', 'validate', 'enrich'],
                        help="Run only a specific phase")
    parser.add_argument("--existing-data", type=str,
                        default=os.path.join(PROJECT_ROOT, "sub genre analysis", "Sub genre analysis- Self Pub universe.xlsx"),
                        help="Path to existing data for deduplication")
    parser.add_argument("--workers", type=int, default=4, help="Parallel workers for validation")
    parser.add_argument("--visible", action="store_true", help="Run browser in visible mode")
    parser.add_argument("--list", action="store_true", help="List available subgenres")
    args = parser.parse_args()

    global WORKER_COUNT, HEADLESS  # noqa
    WORKER_COUNT = args.workers
    if args.visible:
        HEADLESS = False

    if args.list:
        print("\nAvailable subgenres:")
        for i, sg in enumerate(SUBGENRES, 1):
            print(f"  {i}. {sg}")
        return

    if args.all:
        subgenres = SUBGENRES
    elif args.genre:
        # Fuzzy match genre name
        matched = [sg for sg in SUBGENRES if args.genre.lower() in sg.lower()]
        if not matched:
            logger.error(f"No matching subgenre for '{args.genre}'. Use --list to see options.")
            return
        subgenres = matched
    else:
        parser.print_help()
        return

    asyncio.run(run_all(subgenres, args.existing_data, args.phase))


if __name__ == "__main__":
    main()
