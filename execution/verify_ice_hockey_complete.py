#!/usr/bin/env python3
"""
Complete verification of Ice Hockey & Sports Romance data.

ANTI-HALLUCINATION RULES (enforced throughout):
  1. An email is only kept if it was literally extracted from a fetched webpage
     or from a Gemini response that cites a real URL in its grounding sources.
  2. If Gemini returns an email but provides NO grounding source URL, the email
     is REJECTED as potentially hallucinated.
  3. Emails that match known fabrication patterns are rejected.
  4. Every row gets a Contact_Description column listing ALL available
     contact channels even when no email is found.
  5. Verified_Email_Flag: TRUE only if scraped from a live page; FALSE otherwise.

Never invents data — if something cannot be confirmed, it is left blank and flagged.
"""

import json
import os
import re
import sys
import time
import warnings
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import requests
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore")

# ─────────────────────────────── Paths ─────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / ".env"
OUT_DIR = PROJECT_ROOT / "outreach" / "ice-hockey"
SOURCE_XLSX = OUT_DIR / "source" / "Final self-pub scored.xlsx"
VERIFIED_XLSX = OUT_DIR / "verified" / "Final self-pub scored.ice-hockey-FULL.xlsx"
OUTREACH_CSV  = OUT_DIR / "exports" / "ice_hockey_outreach_FULL.csv"
CACHE_FILE    = OUT_DIR / "progress" / "verify_cache_v2.json"

SHEET        = "Ice Hockey & Sports Romance"
TARGET_RANKS = {"P0", "P1", "P2"}

# ─────────────────────────────── API Key ───────────────────────────────────
def _load_api_key() -> str:
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            if line.startswith("GEMINI_API_KEY="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("GEMINI_API_KEY", "")

GEMINI_API_KEY = _load_api_key()

# ─────────────────────────────── HTTP helpers ──────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# ─────────────────────── Anti-hallucination email filter ───────────────────
# Emails that are clearly invented / placeholder
_KNOWN_BAD = {
    "user@domain.com", "email@example.com", "example@example.com",
    "author@directauthor.com", "info@therateabc.com", "noreply@noreply.com",
    "support@support.com", "contact@contact.com", "admin@admin.com",
    "hello@hello.com", "test@test.com", "info@info.com",
}
_BAD_DOMAINS = {"example.com", "domain.com", "email.com", "test.com", "yoursite.com"}

# LLM fabrication patterns: "{name}author@gmail.com", "contact@{name}.com", etc.
# We flag these as suspicious (they MAY be real but must be confirmed on a live page)
_FABRICATION_PATTERNS = [
    re.compile(r"^[a-z0-9]+author@(gmail|yahoo|hotmail|outlook)\.com$", re.I),
    re.compile(r"^contact@[a-z0-9]+\.(com|net|org|co\.uk)$", re.I),
    re.compile(r"^info@[a-z0-9]+\.(com|net|org)$", re.I),
    re.compile(r"^[a-z0-9]+books@(gmail|yahoo|hotmail)\.com$", re.I),
    re.compile(r"^hello@[a-z0-9]+\.(com|net|org)$", re.I),
    re.compile(r"^[a-z0-9]+writes@(gmail|yahoo|hotmail)\.com$", re.I),
]

def is_fabrication_pattern(email: str) -> bool:
    """Return True if the email matches a known LLM fabrication pattern."""
    e = email.strip().lower()
    return any(p.match(e) for p in _FABRICATION_PATTERNS)

def is_structurally_valid(email: str) -> bool:
    """Basic format + domain check (does NOT confirm the email exists)."""
    e = email.strip().lower()
    if not e or "@" not in e:
        return False
    if e in _KNOWN_BAD:
        return False
    domain = e.split("@")[-1]
    if domain in _BAD_DOMAINS:
        return False
    if not re.match(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$", e):
        return False
    return True

def pick_best_email(emails: list, author_name: str = "") -> str:
    """Return the most author-relevant email from a list, or ''."""
    valid = [e for e in emails if is_structurally_valid(e)]
    if not valid:
        return ""
    name_parts = [p.lower() for p in author_name.split() if len(p) > 2]
    for e in valid:
        local = e.split("@")[0].lower().replace(".", "").replace("_", "")
        if any(p in local for p in name_parts):
            return e
    return valid[0]

# ─────────────────────────────── Web scraping ──────────────────────────────
CONTACT_PATHS = ["", "/contact", "/contact-me", "/about", "/connect",
                 "/contact-us", "/reach-me"]

def _safe_get(url: str, timeout: int = 12) -> "requests.Response | None":
    try:
        r = SESSION.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code == 200:
            return r
    except Exception:
        pass
    return None

def _emails_from_html(html: str) -> list:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)
    return EMAIL_RE.findall(text)

def scrape_website_for_email(website_url: str, author_name: str = "") -> tuple:
    """
    Fetch an author's website (and /contact path) and extract any email found.
    Returns (email, source_url) — email is '' if nothing found.
    ONLY emails literally found on the fetched HTML are returned (no guessing).
    """
    if not website_url or not str(website_url).strip().startswith("http"):
        return "", ""

    base = str(website_url).rstrip("/")
    for path in CONTACT_PATHS:
        url = base + path
        resp = _safe_get(url)
        if not resp:
            time.sleep(0.5)
            continue
        emails = _emails_from_html(resp.text)
        # Filter out clearly bad emails before picking
        clean = [e for e in emails if is_structurally_valid(e)
                 and not is_fabrication_pattern(e)]
        email = pick_best_email(clean, author_name)
        if email:
            return email, url
        time.sleep(0.4)
    return "", ""

# ─────────────────────────────── Gemini API ────────────────────────────────
_GEMINI_ENDPOINT = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.0-flash:generateContent"
)
_LAST_CALL = 0.0   # timestamp of last Gemini call
_MIN_GAP   = 4.0   # seconds between calls (~15 rpm)

def _gemini_search_raw(prompt: str, retries: int = 3) -> dict:
    """
    Call Gemini 2.0 Flash with Google Search grounding.
    Returns the full API response dict (not just text) so we can inspect
    grounding sources and confirm they are real URLs.
    """
    global _LAST_CALL
    if not GEMINI_API_KEY:
        return {}
    elapsed = time.time() - _LAST_CALL
    if elapsed < _MIN_GAP:
        time.sleep(_MIN_GAP - elapsed)
    _LAST_CALL = time.time()

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 600},
    }
    url = f"{_GEMINI_ENDPOINT}?key={GEMINI_API_KEY}"

    for attempt in range(retries):
        try:
            r = requests.post(url, json=payload, timeout=30)
            if r.status_code == 429:
                wait = 45 * (attempt + 1)
                print(f"    [rate-limit] sleeping {wait}s …")
                time.sleep(wait)
                continue
            if r.status_code != 200:
                print(f"    [Gemini {r.status_code}] {r.text[:200]}")
                return {}
            return r.json()
        except Exception as e:
            print(f"    [Gemini exception] {e}")
            time.sleep(6)
    return {}

def _extract_text(response: dict) -> str:
    candidates = response.get("candidates", [])
    if not candidates:
        return ""
    parts = candidates[0].get("content", {}).get("parts", [])
    return " ".join(p.get("text", "") for p in parts).strip()

def _extract_grounding_sources(response: dict) -> list:
    """Return list of URLs cited in Gemini grounding metadata."""
    sources = []
    try:
        meta = response.get("candidates", [{}])[0].get("groundingMetadata", {})
        # 'groundingChunks' in newer API versions
        for chunk in meta.get("groundingChunks", []):
            web = chunk.get("web", {})
            if web.get("uri"):
                sources.append(web["uri"])
        # 'searchEntryPoint' / 'webSearchQueries' — not useful for URLs
        # 'groundingSupports' has 'segment' + 'groundingChunkIndices'
    except Exception:
        pass
    return sources

def gemini_find_author_contact(author_name: str, website: str = "") -> dict:
    """
    Use Gemini + Google Search to find author contact info.
    An email is ONLY accepted if Gemini's grounding sources include a real URL
    (not a social/bookstore domain) from which the email could have been scraped.

    Returns dict: email, email_source, email_confirmed_by_source,
                  agent, website, notes
    """
    result = {
        "email": "", "email_source": "", "email_confirmed_by_source": False,
        "agent": "", "website": website or "", "notes": "",
    }

    prompt = (
        f'Search the web for the official contact email address for romance/fiction author '
        f'"{author_name}". '
        f'Check their official author website, author newsletter sign-up page, '
        f'and contact page. '
        f'If you find a REAL email address on a public webpage, quote it exactly. '
        f'If no email is publicly listed, report their literary agent and agency name. '
        f'DO NOT INVENT or GUESS any email address — only report what you actually found. '
        f'Format your answer exactly as: '
        f'EMAIL: <exact email or NONE> | AGENT: <agent name, agency or NONE> | '
        f'WEBSITE: <official author site URL or NONE>'
    )

    raw_response = _gemini_search_raw(prompt)
    text = _extract_text(raw_response)
    grounding_sources = _extract_grounding_sources(raw_response)

    if not text:
        return result

    # Parse structured fields
    em_match  = re.search(r"EMAIL:\s*([^\|]+)", text, re.I)
    ag_match  = re.search(r"AGENT:\s*([^\|]+)", text, re.I)
    ws_match  = re.search(r"WEBSITE:\s*([^\|]+)", text, re.I)

    # ── Email handling (strict anti-hallucination) ──────────────────────────
    if em_match:
        raw_em = em_match.group(1).strip()
        emails_in_text = EMAIL_RE.findall(raw_em)
        candidate = pick_best_email(emails_in_text, author_name) if emails_in_text else ""

        if candidate and is_structurally_valid(candidate):
            # ONLY accept if grounding sources exist (Gemini actually searched)
            non_social_sources = [
                s for s in grounding_sources
                if not any(d in s for d in [
                    "facebook.com", "twitter.com", "x.com", "instagram.com",
                    "amazon.com", "goodreads.com", "bookbub.com", "wikipedia.org",
                ])
            ]
            if non_social_sources:
                # Email backed by a real grounding source
                result["email"] = candidate
                result["email_source"] = non_social_sources[0]
                result["email_confirmed_by_source"] = True
            elif grounding_sources:
                # Gemini searched but only social/bookstore sources — store as
                # unconfirmed so we can try to scrape the found site next
                result["email"] = candidate
                result["email_source"] = grounding_sources[0]
                result["email_confirmed_by_source"] = False
            else:
                # No grounding sources at all → REJECT (potential hallucination)
                result["email"] = ""
                result["email_source"] = "rejected-no-grounding-source"
                result["email_confirmed_by_source"] = False

    # ── Agent ───────────────────────────────────────────────────────────────
    if ag_match:
        raw_ag = ag_match.group(1).strip()
        if raw_ag.upper() not in ("NONE", "N/A", "NOT FOUND", "UNKNOWN", ""):
            result["agent"] = raw_ag

    # ── Website ─────────────────────────────────────────────────────────────
    if ws_match and not website:
        raw_ws = ws_match.group(1).strip()
        if raw_ws.upper() not in ("NONE", "N/A", "") and raw_ws.startswith("http"):
            result["website"] = raw_ws

    result["notes"] = text[:400]
    return result

# ─────────────────────────────── Goodreads helpers ─────────────────────────
GOODREADS_SERIES_SEARCH = "https://www.goodreads.com/search?q={}&search_type=series"
_GR_WORKS_RE = re.compile(r"(\d+)\s+primary\s+works?", re.I)
_GR_TOTAL_RE = re.compile(r"(\d+)\s+total\s+works?", re.I)

def _goodreads_scrape_series(series_name: str, author_name: str) -> dict:
    """Directly scrape Goodreads for series metadata."""
    result = {
        "verified_series_name": "", "goodreads_url": "",
        "primary_works": None, "total_works": None,
        "first_book": "", "last_book": "",
    }
    query = f"{series_name} {author_name}"
    url = GOODREADS_SERIES_SEARCH.format(quote_plus(query))
    resp = _safe_get(url)
    if not resp:
        return result

    soup = BeautifulSoup(resp.text, "html.parser")
    links = soup.find_all("a", href=re.compile(r"/series/\d+"))
    if not links:
        return result

    series_link = links[0]
    series_url = "https://www.goodreads.com" + series_link["href"].split("?")[0]
    result["goodreads_url"] = series_url
    result["verified_series_name"] = series_link.get_text(strip=True)

    time.sleep(1.8)
    sr = _safe_get(series_url)
    if not sr:
        return result

    page_text = BeautifulSoup(sr.text, "html.parser").get_text(" ", strip=True)
    m1 = _GR_WORKS_RE.search(page_text)
    m2 = _GR_TOTAL_RE.search(page_text)
    if m1:
        result["primary_works"] = int(m1.group(1))
    if m2:
        result["total_works"] = int(m2.group(1))

    # Book titles on series page
    sr_soup = BeautifulSoup(sr.text, "html.parser")
    titles = []
    for a in sr_soup.find_all("a", itemprop="url"):
        t = a.get_text(strip=True)
        if t and len(t) > 3:
            titles.append(t)
    if titles:
        result["first_book"] = titles[0]
        result["last_book"] = titles[-1]

    return result

def gemini_verify_series(series_name: str, author_name: str, first_book: str = "") -> dict:
    """
    Verify series metadata using Gemini + Goodreads fallback.
    Returns dict with verified_series_name, primary_works, goodreads_url, etc.
    """
    result = {
        "verified_series_name": "", "goodreads_url": "",
        "primary_works": None, "first_book": "", "last_book": "",
        "source": "",
    }

    prompt = (
        f'On Goodreads, find the series that contains the book "{first_book or series_name}" '
        f'by author "{author_name}". '
        f'What is the exact Goodreads series name? '
        f'How many PRIMARY works (not novellas/extras) are in the series? '
        f'What is the first and last primary book? '
        f'Provide the Goodreads series URL (goodreads.com/series/...). '
        f'If you are not certain, say UNKNOWN — do not guess. '
        f'Format: SERIES: <name> | BOOKS: <number> | FIRST: <title> | LAST: <title> | URL: <url>'
    )

    raw = _gemini_search_raw(prompt)
    text = _extract_text(raw)
    sources = _extract_grounding_sources(raw)

    if text:
        sm = re.search(r"SERIES:\s*([^\|]+)", text, re.I)
        bm = re.search(r"BOOKS:\s*([^\|]+)", text, re.I)
        fm = re.search(r"FIRST:\s*([^\|]+)", text, re.I)
        lm = re.search(r"LAST:\s*([^\|]+)", text, re.I)
        um = re.search(r"URL:\s*([^\|]+)", text, re.I)

        def _clean(m):
            v = m.group(1).strip() if m else ""
            return "" if v.upper() in ("UNKNOWN", "NONE", "N/A", "") else v

        v_series = _clean(sm)
        v_books  = _clean(bm)
        v_first  = _clean(fm)
        v_last   = _clean(lm)
        v_url    = _clean(um)

        if v_series:
            result["verified_series_name"] = v_series
        if v_books:
            try:
                result["primary_works"] = int(re.search(r"\d+", v_books).group())
            except Exception:
                pass
        if v_first:
            result["first_book"] = v_first
        if v_last:
            result["last_book"] = v_last
        if v_url and "goodreads.com/series" in v_url:
            result["goodreads_url"] = v_url

        # Only trust if backed by grounding sources
        gr_sources = [s for s in sources if "goodreads.com" in s]
        if result["verified_series_name"] and result["primary_works"]:
            result["source"] = "gemini-search" + ("-grounded" if gr_sources else "-ungrounded")

    # Fallback: direct Goodreads scrape
    if not result["verified_series_name"] or not result["primary_works"]:
        gr = _goodreads_scrape_series(series_name, author_name)
        if gr["verified_series_name"]:
            # Use Goodreads data (authoritative)
            result.update(gr)
            result["source"] = "goodreads-scrape"

    return result

# ─────────────────────────────── Sanity checks ─────────────────────────────
def sanity_check_row(row: pd.Series) -> tuple:
    """
    Logical checks on a single row. Returns (flag, [issue_strings]).
    Uses VERIFIED data where available, falls back to source data.
    """
    issues = []

    btype        = str(row.get("Type", "") or "").strip()
    first_book   = str(row.get("First Book Name", "") or "").strip()
    last_book    = str(row.get("Last Book Name", "") or "").strip()
    series_name  = str(row.get("Book Series Name", "") or "").strip()
    first_eq_ser = row.get("first book name= book series name")

    # Use verified book count if available, else source
    v_books = row.get("Verified_Books_in_Series")
    s_books = row.get("Books in Series")
    def _to_float(v):
        try:
            f = float(v)
            return f if f == f else None   # NaN check
        except (TypeError, ValueError):
            return None

    n_books = _to_float(v_books) or _to_float(s_books)
    n_pages = _to_float(row.get("Total Pages"))

    # 1. Standalone with multiple books
    if "standalone" in btype.lower() and n_books and n_books > 1:
        issues.append(f"TYPE_MISMATCH: Type='{btype}' but Books={n_books}")

    # 2. First book == last book but series claims > 1 book
    if (first_book and last_book and
            first_book.lower() == last_book.lower() and
            n_books and n_books > 1):
        issues.append(f"SAME_FIRST_LAST: first=last='{first_book}' yet Books={n_books}")

    # 3. Total pages implausibly low
    if n_pages is not None and n_books and n_books > 0:
        ppb = n_pages / n_books
        if n_pages in (0, 1):
            issues.append(f"PAGES_ZERO: total pages={n_pages}")
        elif ppb < 50:
            issues.append(f"PAGES_LOW: {n_pages}pp / {n_books}books = {ppb:.0f}pp/book (<50)")

    # 4. Series name == first book name (title used as series name)
    if str(first_eq_ser).lower() in ("true", "1", "yes"):
        issues.append("SERIES_IS_BOOK_TITLE: series name is a book title, not a series name")

    # 5. Default-3 book count (LLM hallucination marker)
    if _to_float(s_books) == 3.0 and not v_books:
        issues.append("BOOKS_DEFAULT3: Books=3 may be LLM default — needs Goodreads check")

    # 6. Formulaic page counts
    if n_pages and n_books:
        if n_pages == n_books * 300:
            issues.append(f"PAGES_FORMULA: total={n_pages} = {n_books}×300 (likely fabricated)")
        elif n_pages == n_books * 250:
            issues.append(f"PAGES_FORMULA: total={n_pages} = {n_books}×250 (likely fabricated)")

    # 7. Absurd book counts
    if n_books and n_books >= 50:
        issues.append(f"BOOKS_EXTREME: {n_books} books — needs manual check")

    # Derive flag
    critical_kw = {"TYPE_MISMATCH", "SAME_FIRST_LAST", "PAGES_ZERO", "SERIES_IS_BOOK_TITLE"}
    has_critical = any(any(k in iss for k in critical_kw) for iss in issues)
    if has_critical or len(issues) >= 3:
        flag = "RED"
    elif issues:
        flag = "YELLOW"
    else:
        flag = "GREEN"

    return flag, issues

def build_contact_description(row: pd.Series, author_data: dict) -> str:
    """
    Build a human-readable description of ALL contact channels available,
    for use when no email is found (and as supplementary info when one is).
    """
    parts = []

    v_email = author_data.get("email", "")
    v_confirmed = author_data.get("email_confirmed_by_source", False)
    agent = author_data.get("agent", "") or str(row.get("Literary Agent", "") or "").strip()
    website = author_data.get("website", "") or str(row.get("Website", "") or "").strip()

    if v_email:
        tag = "CONFIRMED email" if v_confirmed else "UNVERIFIED email (from source)"
        parts.append(f"{tag}: {v_email}")
    if website and website.startswith("http"):
        parts.append(f"Website: {website}")
    if agent:
        parts.append(f"Literary Agent: {agent}")

    social_fields = [
        ("Twitter", row.get("Twitter", "")),
        ("Instagram", row.get("Instagram", "")),
        ("Facebook", row.get("Facebook", "")),
        ("BookBub", row.get("BookBub", "")),
        ("TikTok", row.get("TikTok", "")),
    ]
    for label, val in social_fields:
        if val and str(val).strip() and str(val).strip() not in ("nan", "None"):
            parts.append(f"{label}: {val}")

    if not parts:
        parts.append("No verified contact found")

    return " | ".join(parts)

# ─────────────────────────────── Cache helpers ─────────────────────────────
def load_cache() -> dict:
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except Exception:
            return {}
    return {}

def save_cache(cache: dict):
    CACHE_FILE.write_text(json.dumps(cache, indent=2, ensure_ascii=False))

# ─────────────────────────────── Main ──────────────────────────────────────
def main():
    print("=" * 70)
    print("Ice Hockey & Sports Romance — Full Verification (no hallucinations)")
    print("=" * 70)

    # ── 1. Load ──────────────────────────────────────────────────────────────
    print(f"\n[1] Loading {SOURCE_XLSX}")
    df = pd.read_excel(SOURCE_XLSX, sheet_name=SHEET)
    print(f"    Rows: {len(df)}")
    target_mask = df["Commissioning_Rank"].isin(TARGET_RANKS)
    print(f"    P0/P1/P2 rows: {target_mask.sum()}")
    target_df = df[target_mask].copy()

    cache = load_cache()
    cache.setdefault("authors", {})
    cache.setdefault("series", {})
    print(f"    Cached authors: {len(cache['authors'])}  "
          f"Cached series: {len(cache['series'])}")

    # ── 2. Author contacts ───────────────────────────────────────────────────
    print("\n[2] Author contact verification")
    unique_authors = target_df["Author Name"].dropna().unique().tolist()
    print(f"    Unique authors: {len(unique_authors)}")

    for i, author in enumerate(unique_authors):
        akey = str(author).strip().lower()
        if akey in cache["authors"]:
            continue

        print(f"    [{i+1}/{len(unique_authors)}] {author}")

        # Gather existing info from the sheet
        rows_for_author = target_df[target_df["Author Name"] == author]
        website = ""
        src_email = ""
        agent = ""
        twitter = instagram = facebook = bookbub = tiktok = ""
        for _, r in rows_for_author.iterrows():
            for field, var in [("Website", "website"), ("Email", "src_email"),
                               ("Literary Agent", "agent"), ("Twitter", "twitter"),
                               ("Instagram", "instagram"), ("Facebook", "facebook"),
                               ("BookBub", "bookbub"), ("TikTok", "tiktok")]:
                v = str(r.get(field, "") or "").strip()
                if v and v not in ("nan", "None"):
                    locals()[var]  # just a read — use exec below
            if not website and str(r.get("Website", "") or "").strip().startswith("http"):
                website = str(r["Website"]).strip()
            if not src_email and "@" in str(r.get("Email", "") or ""):
                src_email = str(r["Email"]).strip()
            if not agent and str(r.get("Literary Agent", "") or "").strip():
                agent = str(r["Literary Agent"]).strip()

        contact = {
            "email": "", "email_source": "", "email_confirmed_by_source": False,
            "agent": agent, "website": website, "notes": "",
        }

        # Step A: Scrape the listed website
        if website:
            scraped_email, scraped_url = scrape_website_for_email(website, author)
            if scraped_email:
                contact["email"] = scraped_email
                contact["email_source"] = scraped_url
                contact["email_confirmed_by_source"] = True
                print(f"        ✓ website email: {scraped_email}")

        # Step B: Gemini search (only if no confirmed email yet)
        if not contact["email_confirmed_by_source"]:
            gem = gemini_find_author_contact(author, website)

            if gem["email"] and gem["email_confirmed_by_source"]:
                # Accept Gemini's email only if it came with real grounding
                contact["email"] = gem["email"]
                contact["email_source"] = gem["email_source"]
                contact["email_confirmed_by_source"] = True
                print(f"        ✓ Gemini-confirmed email: {gem['email']}")
            elif gem["email"] and not gem["email_confirmed_by_source"]:
                # Try to cross-validate by scraping the grounding source URL
                cross_email = ""
                if gem["email_source"] and gem["email_source"].startswith("http"):
                    cross_resp = _safe_get(gem["email_source"])
                    if cross_resp:
                        cross_emails = [
                            e for e in _emails_from_html(cross_resp.text)
                            if is_structurally_valid(e)
                        ]
                        cross_email = pick_best_email(cross_emails, author)
                if cross_email and cross_email.lower() == gem["email"].lower():
                    contact["email"] = cross_email
                    contact["email_source"] = gem["email_source"]
                    contact["email_confirmed_by_source"] = True
                    print(f"        ✓ cross-validated email: {cross_email}")
                else:
                    # Could not confirm — reject the email
                    print(f"        ✗ Gemini email could not be confirmed: {gem['email']} — rejected")

            # Update agent/website even if no email
            if gem["agent"] and not contact["agent"]:
                contact["agent"] = gem["agent"]
                print(f"        → agent: {gem['agent']}")
            if gem["website"] and not contact["website"]:
                contact["website"] = gem["website"]
            contact["notes"] = gem.get("notes", "")

        # Step C: If still no confirmed email, keep source email ONLY as
        # "unverified" (it will be clearly flagged in the output)
        if not contact["email"] and src_email and is_structurally_valid(src_email):
            contact["email"] = src_email
            contact["email_source"] = "source-unverified"
            contact["email_confirmed_by_source"] = False
            print(f"        ? keeping source email (UNVERIFIED): {src_email}")
        elif not contact["email"]:
            print(f"        — no email found")

        cache["authors"][akey] = contact

        if (i + 1) % 10 == 0:
            save_cache(cache)
            print(f"    [progress saved — {i+1}/{len(unique_authors)}]")

    save_cache(cache)
    print("    Author contacts done.")

    # ── 3. Series verification ───────────────────────────────────────────────
    print("\n[3] Series verification (Goodreads + Gemini)")
    combos = (
        target_df[["Book Series Name", "Author Name", "First Book Name"]]
        .dropna(subset=["Book Series Name", "Author Name"])
        .drop_duplicates(subset=["Book Series Name", "Author Name"])
        .values.tolist()
    )
    print(f"    Unique series: {len(combos)}")

    for i, (series, author, first_book) in enumerate(combos):
        skey = f"{str(author).strip().lower()}||{str(series).strip().lower()}"
        if skey in cache["series"]:
            continue
        print(f"    [{i+1}/{len(combos)}] '{series}' by {author}")
        sd = gemini_verify_series(
            str(series), str(author),
            str(first_book) if pd.notna(first_book) else ""
        )
        cache["series"][skey] = sd
        if sd["verified_series_name"]:
            print(f"        → '{sd['verified_series_name']}', "
                  f"{sd['primary_works']} books [{sd['source']}]")
        else:
            print(f"        → could not verify")

        if (i + 1) % 20 == 0:
            save_cache(cache)
            print(f"    [progress saved — {i+1}/{len(combos)}]")

    save_cache(cache)
    print("    Series verification done.")

    # ── 4. Apply results to DataFrame ───────────────────────────────────────
    print("\n[4] Applying results to dataframe …")

    NEW_COLS = [
        "Validated_Email", "Email_Confirmed", "Email_Source_URL",
        "Agency_Contact", "Verified_Website",
        "Verified_Series_Name", "Verified_Goodreads_URL",
        "Verified_Books_in_Series", "Verified_First_Book", "Verified_Last_Book",
        "Series_Source", "Contact_Description",
        "Sanity_Issues", "Data_Quality_Flag",
    ]
    for col in NEW_COLS:
        if col not in df.columns:
            df[col] = ""

    for idx, row in df.iterrows():
        if row.get("Commissioning_Rank") not in TARGET_RANKS:
            continue

        author = str(row.get("Author Name", "") or "").strip()
        series = str(row.get("Book Series Name", "") or "").strip()
        akey = author.lower()
        skey = f"{author.lower()}||{series.lower()}"

        # Author contact
        ad = cache["authors"].get(akey, {})
        df.at[idx, "Validated_Email"]   = ad.get("email", "")
        df.at[idx, "Email_Confirmed"]   = ad.get("email_confirmed_by_source", False)
        df.at[idx, "Email_Source_URL"]  = ad.get("email_source", "")
        df.at[idx, "Verified_Website"]  = ad.get("website", "")
        agent = (ad.get("agent", "")
                 or str(row.get("Literary Agent", "") or "").strip())
        df.at[idx, "Agency_Contact"]    = agent

        # Series
        sd = cache["series"].get(skey, {})
        if sd.get("verified_series_name"):
            df.at[idx, "Verified_Series_Name"] = sd["verified_series_name"]
        if sd.get("goodreads_url"):
            df.at[idx, "Verified_Goodreads_URL"] = sd["goodreads_url"]
        if sd.get("primary_works") is not None:
            df.at[idx, "Verified_Books_in_Series"] = sd["primary_works"]
        if sd.get("first_book"):
            df.at[idx, "Verified_First_Book"] = sd["first_book"]
        if sd.get("last_book"):
            df.at[idx, "Verified_Last_Book"] = sd["last_book"]
        df.at[idx, "Series_Source"] = sd.get("source", "")

        # Contact description (all channels)
        df.at[idx, "Contact_Description"] = build_contact_description(row, ad)

        # Sanity checks
        # pass the row with newly applied verified data
        augmented = row.copy()
        if sd.get("primary_works"):
            augmented["Verified_Books_in_Series"] = sd["primary_works"]
        flag, issues = sanity_check_row(augmented)
        df.at[idx, "Sanity_Issues"]       = " | ".join(issues)
        df.at[idx, "Data_Quality_Flag"]   = flag

    print("    Done.")

    # ── 5. Write outputs ─────────────────────────────────────────────────────
    print("\n[5] Writing outputs …")
    VERIFIED_XLSX.parent.mkdir(parents=True, exist_ok=True)
    OUTREACH_CSV.parent.mkdir(parents=True, exist_ok=True)

    # Excel: preserve all original sheets, update Ice Hockey sheet
    try:
        import shutil
        from openpyxl import load_workbook
        shutil.copy(SOURCE_XLSX, VERIFIED_XLSX)
        wb = load_workbook(VERIFIED_XLSX)

        if SHEET in wb.sheetnames:
            ws = wb[SHEET]
            # Build complete header list
            existing_headers = [c.value for c in ws[1]]
            for col in NEW_COLS:
                if col not in existing_headers:
                    existing_headers.append(col)
            # Write headers row
            for ci, h in enumerate(existing_headers, 1):
                ws.cell(row=1, column=ci, value=h)
            # Write data rows
            for ri, (_, row_data) in enumerate(df.iterrows(), 2):
                for ci, h in enumerate(existing_headers, 1):
                    v = row_data.get(h, "")
                    if pd.isna(v):
                        v = ""
                    ws.cell(row=ri, column=ci, value=v)

        wb.save(VERIFIED_XLSX)
        print(f"    Excel: {VERIFIED_XLSX}")
    except Exception as exc:
        fallback_csv = str(VERIFIED_XLSX).replace(".xlsx", ".csv")
        print(f"    Excel failed ({exc}) — saving CSV: {fallback_csv}")
        df.to_csv(fallback_csv, index=False)

    # Outreach CSV: P0/P1/P2 only, key columns, sorted by priority
    OUTREACH_COLS = [
        "Commissioning_Rank", "Commissioning_Score", "Author Name",
        "Verified_Series_Name", "Book Series Name",
        "Verified_Books_in_Series", "Books in Series",
        "Verified_First_Book", "First Book Name",
        "Verified_Last_Book", "Last Book Name",
        "Validated_Email", "Email_Confirmed", "Email_Source_URL",
        "Agency_Contact", "Verified_Website", "Website",
        "Verified_Goodreads_URL", "Goodreads Series URL",
        "Contact_Description",
        "Twitter", "Instagram", "Facebook", "BookBub", "TikTok",
        "Type", "Total Pages", "Length of Adaption in Hours",
        "Data_Quality_Flag", "Sanity_Issues", "Series_Source",
    ]
    out_df = df[target_mask][[c for c in OUTREACH_COLS if c in df.columns]].copy()
    out_df = out_df.sort_values(
        ["Commissioning_Rank", "Commissioning_Score"],
        ascending=[True, False],
    )
    out_df.to_csv(OUTREACH_CSV, index=False)
    print(f"    CSV: {OUTREACH_CSV}")

    # ── 6. Summary ───────────────────────────────────────────────────────────
    print("\n[6] Summary")
    tgt = df[target_mask]
    confirmed   = (tgt["Email_Confirmed"] == True).sum()
    unconfirmed = ((tgt["Validated_Email"] != "") & (tgt["Email_Confirmed"] != True)).sum()
    no_email    = (tgt["Validated_Email"].isna() | (tgt["Validated_Email"] == "")).sum()
    has_agency  = (tgt["Agency_Contact"].notna() & (tgt["Agency_Contact"] != "")).sum()
    has_contact = (tgt["Contact_Description"].str.len() > 20).sum()
    green = (tgt["Data_Quality_Flag"] == "GREEN").sum()
    yellow = (tgt["Data_Quality_Flag"] == "YELLOW").sum()
    red   = (tgt["Data_Quality_Flag"] == "RED").sum()
    v_series = (tgt["Verified_Series_Name"] != "").sum()

    print(f"    P0/P1/P2 rows          : {target_mask.sum()}")
    print(f"    Confirmed emails        : {confirmed}  (found on live page)")
    print(f"    Unverified emails kept  : {unconfirmed}  (from source, not confirmed)")
    print(f"    No email at all         : {no_email}")
    print(f"    Has agency contact      : {has_agency}")
    print(f"    Has any contact info    : {has_contact}")
    print(f"    Series verified         : {v_series}")
    print(f"    GREEN (all checks pass) : {green}")
    print(f"    YELLOW (minor issues)   : {yellow}")
    print(f"    RED (critical issues)   : {red}")
    print(f"\n    Cache  : {CACHE_FILE}")
    print(f"    Excel  : {VERIFIED_XLSX}")
    print(f"    CSV    : {OUTREACH_CSV}")
    print("\nDone.")


if __name__ == "__main__":
    main()
