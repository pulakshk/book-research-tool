#!/usr/bin/env python3
"""
Amazon Full Top 100 Crawler — Paid & Free across all categories
================================================================
Crawls Amazon bestseller lists (both paid and free) across all relevant
romance/fiction categories. Outputs a flat reference table and adds it
as a new tab to FINAL_SELFPUB_SCORED.xlsx.

Output columns:
  1. Amazon Top 100 List  (e.g., "Historical Romance Top 100 Paid")
  2. Rank                 (#1, #2, ...)
  3. Book Name
  4. Book Series Name     (extracted if present)
  5. Author Name
  6. Crawl Date           (today's date)
"""

import asyncio
import re
import random
import logging
import json
from pathlib import Path
from datetime import datetime

import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "output"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(DATA_DIR / "amazon_full_crawl.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("amazon_full_crawl")

TODAY = datetime.now().strftime("%Y-%m-%d")

# ── Amazon Browse Node Categories ─────────────────────────────────────
# Format: (display_name, node_id, path_type)
# path_type: "digital-text" for Kindle, "books" for print
CATEGORIES = [
    # ── Top-level Romance (Kindle) ────────────────────────────
    ("Romance", "158566011", "digital-text"),
    ("Contemporary Romance", "6487818011", "digital-text"),
    ("Romantic Suspense", "6487830011", "digital-text"),
    ("Historical Romance", "6487814011", "digital-text"),
    ("Military Romance", "6487822011", "digital-text"),
    ("Christian Romance", "6487826011", "digital-text"),
    ("Sports Romance", "7588834011", "digital-text"),
    ("Dark Romance", "23488356011", "digital-text"),
    ("Small Town Romance", "25381371011", "digital-text"),
    ("Crime Fiction Romance", "7588814011", "digital-text"),
    ("Political Romance", "7588826011", "digital-text"),
    ("New Adult Romance", "17741157011", "digital-text"),
    ("Multicultural Romance", "7588822011", "digital-text"),
    ("Gothic Romance", "6487812011", "digital-text"),
    ("Paranormal Romance", "6487828011", "digital-text"),
    ("Western Romance", "6487832011", "digital-text"),
    ("Holiday Romance", "10197", "digital-text"),
    ("Romantic Comedy", "6487820011", "digital-text"),
    # ── Broader Fiction (Kindle) ──────────────────────────────
    ("Literature & Fiction", "17", "digital-text"),
    ("Women's Fiction", "542654", "digital-text"),
    ("Mystery Thriller & Suspense", "18", "digital-text"),
    ("Christian Fiction", "12290", "digital-text"),
    ("Historical Fiction", "10177", "digital-text"),
    ("Action & Adventure", "720", "digital-text"),
    ("African American Romance", "7588806011", "digital-text"),
    ("Clean & Wholesome Romance", "23488354011", "digital-text"),
    ("Inspirational Romance", "6487816011", "digital-text"),
    # ── Print Books - Romance ─────────────────────────────────
    ("Romance (Print)", "23", "books"),
    ("Contemporary Romance (Print)", "23747", "books"),
    ("Romantic Suspense (Print)", "10138", "books"),
    ("Historical Romance (Print)", "10136", "books"),
    ("Military Romance (Print)", "10139", "books"),
    ("Western Romance (Print)", "25", "books"),
    # ── Print Books - Related Fiction ─────────────────────────
    ("Women's Fiction (Print)", "542654", "books"),
    ("Mystery & Thriller (Print)", "18", "books"),
    ("Historical Fiction (Print)", "10177", "books"),
    ("Christian Fiction (Print)", "12290", "books"),
]

# URL templates - proven working format
URL_PAID = "https://www.amazon.com/Best-Sellers/zgbs/{path}/{node}/?pg={pg}"
URL_FREE = "https://www.amazon.com/Best-Sellers-Free/zgbs/free/{path}/{node}/?pg={pg}"


async def extract_items_from_page(page):
    """Try multiple selectors to find bestseller items on the page."""
    # Try selectors in order of preference
    selectors = [
        "#gridItemRoot",
        "div[id^='p13n-asin-index-']",
        "div.a-cardui._cDEzb_grid-cell_1uMOS",
        "div[class*='zg-grid-general-faceout']",
        "li.zg-item-immersion",
        "div.zg-item",
        "span.zg-item",
        "div[data-asin]",
    ]

    for selector in selectors:
        items = await page.query_selector_all(selector)
        if items and len(items) > 3:
            return items, selector

    return [], "none"


def parse_item_text(full_text):
    """Parse rank, title, author from an item's text content."""
    full_text = full_text.strip()

    # Extract rank from text (#1, #2, etc.)
    rank_match = re.match(r'#(\d+)', full_text)
    rank = int(rank_match.group(1)) if rank_match else 0

    title = ""
    author = ""

    # Try splitting by newlines first
    lines = [l.strip() for l in full_text.split('\n') if l.strip()]

    skip_words = ['star', 'Kindle', 'INR', '$', 'Hardcover', 'Paperback',
                  'Audio', 'out of', 'rating', 'Price', '₹', 'FREE', 'format']

    # If all content is in a single line (common Amazon format), parse it differently
    non_hash_lines = [l for l in lines if not l.startswith('#')]
    if len(non_hash_lines) == 0 and len(lines) > 0:
        # Single line starting with #N - strip the rank prefix and parse
        blob = re.sub(r'^#\d+\s*', '', lines[0])
        # Pattern: "TitleAuthorRating..." - split on rating pattern
        # Rating usually looks like "4.7 out of 5 stars" or "X.X out of"
        parts = re.split(r'(\d+\.\d+\s+out\s+of\s+5\s+star)', blob)
        if len(parts) >= 2:
            title_author = parts[0].strip()
            # Split title from author - they're concatenated without space
            # Key heuristic: find where lowercase meets uppercase (e.g., "NovelColleen")
            # But first check for ) followed by uppercase (series pattern)
            paren_match = re.search(r'\)([A-Z])', title_author)
            if paren_match:
                split_pos = paren_match.start() + 1
                title = title_author[:split_pos].strip()
                author = title_author[split_pos:].strip()
            else:
                # Find the LAST occurrence of lowercase->uppercase boundary
                # This typically marks where the title ends and author name begins
                # e.g., "Reminders of Him: A NovelColleen Hoover"
                #                                    ^--- boundary here
                boundaries = []
                for i in range(1, len(title_author)):
                    prev = title_author[i-1]
                    curr = title_author[i]
                    # Boundary: lowercase/digit/punctuation followed by uppercase
                    if curr.isupper() and (prev.islower() or prev in '.!?,;:'):
                        boundaries.append(i)

                if boundaries:
                    # Use the last boundary that's past the first 10 chars
                    # (to avoid splitting too early)
                    valid_bounds = [b for b in boundaries if b > 10]
                    if valid_bounds:
                        split_pos = valid_bounds[-1]
                        title = title_author[:split_pos].strip()
                        author = title_author[split_pos:].strip()
                    else:
                        title = title_author
                else:
                    title = title_author
        else:
            # No rating found, just use the blob as title
            title = blob[:200] if len(blob) > 5 else ""
        return rank, title, author

    # Multi-line parsing
    for line in lines:
        line = line.strip()
        if line.startswith('#'):
            continue
        if len(line) < 3:
            continue
        if any(sw in line for sw in skip_words):
            continue
        if re.match(r'^[\d.,]+$', line):
            continue

        if not title and len(line) > 3:
            title = line
        elif title and not author and len(line) > 2 and line != title:
            author = line
            break

    return rank, title, author


def extract_series_from_title(title):
    """Try to extract series name from a book title."""
    # Patterns like "Title (Series Name Book 3)" or "Title: A Series Name Novel"
    patterns = [
        r'\((.+?)\s+(?:Book|#|Vol|Series|Bk)\s*\d',
        r'\((.+?)\s+\d+\)',
        r':\s+(?:A\s+)?(.+?)\s+(?:Novel|Romance|Story|Series)',
        r'\((.+?)\)',
    ]
    for pat in patterns:
        m = re.search(pat, title)
        if m:
            series = m.group(1).strip()
            if len(series) > 3 and len(series) < 100:
                return series
    return ""


async def scrape_list(page, category_name, node_id, path_type, list_type="Paid"):
    """Scrape a single bestseller list (paid or free)."""
    books = []

    for pg in [1, 2]:
        if list_type == "Paid":
            url = URL_PAID.format(path=path_type, node=node_id, pg=pg)
        else:
            url = URL_FREE.format(path=path_type, node=node_id, pg=pg)

        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(random.uniform(3, 5))

            items, selector = await extract_items_from_page(page)

            for item in items:
                try:
                    full_text = (await item.text_content()).strip()
                    rank, title, author = parse_item_text(full_text)

                    # Fallback: try child elements if parsing failed
                    if not title and rank > 0:
                        try:
                            divs = await item.query_selector_all("div, span, a")
                            texts = []
                            for d in divs:
                                t = (await d.text_content()).strip()
                                if t and 3 < len(t) < 200 and t not in texts:
                                    texts.append(t)
                            for t in texts:
                                if not t.startswith('#') and 'star' not in t and 'Kindle' not in t and '$' not in t and 'INR' not in t:
                                    if not title:
                                        title = t
                                    elif not author and t != title:
                                        author = t
                                        break
                        except Exception:
                            pass

                    if title and rank > 0:
                        series_name = extract_series_from_title(title)
                        list_label = f"{category_name} Top 100 {list_type}"

                        books.append({
                            "Amazon Top 100 List": list_label,
                            "Rank": rank,
                            "Book Name": title,
                            "Book Series Name": series_name,
                            "Author Name": author,
                            "Crawl Date": TODAY,
                        })
                except Exception:
                    continue

        except Exception as e:
            log.warning(f"  [{category_name} {list_type}] Page {pg} error: {e}")

        await asyncio.sleep(random.uniform(1, 2))

    return books


async def run_crawl():
    log.info("=" * 70)
    log.info("AMAZON FULL TOP 100 CRAWLER — Paid & Free")
    log.info(f"Date: {TODAY}")
    log.info(f"Categories: {len(CATEGORIES)}")
    log.info("=" * 70)

    from playwright.async_api import async_playwright

    all_books = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        async def make_stealth_context():
            ctx = await browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
                timezone_id="America/New_York",
                extra_http_headers={
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Accept-Language": "en-US,en;q=0.9",
                    "DNT": "1",
                },
            )
            await ctx.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                window.chrome = { runtime: {} };
            """)
            return ctx

        context = await make_stealth_context()
        page = await context.new_page()

        total_cats = len(CATEGORIES) * 2  # paid + free
        done = 0

        for cat_name, node_id, path_type in CATEGORIES:
            # Scrape PAID list
            log.info(f"\n  [{done+1}/{total_cats}] Scraping: {cat_name} (Paid)")
            paid_books = await scrape_list(page, cat_name, node_id, path_type, "Paid")
            all_books.extend(paid_books)
            log.info(f"    -> {len(paid_books)} books")
            done += 1
            await asyncio.sleep(random.uniform(1, 3))

            # Scrape FREE list (only for digital-text)
            if path_type == "digital-text":
                log.info(f"  [{done+1}/{total_cats}] Scraping: {cat_name} (Free)")
                free_books = await scrape_list(page, cat_name, node_id, path_type, "Free")
                all_books.extend(free_books)
                log.info(f"    -> {len(free_books)} books")
            else:
                log.info(f"  [{done+1}/{total_cats}] Skipping free list for print books")
            done += 1
            await asyncio.sleep(random.uniform(2, 4))

            # Rotate context every 10 categories to avoid detection
            if done % 20 == 0:
                try:
                    await page.close()
                    await context.close()
                except Exception:
                    pass
                context = await make_stealth_context()
                page = await context.new_page()
                log.info("    [Context rotated]")

        await browser.close()

    # Save raw data
    if all_books:
        crawl_df = pd.DataFrame(all_books)
        raw_path = DATA_DIR / "amazon_full_crawl_raw.csv"
        crawl_df.to_csv(raw_path, index=False)
        log.info(f"\n  Saved {len(all_books)} entries to {raw_path.name}")

        # Summary by list
        log.info("\n  Summary by list:")
        for list_name in crawl_df["Amazon Top 100 List"].unique():
            count = len(crawl_df[crawl_df["Amazon Top 100 List"] == list_name])
            log.info(f"    {list_name}: {count} books")

        # Add to Excel
        add_to_excel(crawl_df)
    else:
        log.warning("  No books scraped!")


def add_to_excel(crawl_df=None):
    """Add the Amazon crawl data as a new tab to FINAL_SELFPUB_SCORED.xlsx."""
    xlsx_path = DATA_DIR / "FINAL_SELFPUB_SCORED.xlsx"

    if crawl_df is None:
        raw_path = DATA_DIR / "amazon_full_crawl_raw.csv"
        if raw_path.exists():
            crawl_df = pd.read_csv(raw_path)
        else:
            log.error("No crawl data found!")
            return

    if not xlsx_path.exists():
        log.error(f"Excel file not found: {xlsx_path}")
        return

    log.info(f"\n  Adding 'Amazon Top 100 Lists' tab to {xlsx_path.name}...")

    # Order columns
    col_order = [
        "Amazon Top 100 List",
        "Rank",
        "Book Name",
        "Book Series Name",
        "Author Name",
        "Crawl Date",
    ]
    crawl_df = crawl_df[col_order].sort_values(
        ["Amazon Top 100 List", "Rank"]
    ).reset_index(drop=True)

    # Read existing Excel and append new sheet
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)

    # Remove existing sheet if present
    if "Amazon Top 100 Lists" in wb.sheetnames:
        del wb["Amazon Top 100 Lists"]

    wb.save(xlsx_path)

    # Write with ExcelWriter in append mode
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as writer:
        crawl_df.to_excel(writer, sheet_name="Amazon Top 100 Lists", index=False)

        # Auto-size columns
        try:
            ws = writer.sheets["Amazon Top 100 Lists"]
            col_widths = {
                "A": 45,  # List name
                "B": 8,   # Rank
                "C": 60,  # Book Name
                "D": 40,  # Series Name
                "E": 25,  # Author
                "F": 12,  # Date
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
        except Exception:
            pass

    log.info(f"  Added 'Amazon Top 100 Lists' sheet ({len(crawl_df)} rows)")

    # Also do cross-reference mapping
    map_to_dataset(crawl_df)


def map_to_dataset(crawl_df):
    """Map Amazon crawl data against our scored dataset for enhanced tagging."""
    scored_path = DATA_DIR / "FINAL_SELFPUB_SCORED.csv"
    if not scored_path.exists():
        return

    df = pd.read_csv(scored_path, low_memory=False)
    log.info(f"\n  Cross-referencing against {len(df)} scored series...")

    if "Amazon_Bestseller_Tag" not in df.columns:
        df["Amazon_Bestseller_Tag"] = ""
    if "Amazon_Best_Rank" not in df.columns:
        df["Amazon_Best_Rank"] = ""

    # Build lookups
    title_to_idx = {}
    author_to_idx = {}
    for idx, row in df.iterrows():
        for col in ["Book Series Name", "First Book Name"]:
            val = str(row.get(col, "")).lower().strip()
            if val and val != "nan":
                norm = re.sub(r'[^\w\s]', '', val).strip()
                if len(norm) > 3:
                    title_to_idx.setdefault(norm, []).append(idx)
        author = str(row.get("Author Name", "")).lower().strip()
        if author and author != "nan":
            author_to_idx.setdefault(author, []).append(idx)

    matched = 0
    for _, amz in crawl_df.iterrows():
        amz_title = re.sub(r'[^\w\s]', '', str(amz["Book Name"]).lower().strip())
        amz_series = re.sub(r'[^\w\s]', '', str(amz.get("Book Series Name", "")).lower().strip())
        amz_author = str(amz.get("Author Name", "")).lower().strip()
        rank = int(amz["Rank"]) if pd.notna(amz["Rank"]) else 999
        list_name = str(amz["Amazon Top 100 List"])

        if rank <= 10:
            tag = f"Amazon Top 10 {list_name}"
        elif rank <= 50:
            tag = f"Amazon Top 50 {list_name}"
        else:
            tag = f"Amazon Top 100 {list_name}"

        matched_indices = set()

        # Title match
        for search_term in [amz_title, amz_series]:
            if not search_term or len(search_term) < 4:
                continue
            if search_term in title_to_idx:
                matched_indices.update(title_to_idx[search_term])
            elif len(search_term) > 8:
                for our_title, indices in title_to_idx.items():
                    if len(our_title) > 5:
                        if our_title in search_term or search_term in our_title:
                            matched_indices.update(indices)
                            break

        # Author match
        if not matched_indices and amz_author and len(amz_author) > 5:
            if amz_author in author_to_idx:
                matched_indices.update(author_to_idx[amz_author])

        for idx in matched_indices:
            current_tag = str(df.at[idx, "Amazon_Bestseller_Tag"]).strip()
            if current_tag and current_tag not in ["", "nan"]:
                if tag not in current_tag:
                    df.at[idx, "Amazon_Bestseller_Tag"] = current_tag + "; " + tag
            else:
                df.at[idx, "Amazon_Bestseller_Tag"] = tag

            current_rank = df.at[idx, "Amazon_Best_Rank"]
            try:
                cr = int(float(current_rank)) if pd.notna(current_rank) and str(current_rank).strip() not in ["", "nan"] else 999
            except (ValueError, TypeError):
                cr = 999
            if rank < cr:
                df.at[idx, "Amazon_Best_Rank"] = rank
            matched += 1

    has_tag = df["Amazon_Bestseller_Tag"].notna() & (~df["Amazon_Bestseller_Tag"].astype(str).str.strip().isin(["", "nan"]))
    log.info(f"  Matched {matched} entries -> {has_tag.sum()} series tagged")

    df.to_csv(scored_path, index=False)
    log.info(f"  Updated {scored_path.name}")


if __name__ == "__main__":
    start = datetime.now()
    asyncio.run(run_crawl())
    log.info(f"\n  Completed in {datetime.now() - start}")
