#!/usr/bin/env python3
"""
Deep Enrichment — Fill remaining gaps via Gemini
=================================================
Targets the biggest data gaps in FINAL_SELFPUB_SCORED.csv:
  Phase A: Book-level details (Last/Highest/Lowest book names, ratings)
  Phase B: Missing First Book Ratings + Goodreads Series URLs
  Phase C: Remaining missing emails/social links

Usage:
  python3 deep_enrich.py --phase all
  python3 deep_enrich.py --phase A
"""

import os, sys, json, re, time, logging, argparse
from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DATA_DIR = BASE_DIR / "output"
INPUT_FILE = DATA_DIR / "PRIORITY_SELFPUB_ENRICHED.csv"
PARTIAL_FILE = DATA_DIR / "deep_enrich_partial.csv"

# Column name mapping: ENRICHED file uses different names than SCORED file
COL_MAP = {
    "Email": "Author Email",
    "Website": "Author Website",
}

parser = argparse.ArgumentParser()
parser.add_argument("--phase", type=str, default="all", choices=["A", "B", "C", "all"])
args = parser.parse_args()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(DATA_DIR / "deep_enrich.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("deep_enrich")


def get_gemini_key():
    env_path = PROJECT_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith("GEMINI_API_KEY="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("GEMINI_API_KEY", "")


GEMINI_KEY = get_gemini_key()
BATCH_SIZE_A = 50
BATCH_SIZE_B = 50
BATCH_SIZE_C = 50
SAVE_INTERVAL = 5


def _robust_json_parse(text):
    text = text.strip()
    if text.startswith("```json"): text = text[7:]
    elif text.startswith("```"): text = text[3:]
    if text.endswith("```"): text = text[:-3]
    text = text.strip()
    text = re.sub(r',\s*([}\]])', r'\1', text)
    try:
        data = json.loads(text)
        return data if isinstance(data, list) else [data]
    except json.JSONDecodeError:
        pass
    match = re.search(r'\[[\s\S]*\]', text)
    if match:
        try:
            return json.loads(re.sub(r',\s*([}\]])', r'\1', match.group()))
        except json.JSONDecodeError:
            pass
    objects = []
    for m in re.finditer(r'\{[^{}]+\}', text):
        try:
            objects.append(json.loads(m.group()))
        except json.JSONDecodeError:
            continue
    return objects


def is_missing(val):
    if val is None: return True
    if isinstance(val, float) and np.isnan(val): return True
    return str(val).strip().lower() in ("", "nan", "none", "null", "nat")


def call_gemini(model, prompt, label=""):
    backoff = 5
    for attempt in range(1, 6):
        try:
            response = model.generate_content(prompt, request_options={"timeout": 180})
            return _robust_json_parse(response.text.strip())
        except Exception as e:
            err = str(e).lower()
            if any(k in err for k in ["429", "rate", "quota", "resource_exhausted"]) and attempt < 5:
                log.warning(f"    Rate limit on {label} (attempt {attempt}). Backing off {backoff}s...")
                time.sleep(backoff)
                backoff = min(backoff * 2, 60)
            else:
                log.warning(f"    Gemini error on {label}: {e}")
                return []
    return []


# ══════════════════════════════════════════════════════════════
# PHASE A: Book-level details (Last/Highest/Lowest)
# ══════════════════════════════════════════════════════════════
def phase_A(df, model):
    log.info("=" * 70)
    log.info("  PHASE A: Book-level details (Last/Highest/Lowest book)")
    log.info("=" * 70)

    def needs_A(row):
        return (is_missing(row.get("Last Book Name"))
                or is_missing(row.get("Highest Rated Book Name"))
                or is_missing(row.get("Lowest Rated Book Name")))

    mask = df.apply(needs_A, axis=1)
    indices = df[mask].index.tolist()
    log.info(f"  Need book-level details: {len(indices)}")

    if not indices:
        return df

    total_batches = 0
    total_updated = 0

    for batch_start in range(0, len(indices), BATCH_SIZE_A):
        batch_end = min(batch_start + BATCH_SIZE_A, len(indices))
        batch_idx = indices[batch_start:batch_end]
        batch_num = batch_start // BATCH_SIZE_A + 1
        total_needed = (len(indices) + BATCH_SIZE_A - 1) // BATCH_SIZE_A

        entries = []
        for i, idx in enumerate(batch_idx):
            row = df.loc[idx]
            series = str(row.get("Book Series Name", ""))
            author = str(row.get("Author Name", ""))
            books_n = row.get("Books in Series", "")
            first_book = str(row.get("First Book Name", ""))
            entries.append(
                f'{i+1}. "{series}" by {author} ({books_n} books, First book: "{first_book}")'
            )

        prompt = f"""You are a book database expert. For each series below, provide book-level details.

Series:
{chr(10).join(entries)}

For EACH series return:
- "series": series name as given
- "last_book_name": name of the most recently published book
- "last_book_rating": Goodreads rating of last book (float, or null)
- "last_book_rating_count": number of ratings for last book (integer, or null)
- "highest_rated_book": name of highest rated book in the series
- "highest_rated_rating": its Goodreads rating (float)
- "highest_rated_count": its number of ratings (integer, or null)
- "lowest_rated_book": name of lowest rated book in the series
- "lowest_rated_rating": its Goodreads rating (float)
- "lowest_rated_count": its number of ratings (integer, or null)
- "goodreads_url": Goodreads series URL if known (e.g., "https://www.goodreads.com/series/123456-series-name"), null if unknown

Return ONLY a JSON array."""

        results = call_gemini(model, prompt, f"PhaseA batch {batch_num}")

        batch_updated = 0
        for i, idx in enumerate(batch_idx):
            result = results[i] if i < len(results) else None
            if not result:
                series = str(df.at[idx, "Book Series Name"]).lower().strip()
                for r in results:
                    if str(r.get("series", "")).lower().strip() == series:
                        result = r
                        break
            if not result:
                continue

            updated = False
            # Detect column name variant for "Last Book Rating" (may have \n)
            lb_rating_col = "Last Book Rating"
            for c in df.columns:
                if "Last Book" in c and "Rating" in c and "Count" not in c and "Name" not in c:
                    lb_rating_col = c
                    break
            # Ensure Goodreads Series URL column exists
            if "Goodreads Series URL" not in df.columns:
                df["Goodreads Series URL"] = ""
            for field, col in [
                ("last_book_name", "Last Book Name"),
                ("last_book_rating", lb_rating_col),
                ("last_book_rating_count", "Last Book Rating Count"),
                ("highest_rated_book", "Highest Rated Book Name"),
                ("highest_rated_rating", "Highest Rated Book Rating"),
                ("highest_rated_count", "Highest Rated Book Rating Count"),
                ("lowest_rated_book", "Lowest Rated Book Name"),
                ("lowest_rated_rating", "Lowest Rated Book Rating"),
                ("lowest_rated_count", "Lowest Rated Book Rating Count"),
                ("goodreads_url", "Goodreads Series URL"),
            ]:
                val = result.get(field)
                if val and not is_missing(val) and is_missing(df.at[idx, col]):
                    df.at[idx, col] = val
                    updated = True

            if updated:
                batch_updated += 1

        total_updated += batch_updated
        total_batches += 1
        log.info(f"    Phase A Batch {batch_num}/{total_needed}: {batch_updated}/{len(batch_idx)} updated (total: {total_updated})")

        if total_batches % SAVE_INTERVAL == 0:
            df.to_csv(PARTIAL_FILE, index=False)
            log.info(f"    [Partial save]")

        time.sleep(0.5)

    log.info(f"  Phase A COMPLETE: {total_updated} entries updated")
    return df


# ══════════════════════════════════════════════════════════════
# PHASE B: Missing First Book Ratings + GR URLs
# ══════════════════════════════════════════════════════════════
def phase_B(df, model):
    log.info("=" * 70)
    log.info("  PHASE B: Missing First Book Ratings & Goodreads URLs")
    log.info("=" * 70)

    def needs_B(row):
        rating = row.get("First Book Rating")
        try:
            r = float(rating)
            if r <= 0 or np.isnan(r):
                return True
        except (ValueError, TypeError):
            return True
        return False

    mask = df.apply(needs_B, axis=1)
    indices = df[mask].index.tolist()
    log.info(f"  Missing first book rating: {len(indices)}")

    if not indices:
        return df

    total_batches = 0
    total_fixed = 0

    for batch_start in range(0, len(indices), BATCH_SIZE_B):
        batch_end = min(batch_start + BATCH_SIZE_B, len(indices))
        batch_idx = indices[batch_start:batch_end]
        batch_num = batch_start // BATCH_SIZE_B + 1
        total_needed = (len(indices) + BATCH_SIZE_B - 1) // BATCH_SIZE_B

        entries = []
        for i, idx in enumerate(batch_idx):
            row = df.loc[idx]
            entries.append(
                f'{i+1}. "{row.get("First Book Name", row.get("Book Series Name", ""))}" by {row.get("Author Name", "")}'
            )

        prompt = f"""You are a book rating expert. For each book below, provide its Goodreads rating.

Books:
{chr(10).join(entries)}

For EACH book return:
- "title": title as given
- "rating": Goodreads average rating (float like 4.12, or null if unknown)
- "rating_count": approximate number of Goodreads ratings (integer, or null)

Return ONLY a JSON array."""

        results = call_gemini(model, prompt, f"PhaseB batch {batch_num}")

        batch_fixed = 0
        for i, idx in enumerate(batch_idx):
            result = results[i] if i < len(results) else None
            if not result:
                title = str(df.at[idx, "First Book Name"]).lower().strip()
                for r in results:
                    if str(r.get("title", "")).lower().strip() == title:
                        result = r
                        break

            if result:
                rating = result.get("rating")
                count = result.get("rating_count")
                if rating and not is_missing(rating):
                    try:
                        r = float(rating)
                        if r > 0:
                            df.at[idx, "First Book Rating"] = r
                            batch_fixed += 1
                    except (ValueError, TypeError):
                        pass
                if count and not is_missing(count):
                    try:
                        c = int(float(count))
                        if c > 0 and is_missing(df.at[idx, "First Book Rating Count"]):
                            df.at[idx, "First Book Rating Count"] = c
                    except (ValueError, TypeError):
                        pass

        total_fixed += batch_fixed
        total_batches += 1
        log.info(f"    Phase B Batch {batch_num}/{total_needed}: {batch_fixed} ratings fixed (total: {total_fixed})")

        if total_batches % SAVE_INTERVAL == 0:
            df.to_csv(PARTIAL_FILE, index=False)

        time.sleep(0.5)

    log.info(f"  Phase B COMPLETE: {total_fixed} ratings filled")
    return df


# ══════════════════════════════════════════════════════════════
# PHASE C: Remaining missing contacts
# ══════════════════════════════════════════════════════════════
def phase_C(df, model):
    log.info("=" * 70)
    log.info("  PHASE C: Fill remaining missing author contacts")
    log.info("=" * 70)

    # Use correct column name for this file
    email_col = "Author Email" if "Author Email" in df.columns else "Email"
    website_col = "Author Website" if "Author Website" in df.columns else "Website"

    def needs_C(row):
        return is_missing(row.get(email_col))

    mask = df.apply(needs_C, axis=1)
    indices = df[mask].index.tolist()
    log.info(f"  Missing email: {len(indices)}")

    if not indices:
        return df

    # Deduplicate by author
    author_map = {}
    for idx in indices:
        author = str(df.at[idx, "Author Name"]).strip()
        if author and author.lower() not in ("nan", "none", ""):
            author_map.setdefault(author, []).append(idx)

    unique_authors = list(author_map.keys())
    log.info(f"  Unique authors needing contacts: {len(unique_authors)}")

    total_batches = 0
    total_found = 0

    for batch_start in range(0, len(unique_authors), BATCH_SIZE_C):
        batch_end = min(batch_start + BATCH_SIZE_C, len(unique_authors))
        batch_authors = unique_authors[batch_start:batch_end]
        batch_num = batch_start // BATCH_SIZE_C + 1
        total_needed = (len(unique_authors) + BATCH_SIZE_C - 1) // BATCH_SIZE_C

        entries = [f"{i+1}. {a}" for i, a in enumerate(batch_authors)]

        prompt = f"""You are a publishing researcher. For each author, find their public contact info.

Authors:
{chr(10).join(entries)}

For EACH return:
- "author": name
- "email": business/public email or null
- "website": author website URL or null
- "twitter": Twitter/X handle or null
- "instagram": Instagram handle or null
- "tiktok": TikTok handle or null

Return ONLY a JSON array."""

        results = call_gemini(model, prompt, f"PhaseC batch {batch_num}")

        batch_found = 0
        for i, author in enumerate(batch_authors):
            result = results[i] if i < len(results) else None
            if not result:
                for r in results:
                    if str(r.get("author", "")).lower().strip() == author.lower().strip():
                        result = r
                        break

            if not result:
                continue

            email = result.get("email")
            website = result.get("website")
            twitter = result.get("twitter")
            instagram = result.get("instagram")
            tiktok = result.get("tiktok")

            has_info = any(not is_missing(v) for v in [email, website, twitter, instagram, tiktok])
            if has_info:
                batch_found += 1

            for idx in author_map.get(author, []):
                if email and not is_missing(email) and is_missing(df.at[idx, email_col]):
                    df.at[idx, email_col] = str(email)
                if website and not is_missing(website) and is_missing(df.at[idx, website_col]):
                    df.at[idx, website_col] = str(website)
                if twitter and not is_missing(twitter) and is_missing(df.at[idx, "Twitter"]):
                    df.at[idx, "Twitter"] = str(twitter)
                if instagram and not is_missing(instagram) and is_missing(df.at[idx, "Instagram"]):
                    df.at[idx, "Instagram"] = str(instagram)
                if tiktok and not is_missing(tiktok) and is_missing(df.at[idx, "TikTok"]):
                    df.at[idx, "TikTok"] = str(tiktok)

        total_found += batch_found
        total_batches += 1
        log.info(f"    Phase C Batch {batch_num}/{total_needed}: {batch_found} found (total: {total_found})")

        if total_batches % SAVE_INTERVAL == 0:
            df.to_csv(PARTIAL_FILE, index=False)

        time.sleep(0.5)

    log.info(f"  Phase C COMPLETE: {total_found} authors enriched")
    return df


# ══════════════════════════════════════════════════════════════
def run():
    start = datetime.now()
    log.info("=" * 70)
    log.info("  DEEP ENRICHMENT — Filling remaining gaps")
    log.info(f"  Phase(s): {args.phase}")
    log.info("=" * 70)

    if not GEMINI_KEY:
        log.error("No GEMINI_API_KEY!")
        sys.exit(1)

    df = pd.read_csv(INPUT_FILE, low_memory=False)
    log.info(f"  Loaded {len(df)} rows from {INPUT_FILE.name}")

    import google.generativeai as genai
    genai.configure(api_key=GEMINI_KEY)
    model = genai.GenerativeModel("gemini-2.5-flash")

    if args.phase in ("A", "all"):
        df = phase_A(df, model)
        df.to_csv(INPUT_FILE, index=False)
        log.info(f"  Saved after Phase A")

    if args.phase in ("B", "all"):
        df = phase_B(df, model)
        df.to_csv(INPUT_FILE, index=False)
        log.info(f"  Saved after Phase B")

    if args.phase in ("C", "all"):
        df = phase_C(df, model)
        df.to_csv(INPUT_FILE, index=False)
        log.info(f"  Saved after Phase C")

    if PARTIAL_FILE.exists():
        PARTIAL_FILE.unlink()

    # Update Excel too
    update_excel(df)

    log.info(f"\n  DONE in {datetime.now() - start}")


def update_excel(df):
    """Update the main Excel file with the enriched data."""
    xlsx_path = DATA_DIR / "FINAL_SELFPUB_SCORED.xlsx"
    if not xlsx_path.exists():
        return

    log.info("  Updating Excel workbook...")
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    # Remove and recreate Combined sheet
    if "Combined" in wb.sheetnames:
        del wb["Combined"]
    wb.save(xlsx_path)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as writer:
        df.to_excel(writer, sheet_name="Combined", index=False)

    # Move Combined to first position
    wb = load_workbook(xlsx_path)
    wb.move_sheet("Combined", offset=-len(wb.sheetnames)+1)
    wb.save(xlsx_path)
    log.info("  Excel updated")


if __name__ == "__main__":
    run()
